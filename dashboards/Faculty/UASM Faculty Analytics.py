# ===========================================================================
#  UASM · Faculty Analytics  ·  App multipágina
#  Página 1: Full-time Faculty Composition
#  Página 2: Full-time Faculty Staffing Levels
#  Página 3: Distribution by Academic Area
#  Página 4: Faculty Demographics
#  Página 5: Full-time Faculty Activities
#  Página 6: Faculty Qualifications
#  Página 7: Update Data — sube las templates y escribe directo en los
#            archivos .xlsx de Drive (BD_profesores.xlsx, BD_cartelera.xlsx)
#            vía Service Account + Drive API.
# ===========================================================================
from __future__ import annotations

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import re
import io
import base64
import time
import math
import datetime
import requests
from typing import Optional, Tuple, List, Dict

try:
    from google.oauth2.service_account import Credentials
    _GSPREAD_OK = True
    _GSPREAD_IMPORT_ERR = None
except ImportError as _e:
    _GSPREAD_OK = False
    _GSPREAD_IMPORT_ERR = str(_e)

try:
    import openpyxl
    from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
    from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
    from openpyxl.formula.translate import Translator
    from openpyxl.worksheet.formula import ArrayFormula
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

# 1) CONFIGURACIÓN GLOBAL (una sola vez para toda la app)
st.set_page_config(
    page_title="UASM Faculty Analytics",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded",
)
# CSS compartido por todas las páginas
st.markdown(
    "<style>"
    ".suite-header{display:flex;flex-direction:column;align-items:center;"
    "padding:16px 24px 12px;"
    "background:linear-gradient(135deg,#004d47 0%,#21877D 60%,#2EC4B6 100%);"
    "border-radius:12px;box-shadow:0 2px 8px rgba(0,77,71,.18);margin-bottom:14px;}"
    ".sh-super{font-size:11px;font-weight:700;letter-spacing:2px;"
    "color:#56D6C9;text-transform:uppercase;margin-bottom:2px;}"
    ".sh-title{font-size:26px;font-weight:800;color:#fff;text-align:center;line-height:1.2;}"
    ".sh-sub{font-size:13px;color:rgba(255,255,255,.75);margin-top:4px;text-align:center;}"
    ".kv{font-size:28px;font-weight:800;color:#21877D;line-height:1.1;}"
    ".kl{font-size:11px;font-weight:600;color:#6B7280;"
    "text-transform:uppercase;letter-spacing:.5px;margin-top:3px;}"
    ".period-label{text-align:center;font-weight:700;font-size:1.05rem;color:#21877D;}"
    "a.dl-min,a.dl-min:link,a.dl-min:visited{color:#00A896 !important;"
    "text-decoration:underline !important;font-size:13px;"
    "display:inline-block;margin-top:6px;}"
    "a.dl-min:hover{opacity:.85;}"
    "div.stDownloadButton>button{background:transparent !important;"
    "border:none !important;box-shadow:none !important;"
    "color:#21877D !important;font-size:13px !important;"
    "padding:0 !important;text-decoration:underline !important;}"
    "div.stDownloadButton{margin:2px 0 8px 0;}"
    "thead th{background:#dff7f2 !important;color:#004d47 !important;"
    "font-weight:700 !important;}"
    "section[data-testid='stSidebar']{background:#F0F7F6 !important;}"
    "#mode-pill [role='radiogroup']{display:flex;gap:8px;margin-top:0;}"
    "#mode-pill [role='radio']{flex:1;justify-content:center;"
    "border:1px solid #d0d4d9;border-radius:999px;padding:8px 12px;"
    "background:#f0f2f6;color:#666;font-weight:600;cursor:pointer;text-align:center;}"
    "#mode-pill [role='radio'][aria-checked='true']{"
    "background:#dff7f2;color:#004d47;border-color:#8fd7cc;}"
    "#mode-pill [data-baseweb='radio'] input{display:none !important;}"
    "div[data-testid='stButton'] button{background:#FFFFFF !important;"
    "border:1px solid #D1E8E4 !important;border-radius:10px !important;"
    "color:#374151 !important;font-size:14px !important;"
    "font-weight:600 !important;height:48px !important;"
    "box-shadow:0 1px 3px rgba(0,0,0,.04) !important;}"
    "div[data-testid='stButton'] button:hover{"
    "background:#F8FFFE !important;border-color:#B7DCD6 !important;}"
    ".st-key-nav_toggle{position:fixed;top:0.25rem;left:50%;transform:translateX(-50%);"
    "z-index:999999;width:70vw;max-width:900px;}"
    ".st-key-nav_toggle div[data-testid='stHorizontalBlock']{"
    "display:flex !important;flex-wrap:nowrap !important;width:100% !important;"
    "justify-content:space-between !important;gap:8px !important;}"
    ".st-key-nav_toggle div[data-testid='column']{width:auto !important;min-width:fit-content !important;flex:none !important;}"
    ".st-key-nav_toggle div[data-testid='stPageLink']{width:auto !important;min-width:fit-content !important;overflow:visible !important;}"
    ".st-key-nav_toggle div[data-testid='stPageLink'] a{white-space:nowrap !important;overflow:visible !important;text-overflow:unset !important;width:auto !important;min-width:fit-content !important;}"
    ".st-key-nav_toggle div[data-testid='stPageLink'] a p{white-space:nowrap !important;overflow:visible !important;}"
    ".st-key-update_sidebar_group{text-align:center;}"
    "div[class*='st-key-course_box_bad_'] div[data-testid='stExpander'],"
    "div[class*='st-key-prof_box_bad_'] div[data-testid='stExpander']{"
    "border:1.5px solid #DC2626 !important;border-radius:8px !important;}"
    ".st-key-update_sidebar_group img{margin:0 auto;}"
    ".st-key-go_to_dashboard_btn a{"
    "display:flex !important;align-items:center;justify-content:center;gap:10px;text-align:center;"
    "background:#004d47 !important;border:none !important;"
    "color:#FFFFFF !important;font-size:23px !important;font-weight:800 !important;"
    "border-radius:14px !important;padding:22px 10px !important;text-decoration:none !important;"
    "box-shadow:0 4px 14px rgba(0,77,71,.25);transition:transform .15s ease;}"
    ".st-key-go_to_dashboard_btn a span{color:#FFFFFF !important;}"
    ".st-key-go_to_dashboard_btn a:hover{transform:translateY(-2px);background:#00332E !important;}"
    ".st-key-go_to_update_btn a{"
    "display:flex !important;align-items:center;justify-content:center;gap:6px;"
    "background:#FFFFFF !important;border:1px solid #D1E8E4 !important;"
    "color:#374151 !important;font-size:14px !important;font-weight:600 !important;"
    "border-radius:10px !important;padding:12px 8px !important;text-decoration:none !important;"
    "height:48px;box-shadow:0 1px 3px rgba(0,0,0,.04);}"
    ".st-key-go_to_update_btn a:hover{background:#F8FFFE !important;border-color:#B7DCD6 !important;}"
    "</style>",
    unsafe_allow_html=True,
)


def _scroll_table_right_once(container_key: str):
    """Desplaza UNA sola vez, justo al cargar, la tabla dentro del
    st.container(key=container_key) hacia su borde derecho — para que el
    último periodo quede visible sin scroll manual. Solo debe usarse en
    tablas de continuidad temporal (columnas = periodos); reintenta unos
    segundos por si la tabla tarda en pintarse, y luego se detiene solo, para
    no pelear con el scroll manual del usuario."""
    components.html(
        f"""<script>
        (function() {{
            const doc = window.parent.document;
            let tries = 0;
            const t = setInterval(() => {{
                tries++;
                const box = doc.querySelector('.st-key-{container_key} [data-testid="stDataFrame"]');
                if (box) {{
                    const scroller = box.querySelector('.dvn-scroller, [class*="scroll"]') || box;
                    scroller.scrollLeft = scroller.scrollWidth;
                    if (scroller.scrollWidth > scroller.clientWidth) clearInterval(t);
                }}
                if (tries > 10) clearInterval(t);
            }}, 200);
        }})();
        </script>""",
        height=0,
    )

# 2) HELPERS COMPARTIDOS
def _xlsx_bytes(df, sheet_name="Data"):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf) as w:
        df.to_excel(w, index=False, sheet_name=sheet_name[:31])
    buf.seek(0)
    return buf.getvalue()


def _download_link(label, df, filename):
    b64 = base64.b64encode(_xlsx_bytes(df)).decode()
    href = ("data:application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet;base64," + b64)
    st.markdown(
        f'<a class="dl-min" download="{filename}" href="{href}">{label}</a>',
        unsafe_allow_html=True,
    )


def _render_header(title, subtitle=""):
    sub = f'<div class="sh-sub">{subtitle}</div>' if subtitle else ""
    st.markdown(
        f'<div class="suite-header"><div class="sh-super">UASM · Faculty Analytics</div>'
        f'<div class="sh-title">{title}</div>{sub}</div>',
        unsafe_allow_html=True,
    )


def _highlight_band(fig, label, all_labels, color="#D0E5F5"):
    if label in all_labels:
        pos = all_labels.index(label)
        fig.add_shape(type="rect", xref="x", yref="paper",
                      x0=pos - 0.4, x1=pos + 0.4, y0=0, y1=1,
                      fillcolor=color, opacity=0.35, line_width=0)


def _is_inter_label(p) -> bool:
    """True si el período es 'YYYY Intersemestral' (usado en Area y Demographics)."""
    return bool(re.fullmatch(r"\d{4}\s+Intersemestral", str(p)))


# 3) CARGA DE DATOS (compartida por todas las páginas)
# BD_Faculty (el Google Sheet combinado) fue retirado. Ahora la fuente de
# verdad son 3 archivos .xlsx sueltos en Drive — cada uno con varias hojas:
PROFESORES_FILE_ID = "1ncnUk_8VsDt1I0Hui9g0VyoTkA-8P376"      # BD_profesores.xlsx  → hojas: planta, Info. Profesores, Faculty Distribution
CARTELERA_FILE_ID = "14Hongi8a180XTvuZGUf3soixgQpFp0Wl"       # BD_cartelera.xlsx   → hojas: cartelera, programas, cursos, qualifications
QUESTIONNAIRE_FILE_ID = "1u6YTILxGOEq7eq1RE_l5sPg-vM5Wu5jH"   # BD_faculty_questionnaire.xlsx → hoja: Faculty_questionnaire
TEMPLATE_PROFESORES_NUEVOS_FILE_ID = "1EEFfstkupiSD-2YyBPauO2WvzelZnYDl"  # Template_profesores_nuevos.xlsx (carpeta de templates)
TEMPLATE_CURSOS_NUEVOS_FILE_ID = "1UGpwCGf3w3GByDm8Mj_1hJgNw0oYtqBb"  # Template_cursos_nuevos.xlsx (carpeta de templates)
TEMPLATE_PLANTA_FILE_ID = "1bE5cNo1UUUjN34-L799U7KJQarefQl1h"  # Template_planta.xlsx (carpeta de templates)
TEMPLATE_CARTELERA_FILE_ID = "1O5DV0ABy_-G1_t4pjfeZHUi-f9OFJprl"  # Template_cartelera.xlsx (carpeta de templates)

# ── Autenticación (lectura y escritura vía Service Account) ────────────────
# Requiere una service account de Google Cloud, compartida como Editor en los
# 3 archivos de arriba, con su JSON guardado en Streamlit secrets bajo la
# clave "gcp_service_account". Ver docs al final de page_update_data().
_GSPREAD_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


def _get_gspread_access_token() -> Optional[str]:
    """Token de acceso de la service account, para llamar la API de Drive
    ya autenticados (necesario porque estos archivos ya no son públicos)."""
    if not _GSPREAD_OK or "gcp_service_account" not in st.secrets:
        return None
    try:
        from google.auth.transport.requests import Request as _GoogleAuthRequest
        creds = Credentials.from_service_account_info(
            dict(st.secrets["gcp_service_account"]), scopes=_GSPREAD_SCOPES
        )
        creds.refresh(_GoogleAuthRequest())
        return creds.token
    except Exception:
        return None


@st.cache_data(ttl=300)
def _download_drive_file_bytes(file_id: str) -> bytes:
    """Descarga un archivo .xlsx de Drive (autenticado con la service account)
    y devuelve sus bytes crudos. Cacheado 5 min por archivo para evitar
    descargas repetidas entre páginas."""
    token = _get_gspread_access_token()
    if not token:
        if not _GSPREAD_OK:
            st.error(
                "📦 Falta instalar las librerías `gspread` y `google-auth` en el "
                f"entorno (el import falló con: `{_GSPREAD_IMPORT_ERR}`). "
                "Agrégalas a tu `requirements.txt`:\n\n```\ngspread\ngoogle-auth\n```"
            )
        elif "gcp_service_account" not in st.secrets:
            st.error(
                "🔑 No encuentro `st.secrets['gcp_service_account']`. Revisa en "
                "Streamlit Cloud → tu app → Settings → Secrets que el bloque "
                "empiece exactamente con `[gcp_service_account]` y que lo hayas "
                "guardado (la app se reinicia sola al guardar)."
            )
        else:
            st.error(
                "🔑 Las credenciales de `gcp_service_account` están presentes pero "
                "no se pudieron usar para autenticar — revisa que el JSON esté "
                "completo y bien formado en los Secrets (especialmente el campo "
                "`private_key`, que debe conservar los `\\n` tal cual)."
            )
        st.stop()

    url = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"
    headers = {"Authorization": f"Bearer {token}"}

    resp = None
    last_err = None
    for attempt in range(3):
        try:
            resp = requests.get(url, timeout=60, headers=headers)
            if resp.status_code == 200:
                return resp.content
            last_err = RuntimeError(f"HTTP {resp.status_code}: {resp.text[:200]}")
        except Exception as e:
            last_err = e
        time.sleep(2)

    st.error(
        f"🌐 No se pudo descargar el archivo de Drive ({file_id}) tras varios intentos: "
        f"{last_err}\n\nVerifica que el archivo esté compartido con el correo de la "
        "service account (`client_email` de tu JSON), con permiso de Editor."
    )
    st.stop()


@st.cache_data(ttl=300)
def load_data():
    raw = io.BytesIO(_download_drive_file_bytes(PROFESORES_FILE_ID))
    df_ = pd.read_excel(raw, sheet_name="planta")

    def _norm_per(val):
        if pd.isna(val):
            return None
        try:
            s = str(int(float(str(val).strip())))
        except (ValueError, OverflowError):
            s = str(val).strip()
        if re.search(r'inter', s, re.IGNORECASE):
            y = re.search(r'((?:19|20)\d{2})', s)
            return f"{y.group(1)} Intersemestral" if y else None
        m = re.fullmatch(r'((?:19|20)\d{2})(10|20)', s) or \
            re.search(r'((?:19|20)\d{2})[^0-9]+(10|20)', s)
        return f"{m.group(1)}-{m.group(2)}" if m else None

    src = "Periodo" if "Periodo" in df_.columns else df_.columns[0]
    df_["Periodo"] = df_[src].map(_norm_per)
    df_ = df_[df_["Periodo"].astype(str).str.match(
        r'^(?:19|20)\d{2}-(10|20)$|^(?:19|20)\d{2}\s+Intersemestral$'
    )].copy()

    if "ID Nr." in df_.columns and "ID" not in df_.columns:
        df_ = df_.rename(columns={"ID Nr.": "ID"})
    if "ID" not in df_.columns and "ID Nr." in df_.columns:
        df_["ID"] = df_["ID Nr."]

    def _key(p):
        s = str(p)
        return (int(s[:4]), 30 if "Intersemestral" in s else int(s[-2:]))

    return df_.sort_values("Periodo", key=lambda c: c.map(_key))


df = load_data()


# Loaders específicos: página "Distribution by Area"
# (mismo archivo BD_profesores.xlsx, pero conservan exactamente la lógica original de esa página)
@st.cache_data(ttl=0)
def area_load_fulltime() -> pd.DataFrame:
    raw = io.BytesIO(_download_drive_file_bytes(PROFESORES_FILE_ID))
    df_ = pd.read_excel(raw, sheet_name="planta")

    sem = df_["Semestre"].astype(str).str.strip() if "Semestre" in df_.columns else df_.iloc[:, 0].astype(str).str.strip()
    is_inter = sem.str.contains("inter", case=False, na=False)
    df_["Periodo"] = np.where(is_inter, sem.str[:4] + " Intersemestral", sem.str[:4] + "-" + sem.str[-2:])

    if "Academic Area" in df_.columns and "AREA_PROFESOR" not in df_.columns:
        df_["AREA_PROFESOR"] = df_["Academic Area"]
    if "ID Nr." in df_.columns and "ID" not in df_.columns:
        df_ = df_.rename(columns={"ID Nr.": "ID"})
    if "Full Name" not in df_.columns:
        fn = df_.get("First Name", "").astype(str).fillna("")
        ln = df_.get("Last Name", "").astype(str).fillna("")
        df_["Full Name"] = (fn + " " + ln).str.strip()
    return df_


@st.cache_data(ttl=0)
def area_load_parttime() -> pd.DataFrame:
    raw = io.BytesIO(_download_drive_file_bytes(PROFESORES_FILE_ID))
    df_ = pd.read_excel(raw, sheet_name="Faculty Distribution")

    if "PLANTA_CATEDRA" in df_.columns:
        col = df_["PLANTA_CATEDRA"].astype(str).str.strip()
        col = col.str.normalize("NFKD").str.encode("ascii", errors="ignore").str.decode("ascii")
        df_ = df_[col.str.upper().eq("CATEDRA")].copy()

    sem = df_["Semestre"].astype(str).str.strip()
    is_inter = sem.str.contains("inter", case=False, na=False)
    df_.loc[~is_inter, "Periodo"] = sem.str[:4] + "-" + sem.str[-2:]
    df_.loc[is_inter, "Periodo"] = sem.str[:4] + " Intersemestral"

    if "ID Nr." in df_.columns and "ID" not in df_.columns:
        df_ = df_.rename(columns={"ID Nr.": "ID"})
    if "AREA_PROFESOR" not in df_.columns and "Academic Area" in df_.columns:
        df_["AREA_PROFESOR"] = df_["Academic Area"]
    return df_


# Loaders específicos: página "Demographics"
# Nota: esta página usa un formato de Periodo sin guion ("YYYY10"/"YYYY Intersemestral"),
# distinto al de las demás páginas — se conserva igual que en el script original.
@st.cache_data(ttl=0)
def demo_load_fulltime() -> pd.DataFrame:
    raw = io.BytesIO(_download_drive_file_bytes(PROFESORES_FILE_ID))
    df_ = pd.read_excel(raw, sheet_name="planta")

    if "Semestre" in df_.columns:
        sem = df_["Semestre"].astype(str).str.strip()
    else:
        sem = df_.iloc[:, 0].astype(str).str.strip()
    is_inter = sem.str.contains("inter", case=False, na=False)
    df_["Periodo"] = np.where(is_inter, sem.str[:4] + " Intersemestral", sem.str[:4] + sem.str[-2:])

    if "Academic Area" in df_.columns and "AREA_PROFESOR" not in df_.columns:
        df_["AREA_PROFESOR"] = df_["Academic Area"]
    if "ID Nr." in df_.columns and "ID" not in df_.columns:
        df_ = df_.rename(columns={"ID Nr.": "ID"})
    if "Full Name" not in df_.columns:
        fn = df_.get("First Name", "").astype(str).fillna("")
        ln = df_.get("Last Name", "").astype(str).fillna("")
        df_["Full Name"] = (fn + " " + ln).str.strip()
    return df_


def _norm_id(v):
    """Normaliza un ID para hacer merges seguros: números (12345678, 12345678.0,
    '12345678') se normalizan todos al mismo string '12345678'; IDs no numéricos
    (cédulas/pasaportes internacionales como 'XDC641686') se dejan como texto
    limpio. Evita el bug de pandas donde NaN==NaN en un merge multiplica filas
    cuando varios IDs no numéricos se convertían todos a NaN con pd.to_numeric."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        f = float(v)
        return str(int(f)) if f.is_integer() else str(f)
    except (ValueError, TypeError):
        s = str(v).strip().upper()
        return s if s else None


@st.cache_data(ttl=0)
def demo_load_parttime() -> pd.DataFrame:
    raw = io.BytesIO(_download_drive_file_bytes(PROFESORES_FILE_ID))
    df_ = pd.read_excel(raw, sheet_name="Faculty Distribution")

    if "PLANTA_CATEDRA" in df_.columns:
        col = df_["PLANTA_CATEDRA"].astype(str).str.strip()
        col = col.str.normalize("NFKD").str.encode("ascii", errors="ignore").str.decode("ascii")
        df_ = df_[col.str.upper().eq("CATEDRA")].copy()

    sem = df_["Semestre"].astype(str).str.strip()
    is_inter = sem.str.contains("inter", case=False, na=False)
    df_.loc[~is_inter, "Periodo"] = sem.str[:4] + sem.str[-2:]
    df_.loc[is_inter, "Periodo"] = sem.str[:4] + " Intersemestral"

    # 'Faculty Distribution' solo trae 8 columnas (Semestre, Profesor, ID,
    # AREA_PROFESOR, GÉNERO, TIPO, P/S, PLANTA_CATEDRA) — insuficiente para el
    # análisis de demografía completo (título, nacionalidad, fecha de
    # nacimiento, universidad, etc.). Se trae TODA la información adicional
    # desde 'Info. Profesores', unida por ID (normalizado como texto — NO con
    # pd.to_numeric, que convierte las cédulas/pasaportes no numéricos en NaN
    # y como pandas trata NaN==NaN en un merge, eso multiplicaba filas).
    raw2 = io.BytesIO(_download_drive_file_bytes(PROFESORES_FILE_ID))
    df_info = pd.read_excel(raw2, sheet_name="Info. Profesores")
    df_info.columns = df_info.columns.str.strip()
    if "ID" in df_.columns and "ID" in df_info.columns:
        extra_cols = [c for c in df_info.columns
                      if c not in ("Profesor", "ID", "AREA_PROFESOR", "GÉNERO", "TIPO", "P/S")]
        df_info_extra = df_info[["ID"] + extra_cols].copy()
        df_info_extra["_id_key"] = df_info_extra["ID"].map(_norm_id)
        df_info_extra = df_info_extra.dropna(subset=["_id_key"]).drop_duplicates(subset=["_id_key"])
        df_info_extra = df_info_extra.drop(columns=["ID"])

        df_["_id_key"] = df_["ID"].map(_norm_id)
        df_ = df_.merge(df_info_extra, on="_id_key", how="left").drop(columns=["_id_key"])

    if "ID Nr." not in df_.columns and "ID" in df_.columns:
        df_ = df_.rename(columns={"ID": "ID Nr."})
    if "AREA_PROFESOR" not in df_.columns and "Academic Area" in df_.columns:
        df_["AREA_PROFESOR"] = df_["Academic Area"]
    return df_


# Loaders específicos: página "Qualifications"
@st.cache_data(ttl=0)
def qual_load_planta() -> pd.DataFrame:
    try:
        raw = io.BytesIO(_download_drive_file_bytes(PROFESORES_FILE_ID))
        dfp = pd.read_excel(raw, sheet_name="planta")
        dfp.columns = dfp.columns.str.strip()
        return dfp
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=0)
def qual_load_faculty_distribution() -> pd.DataFrame:
    raw = io.BytesIO(_download_drive_file_bytes(PROFESORES_FILE_ID))
    df_ = pd.read_excel(raw, sheet_name="Faculty Distribution")
    df_.columns = df_.columns.str.strip()

    # 'Faculty Distribution' solo trae 8 columnas base — Qualifications necesita
    # 'Highest Degree' (y otros campos), que solo viven en 'Info. Profesores'.
    # Mismo merge por ID normalizado que en demo_load_parttime().
    raw2 = io.BytesIO(_download_drive_file_bytes(PROFESORES_FILE_ID))
    df_info = pd.read_excel(raw2, sheet_name="Info. Profesores")
    df_info.columns = df_info.columns.str.strip()
    if "ID" in df_.columns and "ID" in df_info.columns:
        extra_cols = [c for c in df_info.columns
                      if c not in ("Profesor", "ID", "AREA_PROFESOR", "GÉNERO", "TIPO", "P/S")]
        df_info_extra = df_info[["ID"] + extra_cols].copy()
        df_info_extra["_id_key"] = df_info_extra["ID"].map(_norm_id)
        df_info_extra = df_info_extra.dropna(subset=["_id_key"]).drop_duplicates(subset=["_id_key"])
        df_info_extra = df_info_extra.drop(columns=["ID"])
        df_["_id_key"] = df_["ID"].map(_norm_id)
        df_ = df_.merge(df_info_extra, on="_id_key", how="left").drop(columns=["_id_key"])
    return df_


@st.cache_data(ttl=0)
def qual_load_cartelera() -> pd.DataFrame:
    raw = io.BytesIO(_download_drive_file_bytes(CARTELERA_FILE_ID))
    df_ = pd.read_excel(raw, sheet_name="cartelera")
    df_.columns = df_.columns.str.strip()
    return df_



# 5) PÁGINA 1 — Full-time Faculty Composition
def page_composition():
    all_periods = df["Periodo"].astype(str).unique().tolist()
    sem_periods = sorted(
        [p for p in all_periods if re.fullmatch(r'(?:19|20)\d{2}-(10|20)', p)],
        key=_period_sort_key,
    )  # cronológico — alimenta tablas/gráficas de evolución
    inter_periods = sorted(
        [p for p in all_periods if re.fullmatch(r'(?:19|20)\d{2}\sIntersemestral', p)],
        key=_period_sort_key,
    )
    years = sorted(pd.Series(all_periods).str[:4].unique().tolist())

    # Sidebar específico de esta página
    with st.sidebar:
        st.markdown("#### Timeframe")
        tmode = st.radio("", ["Semestral", "Anual", "Intersemestral"], key="ft_comp_timeframe")

        sem_periods_desc = sorted(sem_periods, key=_period_sort_key, reverse=True)
        inter_periods_desc = sorted(inter_periods, key=_period_sort_key, reverse=True)
        years_desc = sorted(years, reverse=True)

        if tmode == "Semestral":
            vis = [p.replace("-", "") for p in sem_periods_desc]
            sel_vis = st.selectbox("Periodo", vis, index=0 if vis else 0,
                                    key="ft_comp_periodo_sem")
            sel_period_internal = sem_periods_desc[vis.index(sel_vis)] if vis else None
            sel_period_label = sel_vis
        elif tmode == "Anual":
            sel_period_internal = st.selectbox("Periodo", years_desc, index=0 if years_desc else 0,
                                                key="ft_comp_periodo_anual")
            sel_period_label = sel_period_internal
        else:
            sel_period_internal = st.selectbox("Periodo", inter_periods_desc,
                                                index=0 if inter_periods_desc else 0,
                                                key="ft_comp_periodo_inter")
            sel_period_label = sel_period_internal

    _render_header("Full-time Faculty Composition",
                   "Evolution and distribution of full-time faculty by ranking")

    # Ranking order & color map
    base_order = ["Full Professor", "Associate Professor", "Assistant Professor",
                  "Instructor", "Adjunct Faculty", "Distinguished Practitioner", "Emeritus Professor"]
    uniq_ranks = df["Faculty Ranking"].dropna().astype(str).unique().tolist()
    ranking_order = [x for x in base_order if x in uniq_ranks] + \
                     [x for x in uniq_ranks if x not in base_order]
    if "Faculty Ranking" in df.columns:
        df["Faculty Ranking"] = pd.Categorical(df["Faculty Ranking"],
                                                categories=ranking_order, ordered=True)

    palette = ["#037C70", "#27BDAE", "#4FFF98", "#FFD166",
               "#F4A261", "#E76F51", "#9D4EDD", "#6D597A",
               "#118AB2", "#073B4C", "#8AC926", "#FF70A6"]
    color_map_rk = {rk: palette[i % len(palette)] for i, rk in enumerate(ranking_order)}

    # Helpers de filtrado por periodo
    def periods_for_tables():
        if tmode == "Semestral":
            return sem_periods
        if tmode == "Intersemestral":
            return inter_periods
        return years

    def df_active():
        if sel_period_internal is None:
            return df.iloc[0:0].copy()
        if tmode in ("Semestral", "Intersemestral"):
            return df[df["Periodo"].astype(str).eq(sel_period_internal)].copy()
        dfa = df[df["Periodo"].astype(str).str.startswith(str(sel_period_internal))].copy()
        return dfa.sort_values("Periodo").drop_duplicates(subset=["ID"], keep="last")

    def pivot_counts():
        cols = periods_for_tables()
        if tmode in ("Semestral", "Intersemestral"):
            return pd.pivot_table(
                df[df["Periodo"].isin(cols)],
                index="Faculty Ranking", columns="Periodo",
                values="ID", aggfunc="count", fill_value=0
            ).reindex(ranking_order)
        out = {rk: {y: 0 for y in cols} for rk in ranking_order}
        for y in cols:
            dfa = df[df["Periodo"].astype(str).str.startswith(str(y))].copy()
            if dfa.empty:
                continue
            dfa = dfa.sort_values("Periodo").drop_duplicates(subset=["ID"], keep="last")
            for rk, v in dfa.groupby("Faculty Ranking")["ID"].count().items():
                out.setdefault(rk, {})[y] = int(v)
        return pd.DataFrame(out).T.reindex(ranking_order).reindex(columns=cols, fill_value=0)

    def line_source_all():
        cols = periods_for_tables()
        if tmode in ("Semestral", "Intersemestral"):
            dat = (df[df["Periodo"].isin(cols)]
                   .groupby(["Periodo", "Faculty Ranking"])["ID"]
                   .count().reset_index(name="Count"))
            dat["Periodo"] = pd.Categorical(dat["Periodo"], categories=cols, ordered=True)
            return dat, cols
        rows = []
        for y in cols:
            dfa = df[df["Periodo"].astype(str).str.startswith(str(y))].copy()
            dfa = dfa.sort_values("Periodo").drop_duplicates(subset=["ID"], keep="last")
            cts = (dfa.groupby("Faculty Ranking")["ID"].count()
                      .reindex(ranking_order, fill_value=0).reset_index()
                      .rename(columns={"ID": "Count"}))
            cts["Periodo"] = str(y)
            rows.append(cts)
        out = pd.concat(rows, ignore_index=True) if rows else \
              pd.DataFrame(columns=["Faculty Ranking", "Count", "Periodo"])
        out["Periodo"] = pd.Categorical(out["Periodo"], categories=cols, ordered=True)
        return out, cols

    def line_source_single(rank):
        cols = periods_for_tables()
        if tmode in ("Semestral", "Intersemestral"):
            dat = (df[df["Periodo"].isin(cols) & (df["Faculty Ranking"] == rank)]
                   .groupby("Periodo")["ID"].count()
                   .reindex(cols, fill_value=0).reset_index(name="Count"))
            return dat, cols
        vals = []
        for y in cols:
            dfa = df[df["Periodo"].astype(str).str.startswith(str(y))].copy()
            dfa = dfa.sort_values("Periodo").drop_duplicates(subset=["ID"], keep="last")
            vals.append({"Periodo": str(y),
                         "Count": int(dfa[dfa["Faculty Ranking"] == rank]["ID"].count())})
        return pd.DataFrame(vals), cols

    # Pivot table
    pivot = pivot_counts().reindex(ranking_order)
    pivot.loc["Total"] = pivot.sum(numeric_only=True)

    st.subheader("Number of Full-time Faculty by Ranking")

    def _bold_total(df_):
        s = pd.DataFrame("", index=df_.index, columns=df_.columns)
        if "Total" in df_.index:
            s.loc["Total", :] = "font-weight:700;"
        return s

    def _highlight_last(df_):
        s = pd.DataFrame("", index=df_.index, columns=df_.columns)
        if len(df_.columns) > 0 and "Total" in df_.index:
            candidates = [sel_period_internal, sel_period_label]
            target = next((c for c in candidates if c in df_.columns), df_.columns[-1])
            s.loc["Total", target] = "background-color:#dff7f2;color:#00A896;font-weight:700;"
        return s

    with st.container(key="tct_comp_ranking"):
        st.dataframe(
            pivot.style.apply(_bold_total, axis=None).apply(_highlight_last, axis=None).format(precision=0),
            use_container_width=True)
    _scroll_table_right_once("tct_comp_ranking")
    _download_link("Descargar tabla (Excel)",
                   pivot.reset_index().rename(columns={"index": "Faculty Ranking"}),
                   f"FT_Composition_{tmode}.xlsx")

    # Charts
    periods_sorted = periods_for_tables()
    st.session_state.setdefault("show_all", True)
    st.session_state.setdefault("single_ranking", "Select...")

    def on_select_ranking():
        if st.session_state.single_ranking != "Select...":
            st.session_state.show_all = False

    def on_toggle_show_all():
        if st.session_state.show_all:
            st.session_state.single_ranking = "Select..."

    st.header("Evolution & composition")
    col_left, col_right = st.columns(2)

    with col_right:
        st.subheader("Composition by period")
        st.markdown(f"<div style='text-align:center;font-weight:800;font-size:2rem;"
                    f"padding-top:4px;'>{sel_period_label}</div>", unsafe_allow_html=True)

        if tmode in ("Semestral", "Intersemestral"):
            dfbar = df[df["Periodo"].astype(str).eq(sel_period_internal)]
        else:
            dfa = df[df["Periodo"].astype(str).str.startswith(str(sel_period_internal))].copy()
            dfbar = dfa.sort_values("Periodo").drop_duplicates(subset=["ID"], keep="last")

        bar_counts = (dfbar.groupby("Faculty Ranking")["ID"].count()
                           .reindex(ranking_order).fillna(0).reset_index())
        bar_counts.columns = ["Faculty Ranking", "Count"]
        st.metric("Total Faculty:", int(bar_counts["Count"].sum()))

        fig_bar = px.bar(bar_counts, x="Count", y="Faculty Ranking", orientation="h",
                         text="Count", color="Faculty Ranking", color_discrete_map=color_map_rk,
                         category_orders={"Faculty Ranking": ranking_order[::-1]})
        fig_bar.update_xaxes(range=[0, max(1, int(bar_counts["Count"].max() or 0)) + 5], title=None)
        fig_bar.update_yaxes(title=None)
        fig_bar.update_traces(textposition="outside")
        st.plotly_chart(fig_bar, use_container_width=True)

    with col_left:
        st.subheader("Evolution of rankings")
        st.checkbox("Show all lines", key="show_all", on_change=on_toggle_show_all)
        st.selectbox("Select a ranking:", ["Select..."] + ranking_order,
                     key="single_ranking", on_change=on_select_ranking)

        fig_line = None
        xcats = periods_sorted

        if st.session_state.show_all:
            data_long, xcats = line_source_all()
            y_max = max(1, int(data_long["Count"].max()) if not data_long.empty else 0)
            fig_line = px.line(data_long, x="Periodo", y="Count", color="Faculty Ranking",
                               markers=True, title="Evolution — all rankings",
                               color_discrete_map=color_map_rk,
                               category_orders={"Periodo": xcats, "Faculty Ranking": ranking_order})
            fig_line.update_yaxes(range=[0, y_max + 1], title=None)
            fig_line.update_xaxes(type="category", categoryorder="array", categoryarray=xcats, title=None)
            fig_line.update_layout(height=550, showlegend=False)
        else:
            rk = st.session_state.single_ranking
            if rk != "Select...":
                data_single, xcats = line_source_single(rk)
                y_max = max(1, int(data_single["Count"].max()) if not data_single.empty else 0)
                fig_line = px.line(data_single, x="Periodo", y="Count", markers=True,
                                   title=f"Evolution — {rk}",
                                   color_discrete_sequence=[color_map_rk.get(rk, "#00A896")],
                                   category_orders={"Periodo": xcats})
                fig_line.update_yaxes(range=[0, y_max + 1], title=None)
                fig_line.update_xaxes(type="category", categoryorder="array", categoryarray=xcats, title=None)
                fig_line.update_layout(height=480)
            else:
                st.info("Select a ranking to visualize its evolution.")

        if fig_line is not None:
            _highlight_band(fig_line, sel_period_internal, list(xcats))
            st.plotly_chart(fig_line, use_container_width=True)

    # Detail table
    st.subheader("Faculty Detail")
    active = df_active()
    selected_ranking = (None if st.session_state.show_all or
                        st.session_state.single_ranking == "Select..."
                        else st.session_state.single_ranking)

    if selected_ranking:
        detail_df = active[active["Faculty Ranking"] == selected_ranking].copy()
        title_txt = f"### **{len(detail_df)}** **{selected_ranking}**"
    else:
        detail_df = active.copy()
        title_txt = f"### **{len(detail_df)}** Full-time Faculty"

    col_title, col_gender = st.columns([3, 2])
    with col_title:
        st.markdown(title_txt)
    with col_gender:
        gender_col = "Gender" if "Gender" in detail_df.columns else None
        if gender_col and len(detail_df):
            gen = detail_df[gender_col].value_counts()
            total_g = len(detail_df)
            pct_m = round(gen.get("Male", 0) / total_g * 100, 1) if total_g else 0
            pct_f = round(gen.get("Female", 0) / total_g * 100, 1) if total_g else 0
            df_gen = pd.DataFrame({
                "Gender": ["Male", "Female"], "P": [pct_m, pct_f], "Bar": [" ", " "],
            })
            fig_gen = px.bar(
                df_gen, x="P", y="Bar", color="Gender", text="Gender",
                color_discrete_map={"Male": "#003366", "Female": "#56d6c9"}, orientation="h",
            )
            fig_gen.update_traces(texttemplate='%{text} %{x}%', textposition="inside",
                                   textfont_size=16, width=0.7)
            fig_gen.update_layout(showlegend=False, xaxis_visible=False, yaxis_visible=False,
                                   height=80, margin=dict(l=0, r=0, t=0, b=0))
            st.plotly_chart(fig_gen, use_container_width=True)

    detail_cols = ["Periodo", "ID", "ID Nr.", "Full Name", "Academic Area",
                   "Faculty Ranking", "Subcategorization", "Faculty Qualific.", "P/S",
                   "Highest Earned Degree", "Year", "University", "Normal professional Resp."]
    show_cols = [c for c in detail_cols if c in detail_df.columns]
    detail_display = detail_df[show_cols].reset_index(drop=True)
    detail_display.index += 1
    st.dataframe(detail_display, use_container_width=True)
    _download_link("Descargar detalle (Excel)", detail_df[show_cols],
                   f"FT_Composition_Detail_{sel_period_label}.xlsx")


# 6) PÁGINA 2 — Full-time Faculty Staffing Levels
def page_staffing():
    all_periods = sorted(df["Periodo"].astype(str).unique().tolist())
    sem_periods = sorted(
        [p for p in all_periods if re.fullmatch(r'(?:19|20)\d{2}-(10|20)', p)],
        key=_period_sort_key,
    )  # cronológico (2020→2026) — así queda igual la tabla/gráficas de evolución

    # Sidebar específico de esta página
    with st.sidebar:
        st.markdown("#### Select Semester")
        sem_periods_desc = sorted(sem_periods, key=_period_sort_key, reverse=True)  # solo para el desplegable
        vis_opts = [p.replace("-", "") for p in sem_periods_desc]
        sel_vis = st.selectbox("", vis_opts, index=0 if vis_opts else None, key="ft_staff_periodo")
        sel_period_internal = sem_periods_desc[vis_opts.index(sel_vis)] if vis_opts else None
        sel_period_label = sel_vis

    _render_header("Full-time Faculty Staffing Levels", "New entrants, leavers, and headcount evolution")

    # Helpers (solo semestral)
    def perlist_sem():
        return sem_periods

    def final_count_series_sem(df_):
        return df_.groupby("Periodo")["ID"].nunique()

    def in_out_counts_sem(df_, label):
        counts = {}
        for p in sem_periods:
            flat = p.replace("-", "")
            counts[p] = int(df_["Notes"].astype(str).str.contains(
                fr"\b{label}\s+IN\s+\(?{flat}\)?\b", case=False, na=False
            ).sum())
        return pd.Series(counts)

    # Staffing summary table
    cols_summary = perlist_sem()
    fin_ser = final_count_series_sem(df).reindex(cols_summary, fill_value=0)
    new_ser = in_out_counts_sem(df, "IN").reindex(cols_summary, fill_value=0)
    out_ser = in_out_counts_sem(df, "OUT").reindex(cols_summary, fill_value=0)

    rows = []
    for i, key in enumerate(cols_summary):
        new_hires = int(new_ser.get(key, 0))
        leavers = int(out_ser.get(key, 0))
        if i == 0:
            start_val = int(fin_ser.iloc[0]) - new_hires + leavers
        else:
            start_val = int(fin_ser.iloc[i - 1])
        rows.append({
            "Start": int(start_val),
            "New": new_hires,
            "Leavers": leavers,
            "Final": int(fin_ser.iloc[i])
        })

    summary_df = pd.DataFrame(rows, index=cols_summary).T

    st.subheader("New entrants and leavers")

    def _bold_final_row(df_):
        styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
        if 'Final' in df_.index:
            styles.loc['Final', :] = 'font-weight:700;'
        return styles

    def _highlight_final_latest(df_):
        styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
        if len(df_.columns) > 0 and 'Final' in df_.index:
            candidates = [sel_period_label, sel_period_label.replace("-", ""), sel_period_internal]
            target = next((c for c in candidates if c in df_.columns), df_.columns[-1])
            styles.loc['Final', target] = 'background-color:#dff7f2; color:#00A896; font-weight:700;'
        return styles

    styled_summary = (
        summary_df
        .style
        .apply(_bold_final_row, axis=None)
        .apply(_highlight_final_latest, axis=None)
        .format(precision=0)
    )
    with st.container(key="tct_staff_summary"):
        st.dataframe(styled_summary, use_container_width=True)
    _scroll_table_right_once("tct_staff_summary")

    sum_left, sum_right = st.columns([1, 5])
    with sum_left:
        simple_tbl = summary_df.reset_index().rename(columns={"index": "Metric"})
        _download_link("Descargar tabla (Excel)", simple_tbl, "FT_New_Leavers_Semestral.xlsx")

    # Charts layout
    areas = sorted(df.get("Academic Area", pd.Series(dtype=object)).dropna().unique().tolist())
    col_left, col_right = st.columns([3, 2])

    with col_right:
        st.markdown(f"<div style='text-align:center;font-weight:700'>Period: {sel_period_label}</div>",
                   unsafe_allow_html=True)

        current_period = sel_period_internal or ""
        flat_list = [current_period.replace("-", "")] if current_period else []

        pat_in = "|".join([re.escape(f) for f in flat_list]) if flat_list else r"$^"
        df_in = df[df["Notes"].astype(str).str.contains(fr"\bIN\s+IN\s+\(?({pat_in})\)?\b", case=False, na=False)]
        df_out = df[df["Notes"].astype(str).str.contains(fr"\bOUT\s+IN\s+\(?({pat_in})\)?\b", case=False, na=False)]

        new_by_area = df_in.groupby("Academic Area")["ID"].nunique().reindex(areas, fill_value=0)
        left_by_area = df_out.groupby("Academic Area")["ID"].nunique().reindex(areas, fill_value=0)
        net_by_area = (new_by_area - left_by_area).astype(int)
        order = net_by_area.sort_values(ascending=True).index

        ret_vals = left_by_area.reindex(order).astype(int)
        new_vals = new_by_area.reindex(order).astype(int)

        fig_tornado = go.Figure()
        fig_tornado.add_trace(go.Bar(
            y=order, x=-ret_vals, orientation="h",
            name="Leavers", marker_color="#C0392B",
            text=ret_vals, texttemplate="%{text}", textposition="inside",
            insidetextanchor="middle", textfont=dict(size=14, color="white"),
            hovertemplate="Area: %{y}<br>Leavers: %{customdata}<extra></extra>",
            customdata=ret_vals
        ))
        fig_tornado.add_trace(go.Bar(
            y=order, x=new_vals, orientation="h",
            name="New", marker_color="#56d6c9",
            text=new_vals, texttemplate="%{text}", textposition="inside",
            insidetextanchor="middle", textfont=dict(size=14, color="white"),
            hovertemplate="Area: %{y}<br>New: %{customdata}<extra></extra>",
            customdata=new_vals
        ))

        fig_tornado.update_xaxes(showticklabels=False, showgrid=False, zeroline=False, showline=False)
        fig_tornado.update_yaxes(autorange="reversed")
        fig_tornado.update_layout(
            title=f"New vs Leavers by Area — {sel_period_label}",
            barmode="relative",
            height=max(360, 24 * len(areas)),
            margin=dict(l=10, r=10, t=20, b=80),
            legend=dict(orientation="h", y=-0.25, yanchor="top", x=0.5, xanchor="center"),
            xaxis_title=None, yaxis_title=None
        )
        st.plotly_chart(fig_tornado, use_container_width=True)

    with col_left:
        st.markdown("### Evolution of Faculty (Start vs Final)")

        x_periods = list(summary_df.columns)
        y_start = pd.to_numeric(summary_df.loc["Start"], errors="coerce").tolist() if "Start" in summary_df.index else []
        y_final = pd.to_numeric(summary_df.loc["Final"], errors="coerce").tolist() if "Final" in summary_df.index else []

        fig_line = go.Figure()
        fig_line.add_trace(go.Scatter(
            x=x_periods, y=y_start, mode="lines+markers",
            name="Start", line=dict(shape="linear", width=3, dash="dot")
        ))
        fig_line.add_trace(go.Scatter(
            x=x_periods, y=y_final, mode="lines+markers",
            name="Final", line=dict(shape="linear", width=3, color="#003366")
        ))

        for p, s, f in zip(x_periods, y_start, y_final):
            if pd.isna(s) or pd.isna(f):
                continue
            arrowcolor = "green" if f > s else ("red" if f < s else None)
            if not arrowcolor:
                continue
            fig_line.add_annotation(
                x=p, y=f, ax=p, ay=s,
                xref="x", yref="y", axref="x", ayref="y",
                showarrow=True, arrowhead=2, arrowsize=1.2, arrowwidth=2, arrowcolor=arrowcolor
            )

        fig_line.update_xaxes(type="category", tickangle=45, showgrid=True, title=None)
        fig_line.update_yaxes(showgrid=True, zeroline=False, title=None)
        fig_line.update_layout(
            height=380,
            margin=dict(l=10, r=10, t=10, b=80),
            legend=dict(orientation="h", y=-0.25, yanchor="top", x=0.5, xanchor="center"),
        )

        sel_for_band = sel_period_internal
        if sel_for_band in x_periods:
            pos = x_periods.index(sel_for_band)
            fig_line.add_shape(
                type="rect", xref="x", yref="paper",
                x0=pos - 0.4, x1=pos + 0.4, y0=0, y1=1,
                fillcolor="#D0E5F5", opacity=0.35, line_width=0
            )

        st.plotly_chart(fig_line, use_container_width=True)

    # Faculty details (semestre seleccionado) — el bloque de "View Faculty
    # details" con número grande y gráfica de género se quitó de aquí: el
    # género ahora se muestra en Composition. 'active' se mantiene porque
    # alimenta la tabla completa de abajo.
    active = df[df["Periodo"].astype(str).eq(sel_period_internal)].copy()

    # Full table
    st.markdown("### Complete Full-time table")
    cols_full = [
        "ID Nr.", "ID", "First Name", "Last Name",
        "Date of First Appointment to the School", "Academic Area",
        "Highest Degree", "Year", "Region were degree was obtained",
        "International Degree", "% devoted to Mission", "Faculty Ranking",
        "Subcategorization",
        "Country of Birth", "Double Nationality", "Date of Birth",
        "Age", "Gender", "Faculty Qualific.", "P/S",
        "Normal professional Resp.", "Notes"
    ]
    full = active[[c for c in cols_full if c in active.columns]].copy().reset_index(drop=True)

    if "Date of First Appointment to the School" in full.columns:
        full["Date of First Appointment to the School"] = pd.to_datetime(
            full["Date of First Appointment to the School"], errors="coerce"
        ).dt.date
    if "Date of Birth" in full.columns:
        full["Date of Birth"] = pd.to_datetime(full["Date of Birth"], errors="coerce").dt.date
    if "Year" in full.columns:
        full["Year"] = full["Year"].astype(str).str.extract(r'(\d{4})')

    with st.expander("Show complete table"):
        _download_link("Descargar tabla completa (Excel)", full, f"FT_Complete_Table_{sel_period_label}.xlsx")
        st.dataframe(full, use_container_width=False, hide_index=True)

    # Professor trajectory (PLANTA only)
    def _c(df0, *names):
        if df0 is None or df0.empty:
            return None
        cmap = {str(c).strip().casefold(): c for c in df0.columns}
        for n in names:
            key = str(n).strip().casefold()
            if key in cmap:
                return cmap[key]
        return None

    period_col = _c(df, "Periodo")
    id_col = _c(df, "ID Nr.", "ID Nr", "ID")
    fn_col = _c(df, "First Name")
    ln_col = _c(df, "Last Name")
    area_col = _c(df, "Academic Area", "AREA_PROFESOR")
    deg_col = _c(df, "Highest Degree")
    rank_col = _c(df, "Faculty Ranking")
    subc_col = _c(df, "Subcategorization")
    age_col = _c(df, "Age")
    qual_col = _c(df, "Faculty Qualific.")
    ps_col = _c(df, "P/S", "P - S")
    resp_col = _c(df, "Normal professional Resp.")
    notes_col = _c(df, "Notes")

    vals = sorted(df[period_col].dropna().astype(str).unique().tolist()) if period_col else []
    last_period = vals[-1] if vals else ""
    st.markdown(
        f"<div style='font-size:18px;font-weight:700;margin-top:18px'>"
        f"Search professor trajectory (PLANTA only) 2020-10 – {last_period}"
        f"</div>",
        unsafe_allow_html=True
    )

    if not all([period_col, id_col, fn_col, ln_col]):
        st.warning("Required columns were not found in the PLANTA sheet.")
    else:
        df_ids = df[[id_col, fn_col, ln_col]].copy().dropna(subset=[id_col])
        df_ids = df_ids.drop_duplicates(subset=[id_col], keep="last")
        df_ids["label"] = (
            df_ids[fn_col].astype(str).str.strip() + " " +
            df_ids[ln_col].astype(str).str.strip() +
            " — ID: " + df_ids[id_col].astype(str).str.strip()
        )
        df_ids = df_ids.sort_values("label")

        sel_label = st.selectbox(
            "Select a professor (PLANTA):",
            options=["(Select...)"] + df_ids["label"].tolist(),
            index=0,
            key="ft_staff_trajectory_select"
        )

        if sel_label and sel_label != "(Select...)":
            m = re.search(r"ID:\s*(.+)$", sel_label)
            chosen_id = m.group(1).strip() if m else None

            traj = df[df[id_col].astype(str).str.strip() == chosen_id].copy()

            out_cols_raw = [
                (period_col, "Periodo"),
                (id_col, "ID Nr."),
                (fn_col, "First Name"),
                (ln_col, "Last Name"),
                (area_col, "Academic Area"),
                (deg_col, "Highest Degree"),
                (rank_col, "Faculty Ranking"),
                (subc_col, "Subcategorization"),
                (age_col, "Age"),
                (qual_col, "Faculty Qualific."),
                (ps_col, "P/S"),
                (resp_col, "Normal professional Resp."),
                (notes_col, "Notes"),
            ]

            out_df = pd.DataFrame({
                new: (traj[orig] if (orig in traj.columns) else pd.Series([""] * len(traj), index=traj.index))
                for orig, new in out_cols_raw
            })
            out_df.columns = pd.Index(out_df.columns).map(str)

            OUT_COLOR = "#8B0000"
            IN_COLOR = "#00796B"

            def _matches_tag_for_period(note_upper: str, tag: str, flat_period: str) -> bool:
                pat = rf'\b{tag}\s+IN\s+\(?((?:19|20)\d{{2}}[-_/ ]?\d{{2}})\)?\b'
                m2 = re.search(pat, note_upper, flags=re.IGNORECASE)
                if not m2:
                    return False
                per_txt = m2.group(1)
                per_flat = re.sub(r'\D', '', per_txt)
                return flat_period and (per_flat == flat_period)

            def _color_in_out(row: pd.Series):
                per = str(row.get("Periodo", ""))
                note_upper = str(row.get("Notes", "")).upper()
                flat = re.sub(r'\D', '', per)

                is_out = _matches_tag_for_period(note_upper, "OUT", flat) or ("OUT IN" in note_upper)
                is_in = _matches_tag_for_period(note_upper, "IN", flat) or ("IN IN" in note_upper)

                if is_out:
                    return [f'color:{OUT_COLOR};font-weight:700;' for _ in row.index]
                if is_in:
                    return [f'color:{IN_COLOR};font-weight:700;' for _ in row.index]
                return ['' for _ in row.index]

            st.dataframe(
                out_df.reset_index(drop=True).style.apply(_color_in_out, axis=1).hide(axis="index"),
                use_container_width=True
            )

            _download_link("Descargar trayectoria (Excel)", out_df, f"Trajectory_{chosen_id}.xlsx")


# 7) PÁGINA 3 — Distribution by Academic Area
def page_area():
    MINT = "#00A896"
    HIGHLIGHT = "#D0E5F5"
    PALETTE = [
        "#056D62", "#1CDFCB", "#FF7F50", "#9B59B6", "#F4A261",
        "#1B6CA8", "#0EAD69", "#E76F51", "#3D5A80", "#8D99AE",
        "#78A7A2", "#F6BD60", "#6D597A", "#43AA8B", "#277DA1",
    ]

    def is_sem_label(p: str) -> bool:
        return bool(re.fullmatch(r"\d{4}-(10|20)", str(p)))

    def display_label_sem(p_internal: str) -> str:
        return str(p_internal).replace("-", "")


    def filter_for_timeframe(df_in: pd.DataFrame, time_mode: str, value) -> pd.DataFrame:
        """
        value:
          - Semestral: visible 'YYYY10'/'YYYY20', internamente 'YYYY-10/20'
          - Anual: 'YYYY' -> incluye ambos semestres + intersemestral, deduplicado por profesor/año
          - Intersemestral: 'YYYY Intersemestral'
        """
        if value is None:
            return df_in.iloc[0:0].copy()

        dfb = df_in.copy()

        if time_mode == "Semestral":
            sem_internal = f"{str(value)[:4]}-{str(value)[-2:]}"
            return dfb[dfb["Periodo"].astype(str).eq(sem_internal)].copy()

        if time_mode == "Anual":
            y = str(value)
            dfy = dfb[dfb["Periodo"].astype(str).str.startswith(y)].copy()
            if "ID" in dfy.columns:
                dfy["__Year"] = dfy["Periodo"].astype(str).str[:4]
                dfy = dfy.sort_values(by=["Periodo"]).drop_duplicates(subset=["ID", "__Year"], keep="last")
                dfy = dfy.drop(columns=["__Year"])
            return dfy

        return dfb[dfb["Periodo"].astype(str).eq(str(value))].copy()



    _render_header("Distribution by Academic Area", "Faculty distribution and evolution across academic areas")

    df_full = area_load_fulltime()
    df_part = area_load_parttime()

    st.session_state.setdefault("modo_faculty", "Full-time")


    # Sidebar
    with st.sidebar:
        st.markdown("#### Faculty Type")
        st.markdown('<div id="mode-pill">', unsafe_allow_html=True)
        mode_sidebar = st.radio(
            "Mode", ["Full-time", "Part-time"],
            index=0 if st.session_state.modo_faculty == "Full-time" else 1,
            horizontal=True, label_visibility="collapsed", key="mode_pill_radio",
        )
        st.markdown('</div>', unsafe_allow_html=True)

        if mode_sidebar != st.session_state.modo_faculty:
            st.session_state.modo_faculty = mode_sidebar
            st.rerun()

        df_base = df_full if st.session_state.modo_faculty == "Full-time" else df_part

        st.markdown("#### Timeframe")
        tmode = st.radio("Timeframe", ["Semestral", "Anual", "Intersemestral"],
                          key="time_mode_side", label_visibility="collapsed")

        all_periods = sorted(df_base["Periodo"].astype(str).dropna().unique().tolist())

        if tmode == "Semestral":
            period_opts = sorted([p for p in all_periods if is_sem_label(p)], key=_period_sort_key, reverse=True)
            visible_opts = [display_label_sem(p) for p in period_opts]
            sel_visible = st.selectbox("Periodo", visible_opts, index=0 if period_opts else None)
            sel_value = period_opts[visible_opts.index(sel_visible)] if period_opts else None
            sel_label = sel_visible
        elif tmode == "Anual":
            years = sorted(pd.Series(all_periods).astype(str).str[:4].unique().tolist(), reverse=True)
            sel_value = st.selectbox("Periodo", years, index=0 if years else None)
            sel_label = sel_value
        else:
            inters = sorted([p for p in all_periods if _is_inter_label(p)], key=_period_sort_key, reverse=True)
            sel_value = st.selectbox("Periodo", inters, index=0 if inters else None)
            sel_label = sel_value

        st.session_state["sel_tf_mode"] = tmode
        st.session_state["sel_tf_value"] = sel_value
        st.session_state["sel_tf_label"] = sel_label


    # Active dataset
    df = df_full.copy() if st.session_state.modo_faculty == "Full-time" else df_part.copy()

    tmode_now = st.session_state.get("sel_tf_mode", "Semestral")
    sel_value = st.session_state.get("sel_tf_value")
    sel_label = st.session_state.get("sel_tf_label")
    IDCOL = "ID"


    # Pivot table by academic area
    df_view = df.copy()

    if tmode_now == "Semestral":
        df_view = df_view[df_view["Periodo"].astype(str).apply(is_sem_label)].copy()
        df_view["Periodo_display"] = df_view["Periodo"].astype(str).map(display_label_sem)
        pivot_area = pd.pivot_table(df_view, index="AREA_PROFESOR", columns="Periodo_display",
                                     values=IDCOL, aggfunc="nunique", fill_value=0).sort_index()
        col_order = sorted(df_view["Periodo_display"].unique().tolist())

    elif tmode_now == "Intersemestral":
        df_view = df_view[df_view["Periodo"].astype(str).apply(_is_inter_label)].copy()
        df_view["Periodo_display"] = df_view["Periodo"].astype(str)
        pivot_area = pd.pivot_table(df_view, index="AREA_PROFESOR", columns="Periodo_display",
                                     values=IDCOL, aggfunc="nunique", fill_value=0).sort_index()
        col_order = sorted(df_view["Periodo_display"].unique().tolist())

    else:  # Anual
        df_view["__Year"] = df_view["Periodo"].astype(str).str[:4]
        df_view = df_view.sort_values(by=["Periodo"]).drop_duplicates(subset=[IDCOL, "__Year"], keep="last")
        pivot_area = pd.pivot_table(df_view, index="AREA_PROFESOR", columns="__Year",
                                     values=IDCOL, aggfunc="nunique", fill_value=0).sort_index()
        col_order = sorted(pivot_area.columns.astype(str).tolist())

    areas_palette_order = [a for a in pivot_area.index if a != "Total"]
    color_map_area = {a: PALETTE[i % len(PALETTE)] for i, a in enumerate(areas_palette_order)}

    pivot_area.loc["Total"] = pivot_area.sum(numeric_only=True)


    def style_bold_total(df_):
        styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
        if "Total" in df_.index:
            styles.loc["Total", :] = "font-weight:700;"
        return styles


    def style_total_lastcell(df_):
        styles = pd.DataFrame("", index=df_.index, columns=df_.columns)
        if len(df_.columns) > 0 and "Total" in df_.index:
            _sel = st.session_state.get("sel_tf_label", "")
            candidates = [_sel, str(_sel).replace("-", "")]
            target = next((c for c in candidates if c in df_.columns), df_.columns[-1])
            styles.loc["Total", target] = "background-color:#dff7f2; color:#00A896; font-weight:700;"
        return styles


    styled_area = (
        pivot_area[col_order].style
        .apply(style_bold_total, axis=None)
        .apply(style_total_lastcell, axis=None)
        .format(precision=0)
    )

    st.subheader(f"{st.session_state.modo_faculty} Faculty count")
    with st.container(key="tct_area_count"):
        st.dataframe(styled_area, use_container_width=True)
    _scroll_table_right_once("tct_area_count")

    pivot_download = pivot_area[col_order].reset_index()
    fname_pvt = f"Pivot_{'FT' if st.session_state.modo_faculty == 'Full-time' else 'PT'}_{tmode_now}.xlsx"
    _download_link("Descargar tabla (Excel)", pivot_download, fname_pvt)


    # Charts: evolution line + donut
    st.markdown(f"### Evolution by Academic Area — Number of {st.session_state.modo_faculty} Faculty")

    mode_key = "ft" if st.session_state.modo_faculty == "Full-time" else "pt"
    show_all_key = f"ver_todas_{mode_key}"
    area_sel_key = f"area_sel_{mode_key}"

    st.session_state.setdefault(show_all_key, True)
    st.session_state.setdefault(area_sel_key, "Select...")

    if tmode_now == "Semestral":
        base_line = df[df["Periodo"].astype(str).apply(is_sem_label)].copy()
        base_line["X"] = base_line["Periodo"].astype(str).map(display_label_sem)
        x_to_filter = sel_label
    elif tmode_now == "Intersemestral":
        base_line = df[df["Periodo"].astype(str).apply(_is_inter_label)].copy()
        base_line["X"] = base_line["Periodo"].astype(str)
        x_to_filter = sel_label
    else:  # Anual
        tmp = df.copy()
        tmp["__Year"] = tmp["Periodo"].astype(str).str[:4]
        tmp = tmp.sort_values(by=["Periodo"]).drop_duplicates(subset=[IDCOL, "__Year"], keep="last")
        base_line = tmp.rename(columns={"__Year": "X"}).copy()
        x_to_filter = str(sel_label)

    x_labels = sorted(base_line["X"].unique().tolist())

    totals_period = (
        base_line.groupby("X")[IDCOL].nunique()
        .reindex(x_labels).fillna(0).astype(int)
        .reset_index(name="Total")
    )


    colL, colR = st.columns([3, 2])

    with colL:
        areas_current = sorted([a for a in pivot_area.index if a != "Total"])
        if st.session_state[area_sel_key] not in ["Select...", *areas_current]:
            st.session_state[area_sel_key] = "Select..."

        st.checkbox("Show all lines", key=show_all_key)
        st.selectbox("Select academic area:", ["Select...", *areas_current], key=area_sel_key)

        show_all = st.session_state[show_all_key]
        area_sel_val = st.session_state[area_sel_key]

        if show_all:
            df_counts = base_line.groupby(["X", "AREA_PROFESOR"])[IDCOL].nunique().reset_index(name="Count")
            df_counts["X"] = pd.Categorical(df_counts["X"], categories=x_labels, ordered=True)
            df_counts = df_counts.merge(totals_period, on="X", how="left")
            df_counts["Pct"] = (df_counts["Count"] / df_counts["Total"].replace(0, pd.NA)).fillna(0)

            fig_line = px.line(
                df_counts, x="X", y="Pct", color="AREA_PROFESOR", markers=True,
                category_orders={"X": x_labels, "AREA_PROFESOR": areas_palette_order},
                color_discrete_map=color_map_area,
            )
            fig_line.update_traces(mode="lines+markers", line=dict(width=2),
                                    hovertemplate="<b>%{x}</b><br>%{fullData.name}: %{y:.1%}<extra></extra>")
            fig_line.update_xaxes(type="category", categoryorder="array", categoryarray=x_labels, title=None)
            fig_line.update_yaxes(rangemode="tozero", tickformat=".0%", title=None)
            fig_line.update_layout(showlegend=False)
            _highlight_band(fig_line, x_to_filter, x_labels, color=HIGHLIGHT)
            st.plotly_chart(fig_line, use_container_width=True)

        elif area_sel_val == "Select...":
            st.info("Select an academic area or enable 'Show all lines'.")

        else:
            df_area = (
                base_line[base_line["AREA_PROFESOR"] == area_sel_val]
                .groupby("X")[IDCOL].nunique()
                .reindex(x_labels).fillna(0).astype(int)
                .reset_index(name="Count")
            )
            df_line = df_area.merge(totals_period, on="X", how="left")
            df_line["Pct"] = (df_line["Count"] / df_line["Total"].replace(0, pd.NA)).fillna(0)
            df_line["X"] = pd.Categorical(df_line["X"], categories=x_labels, ordered=True)

            fig_line = px.line(
                df_line, x="X", y="Pct", markers=True, title=f"Evolution (% share) — {area_sel_val}",
                color_discrete_sequence=[color_map_area.get(area_sel_val, MINT)],
            )
            fig_line.update_traces(mode="lines+markers", line=dict(width=2),
                                    hovertemplate="<b>%{x}</b><br>%{y:.1%}<extra></extra>")
            fig_line.update_xaxes(type="category", categoryorder="array", categoryarray=x_labels, title=None)
            fig_line.update_yaxes(rangemode="tozero", tickformat=".0%", title=None)
            fig_line.update_layout(showlegend=False)
            _highlight_band(fig_line, x_to_filter, x_labels, color=HIGHLIGHT)
            st.plotly_chart(fig_line, use_container_width=True)

    with colR:
        st.markdown(f"##### Distribution by academic area — {sel_label or ''}")

        df_donut = filter_for_timeframe(df, tmode_now, sel_value)

        if df_donut.empty:
            st.info("No data for the selected period.")
        else:
            dist = df_donut.groupby("AREA_PROFESOR")[IDCOL].nunique().sort_values(ascending=False)
            donut_df = pd.DataFrame({"Area": dist.index, "Value": dist.values})
            donut_df["Area"] = donut_df["Area"].astype(str).str.replace(" & ", " &<br>", regex=False)

            _SOFT_DONUT_PALETTE = [
                "#8FBFB8", "#A7D8CF", "#F2B880", "#B7A3D1", "#F4C79A",
                "#7FA8C9", "#9BCBB0", "#E8A18C", "#8FA3BF", "#C3CBD6",
            ]
            soft_color_map = {a: _SOFT_DONUT_PALETTE[i % len(_SOFT_DONUT_PALETTE)] for i, a in enumerate(donut_df["Area"])}

            fig_donut = px.pie(
                donut_df, names="Area", values="Value", hole=0.45, color="Area",
                color_discrete_map=soft_color_map,
            )
            fig_donut.update_traces(
                texttemplate="%{label}<br>%{percent}", textposition="inside", insidetextorientation="horizontal",
                pull=[0.04] * len(donut_df), sort=False,
                textfont=dict(size=15, color="#1F2937", family="Arial, sans-serif"),
            )
            fig_donut.update_layout(showlegend=False, margin=dict(l=10, r=10, t=10, b=10), height=420,
                                     uniformtext_minsize=15, uniformtext_mode="show")
            st.plotly_chart(fig_donut, use_container_width=True)

            fname_donut = f"Donut_{'FT' if st.session_state.modo_faculty == 'Full-time' else 'PT'}_{tmode_now}_{str(sel_label).replace(' ', '_')}.xlsx"
            _download_link("Descargar tabla (Excel)", donut_df, fname_donut)


    # Detail table
    detail = filter_for_timeframe(df, tmode_now, sel_value)

    cols_prefer_ft = ["Full Name", "AREA_PROFESOR", "Faculty Ranking", "Faculty Qualific.", "P/S"]
    cols_prefer_pt = ["Profesor", "AREA_PROFESOR", "PLANTA_CATEDRA", "TIPO", "P/S"]
    prefer_cols = cols_prefer_ft if st.session_state.modo_faculty == "Full-time" else cols_prefer_pt

    cols_to_show = [c for c in prefer_cols if c in detail.columns]
    detail_out = detail[cols_to_show].drop_duplicates().reset_index(drop=True)

    count_label = int(detail[IDCOL].nunique()) if IDCOL in detail.columns else len(detail_out)
    faculty_word = "full-time" if st.session_state.modo_faculty == "Full-time" else "part-time"
    st.markdown(f"### There are {count_label} {faculty_word} Faculty in **{sel_label}**")
    detail_out = detail_out.reset_index(drop=True)
    detail_out.index += 1
    st.dataframe(detail_out, use_container_width=True)

    fname_det = f"Detail_{'FT' if st.session_state.modo_faculty == 'Full-time' else 'PT'}_{tmode_now}_{str(sel_label).replace(' ', '_')}.xlsx"
    _download_link("Descargar tabla (Excel)", detail_out, fname_det)


# 8) PÁGINA 4 — Faculty Demographics
def page_demographics():
    COLORS = {
        "primary": "#21877D", "primary_dark": "#004d47", "primary_light": "#dff7f2",
        "accent1": "#2EC4B6", "accent2": "#00A896", "accent3": "#56D6C9",
        "highlight": "#D0E5F5",
    }

    _render_header("Faculty Demographics", "PhD attainment, international diversity, and composition over time")

    df_full = demo_load_fulltime()
    df_part = demo_load_parttime()

    # Timeframe helpers
    def col_id(df_: pd.DataFrame) -> str | None:
        return "ID Nr." if "ID Nr." in df_.columns else ("ID" if "ID" in df_.columns else None)


    def col_degree(df_: pd.DataFrame) -> str | None:
        return next((c for c in ["Highest Degree", "TÍTULO"] if c in df_.columns), None)


    def col_gender(df_: pd.DataFrame) -> str | None:
        return next((c for c in ["Gender", "GÉNERO"] if c in df_.columns), None)


    def col_nationality(df_: pd.DataFrame) -> str | None:
        return next((c for c in ["Country of Birth", "Nationality"] if c in df_.columns), None)


    def is_semester_label(p: str) -> bool:
        return bool(re.fullmatch(r"\d{4}(10|20)", str(p)))

    def normalize_degree(series: pd.Series) -> pd.Series:
        s = series.astype(str).str.strip()
        is_tbd = s.str.upper().eq("TBD") | s.eq("") | s.str.lower().isin(["na", "none"])
        s_norm = s.str.lower().str.replace(".", "", regex=False)
        s_norm = s_norm.str.normalize("NFKD").str.encode("ascii", errors="ignore").str.decode("ascii")
        is_phd = s_norm.str.contains(r"\bphd\b") | s_norm.str.contains("doctor")
        is_master = s_norm.str.contains("master") | s_norm.str.contains(r"\bmsc\b") | s_norm.str.contains(r"\bms\b")
        is_bachelor = (s_norm.str.contains("bachelor") | s_norm.str.contains(r"\bbsc\b")
                       | s_norm.str.contains(r"\bbs\b") | s_norm.str.contains(r"\bba\b")
                       | s_norm.str.contains("licen"))
        out = pd.Series("Other", index=s.index, dtype=object)
        out[is_tbd] = "TBD"
        out[~is_tbd & is_phd] = "PhD"
        out[~is_tbd & ~is_phd & is_master] = "Master"
        out[~is_tbd & ~is_phd & ~is_master & is_bachelor] = "Bachelor"
        return out


    def filter_for_timeframe(df_in: pd.DataFrame, time_mode: str, sel_sem: str | None = None,
                              sel_year: int | None = None, sel_inter_label: str | None = None) -> pd.DataFrame:
        dfb = df_in.copy()
        pcol = "Periodo" if "Periodo" in dfb.columns else None
        scol = "Semestre" if "Semestre" in dfb.columns else None
        idc = col_id(dfb)

        if time_mode == "Semestral" and sel_sem:
            target = str(sel_sem)
            mask = pd.Series(False, index=dfb.index)
            if pcol:
                mask |= dfb[pcol].astype(str).eq(target)
            if scol:
                mask |= dfb[scol].astype(str).str.replace("-", "", regex=False).str.fullmatch(target, na=False)
            dfb = dfb[mask].copy()

        elif time_mode == "Anual" and sel_year is not None:
            y = str(sel_year)
            mask = pd.Series(False, index=dfb.index)
            if pcol:
                mask |= dfb[pcol].astype(str).str.startswith(y)
            if scol:
                mask |= dfb[scol].astype(str).str.startswith(y)
            dfb = dfb[mask].copy()
            if idc:
                sort_key = pcol or scol
                if sort_key:
                    dfb = dfb.sort_values(by=[sort_key])
                dfb = dfb.drop_duplicates(subset=[idc], keep="last")

        elif time_mode == "Intersemestral" and sel_inter_label:
            y = sel_inter_label.split()[0]
            mask = pd.Series(False, index=dfb.index)
            if scol:
                scol_n = dfb[scol].astype(str)
                mask |= scol_n.str.contains("inter", case=False, na=False) & scol_n.str.contains(y, na=False)
            if pcol:
                mask |= dfb[pcol].astype(str).eq(sel_inter_label)
            dfb = dfb[mask].copy()

        return dfb


    def options_for_timeframe(df_src: pd.DataFrame, time_mode: str) -> list:
        per = df_src["Periodo"].dropna().astype(str)
        if time_mode == "Semestral":
            return sorted([p for p in per.unique() if is_semester_label(p)])
        if time_mode == "Intersemestral":
            return sorted([p for p in per.unique() if _is_inter_label(p)])
        return sorted(per.str[:4].unique().tolist())


    def build_time_series(df_src: pd.DataFrame, time_mode: str, idcol: str,
                           degree_col: str | None, nat_col: str | None):
        labels, phd_pct, intl_pct = [], [], []

        def _phd_pct(sub):
            if sub.empty or degree_col is None:
                return 0.0
            sub = sub.copy()
            sub["__deg"] = normalize_degree(sub[degree_col])
            tot = sub[idcol].nunique()
            return 0.0 if tot == 0 else round(100 * sub.loc[sub["__deg"] == "PhD", idcol].nunique() / tot, 1)

        def _intl_pct(sub):
            if sub.empty or nat_col is None:
                return 0.0
            nat = sub[nat_col].astype(str).str.strip()
            is_valid = ~nat.eq("Colombian") & ~nat.str.upper().eq("TBD") & ~nat.eq("")
            tot = sub[idcol].nunique()
            return 0.0 if tot == 0 else round(100 * sub.loc[is_valid, idcol].nunique() / tot, 1)

        if time_mode in ("Semestral", "Intersemestral"):
            label_filter = is_semester_label if time_mode == "Semestral" else _is_inter_label
            periods = sorted([p for p in df_src["Periodo"].dropna().astype(str).unique() if label_filter(p)])
            for p in periods:
                sub = df_src[df_src["Periodo"].astype(str).eq(p)]
                labels.append(p)
                phd_pct.append(_phd_pct(sub))
                intl_pct.append(_intl_pct(sub))
        else:
            years = sorted(df_src["Periodo"].dropna().astype(str).str[:4].unique())
            for y in years:
                sub = filter_for_timeframe(df_src, "Anual", sel_year=int(y))
                labels.append(y)
                phd_pct.append(_phd_pct(sub))
                intl_pct.append(_intl_pct(sub))

        return labels, phd_pct, intl_pct


    AGE_LABELS = ["Under 30", "31-40", "41-50", "51-60", "over 61"]
    AGE_BINS = [-np.inf, 29, 40, 50, 60, np.inf]


    def age_buckets(series: pd.Series) -> pd.Series:
        return pd.cut(pd.to_numeric(series, errors="coerce"), bins=AGE_BINS, labels=AGE_LABELS)


    # Session defaults
    st.session_state.setdefault("modo_faculty", "Full-time")
    st.session_state.setdefault("time_mode_side", "Semestral")
    st.session_state.setdefault("sel_tf_label", None)


    # Sidebar
    with st.sidebar:
        st.markdown("#### Faculty Type")
        st.markdown('<div id="mode-pill">', unsafe_allow_html=True)
        mode_sidebar = st.radio(
            "Mode", ["Full-time", "Part-time"],
            index=0 if st.session_state.modo_faculty == "Full-time" else 1,
            horizontal=True, label_visibility="collapsed", key="mode_pill_radio",
        )
        if mode_sidebar != st.session_state.modo_faculty:
            st.session_state.modo_faculty = mode_sidebar
            st.session_state.sel_tf_label = None
            st.rerun()

        st.markdown("#### Timeframe")
        time_mode_side = st.radio(
            "Timeframe", ["Semestral", "Anual", "Intersemestral"],
            key="time_mode_side", label_visibility="collapsed",
        )

        df_base = df_full if st.session_state.modo_faculty == "Full-time" else df_part
        options_tf = sorted(options_for_timeframe(df_base, time_mode_side), key=_period_sort_key, reverse=True)

        sel_label = st.selectbox(
            "Periodo", options_tf,
            index=(options_tf.index(st.session_state.sel_tf_label)
                   if st.session_state.sel_tf_label in options_tf else 0),
        ) if options_tf else None

        if sel_label != st.session_state.get("sel_tf_label"):
            st.session_state.sel_tf_label = sel_label


    # Active dataset
    mode_now = st.session_state.get("modo_faculty", "Full-time")
    df = (df_full if mode_now == "Full-time" else df_part).copy()
    if "ID Nr." not in df.columns and "ID" in df.columns:
        df["ID Nr."] = df["ID"]

    sel_period_text = st.session_state.get("sel_tf_label") or ""
    st.subheader("Full-time demographics by Faculty ranking" if mode_now == "Full-time" else "Part-time demographic table")
    if sel_period_text and mode_now == "Full-time":
        st.markdown(f"<div class='period-label'>{sel_period_text}</div>", unsafe_allow_html=True)

    col_table, col_side = st.columns([3, 1.2])


    # Main table
    with col_table:
        IDCOL = col_id(df)
        if not IDCOL:
            st.error("ID column not found.")
            st.stop()

        tmode = st.session_state.get("time_mode_side", "Semestral")
        sel_lbl = st.session_state.get("sel_tf_label")
        GROUPS = {"Highest Degree", "Nationality", "Gender", "Age"}

        if not sel_lbl and mode_now == "Full-time":
            st.info("No data for the selected mode/period.")

        elif mode_now == "Full-time":
            if tmode == "Semestral":
                active = filter_for_timeframe(df, "Semestral", sel_sem=sel_lbl)
            elif tmode == "Anual":
                active = filter_for_timeframe(df, "Anual", sel_year=int(sel_lbl))
            else:
                active = filter_for_timeframe(df, "Intersemestral", sel_inter_label=sel_lbl)

            if "Faculty Ranking" in active.columns:
                base_order = ["Full Professor", "Associate Professor", "Assistant Professor", "Instructor"]
                uniq = active["Faculty Ranking"].dropna().unique().tolist()
                ranking_order = [x for x in base_order if x in uniq] + [x for x in uniq if x not in base_order]
            else:
                ranking_order = []

            cols_out = ["Category"] + ranking_order + ["Total"]

            def counts_by_ranking(df_sub: pd.DataFrame) -> pd.Series:
                if ranking_order:
                    s = (df_sub.groupby("Faculty Ranking")[IDCOL].nunique()
                         .reindex(ranking_order, fill_value=0).astype(int))
                else:
                    s = pd.Series(dtype=int)
                s.loc["Total"] = int(s.sum()) if not s.empty else int(df_sub[IDCOL].nunique())
                return s

            rows = []
            dcol = col_degree(active)
            if dcol and not active.empty:
                active["Degree_norm"] = normalize_degree(active[dcol])
                rows.append(pd.Series({"Category": "Highest Degree",
                                        **counts_by_ranking(active[active["Degree_norm"].isin(["PhD", "Master", "Bachelor"])]).to_dict()}))
                for d in ["PhD", "Master", "Bachelor"]:
                    rows.append(pd.Series({"Category": d, **counts_by_ranking(active[active["Degree_norm"] == d]).to_dict()}))

            ncol = col_nationality(active)
            if ncol and not active.empty:
                rows.append(pd.Series({"Category": "Nationality", **counts_by_ranking(active).to_dict()}))
                rows.append(pd.Series({"Category": "Colombian", **counts_by_ranking(active[active[ncol].astype(str).eq("Colombian")]).to_dict()}))
                rows.append(pd.Series({"Category": "International", **counts_by_ranking(active[~active[ncol].astype(str).eq("Colombian")]).to_dict()}))

            gcol = col_gender(active)
            if gcol and not active.empty:
                rows.append(pd.Series({"Category": "Gender", **counts_by_ranking(active[active[gcol].astype(str).isin(["Male", "Female"])]).to_dict()}))
                for g in ["Male", "Female"]:
                    rows.append(pd.Series({"Category": g, **counts_by_ranking(active[active[gcol].astype(str) == g]).to_dict()}))

            if not active.empty and "Age" in active.columns:
                active["Age_bucket"] = age_buckets(active["Age"])
                rows.append(pd.Series({"Category": "Age", **counts_by_ranking(active[active["Age_bucket"].notna()]).to_dict()}))
                for b in AGE_LABELS:
                    rows.append(pd.Series({"Category": b, **counts_by_ranking(active[active["Age_bucket"] == b]).to_dict()}))

            table_df = pd.DataFrame(rows).reindex(columns=cols_out).fillna(0)

            if not table_df.empty:
                numeric_cols = [c for c in (ranking_order + ["Total"]) if c in table_df.columns]
                is_group = table_df["Category"].isin(GROUPS)
                all_zero = (table_df[numeric_cols].sum(axis=1) == 0) if numeric_cols else pd.Series(False, index=table_df.index)
                table_df = table_df.loc[~(all_zero & ~is_group)].copy()
                for c in numeric_cols:
                    table_df[c] = pd.to_numeric(table_df[c], errors="coerce").fillna(0).astype(int)

            if table_df.empty:
                st.info("No rows to display for this selection.")
            else:
                mint_light = "#dff7f2"
                mint_dark = "#004d47"

                def style_group_rows(df_):
                    styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
                    mask = df_["Category"].isin(GROUPS)
                    styles.loc[mask, ["Category"] + ranking_order + ["Total"]] = 'background-color:#f2f2f2;'
                    styles.loc[mask, "Category"] += 'font-weight:700;'
                    return styles

                def style_total_col(df_):
                    styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
                    styles.loc[:, "Total"] = 'font-weight:700;'
                    return styles

                def style_group_totals(df_):
                    styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
                    mask = df_["Category"].isin(GROUPS)
                    styles.loc[mask, "Total"] = f'background-color:{mint_light}; color:{mint_dark}; font-weight:800;'
                    return styles

                styled_table = (
                    table_df.style
                    .apply(style_group_rows, axis=None)
                    .apply(style_total_col, axis=None)
                    .apply(style_group_totals, axis=None)
                    .format(precision=0, na_rep="")
                    .hide(axis="index")
                )
                st.dataframe(styled_table, use_container_width=True, height=48 + 33 * (len(table_df) + 1), hide_index=True)

        else:  # Part-time
            active = df.copy()
            if tmode == "Semestral":
                active = active[active["Periodo"].astype(str).apply(is_semester_label)].copy()
                keys = sorted(active["Periodo"].dropna().astype(str).unique().tolist())
            elif tmode == "Intersemestral":
                active = active[active["Periodo"].astype(str).apply(_is_inter_label)].copy()
                keys = sorted(active["Periodo"].dropna().astype(str).unique().tolist())
            else:
                active["__Year"] = active["Periodo"].astype(str).str[:4]
                active = active.sort_values(by=["Periodo"]).drop_duplicates(subset=[IDCOL, "__Year"], keep="last")
                keys = sorted(active["__Year"].dropna().astype(str).unique().tolist())

            def counts_by_key(df_sub: pd.DataFrame) -> pd.Series:
                group_col = "__Year" if tmode == "Anual" else "Periodo"
                return df_sub.groupby(group_col)[IDCOL].nunique().reindex(keys, fill_value=0).astype(int)

            deg_col = "TÍTULO" if "TÍTULO" in active.columns else ("Highest Degree" if "Highest Degree" in active.columns else None)
            nat_col = "Nationality" if "Nationality" in active.columns else ("Country of Birth" if "Country of Birth" in active.columns else None)
            gen_col = "GÉNERO" if "GÉNERO" in active.columns else ("Gender" if "Gender" in active.columns else None)

            rows = []
            GROUPS_PT = {"Highest Degree", "Nationality", "Gender", "Age"}

            if deg_col and not active.empty:
                active["Degree_norm"] = normalize_degree(active[deg_col])
                rows.append(pd.Series({"Category": "Highest Degree", **counts_by_key(active).to_dict()}))
                for d in ["PhD", "Master", "Bachelor", "TBD"]:
                    rows.append(pd.Series({"Category": d, **counts_by_key(active[active["Degree_norm"] == d]).to_dict()}))

            if nat_col and not active.empty:
                nat = active[nat_col].astype(str).str.strip()
                is_tbd_nat = nat.str.upper().eq("TBD")
                is_col = nat.eq("Colombian")
                is_int = ~is_col & ~is_tbd_nat & nat.ne("")
                rows.append(pd.Series({"Category": "Nationality", **counts_by_key(active).to_dict()}))
                rows.append(pd.Series({"Category": "Colombian", **counts_by_key(active[is_col]).to_dict()}))
                rows.append(pd.Series({"Category": "International", **counts_by_key(active[is_int]).to_dict()}))
                rows.append(pd.Series({"Category": "TBD (Nationality)", **counts_by_key(active[is_tbd_nat]).to_dict()}))

            if gen_col and not active.empty:
                g = active[gen_col].astype(str).str.strip()
                is_tbd_g = g.str.upper().eq("TBD") | g.eq("")
                rows.append(pd.Series({"Category": "Gender", **counts_by_key(active).to_dict()}))
                for gval in ["Male", "Female"]:
                    rows.append(pd.Series({"Category": gval, **counts_by_key(active[g.eq(gval)]).to_dict()}))
                rows.append(pd.Series({"Category": "TBD (Gender)", **counts_by_key(active[is_tbd_g]).to_dict()}))

            if "Age" in active.columns and not active.empty:
                active = active.assign(Age_bucket=age_buckets(active["Age"]))
                is_tbd_age = active["Age_bucket"].isna()
                rows.append(pd.Series({"Category": "Age", **counts_by_key(active).to_dict()}))
                for b in AGE_LABELS:
                    rows.append(pd.Series({"Category": b, **counts_by_key(active[active["Age_bucket"] == b]).to_dict()}))
                rows.append(pd.Series({"Category": "TBD (Age)", **counts_by_key(active[is_tbd_age]).to_dict()}))

            ycol = "Years Industry experience"
            if ycol in active.columns:
                yseries = pd.to_numeric(active[ycol], errors="coerce")
                group_col = "__Year" if tmode == "Anual" else "Periodo"
                avg_series = active.assign(**{ycol: yseries}).groupby(group_col)[ycol].mean().reindex(keys).round(1)
                rows.append(pd.Series({"Category": "Avg years of Work Exp.", **avg_series.to_dict()}))
                rows.append(pd.Series({"Category": "TBD (Work Exp.)", **counts_by_key(active[active[ycol].isna()]).to_dict()}))

            cols_out = ["Category"] + keys
            table_df = pd.DataFrame(rows).reindex(columns=cols_out).fillna(0)

            if not table_df.empty:
                is_group_or_avg = table_df["Category"].isin(GROUPS_PT) | table_df["Category"].eq("Avg years of Work Exp.")
                all_zero = table_df[keys].sum(axis=1) == 0
                table_df = table_df.loc[~(all_zero & ~is_group_or_avg)].copy()

            if table_df.empty:
                st.info("No rows to display.")
                st.stop()

            mask_avg = table_df["Category"].eq("Avg years of Work Exp.")
            display_df = table_df.copy().astype(object)
            for c in keys:
                display_df.loc[mask_avg, c] = pd.to_numeric(table_df.loc[mask_avg, c], errors="coerce").map(
                    lambda x: "" if pd.isna(x) else f"{x:.1f}")
                display_df.loc[~mask_avg, c] = pd.to_numeric(table_df.loc[~mask_avg, c], errors="coerce").fillna(0).astype(int).map(str)

            blue_light, blue_dark = "#dff7f2", "#00A896"
            red_light, red_dark = "#f8d7da", "#721c24"

            def style_gray(df_):
                styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
                gray_rows = df_["Category"].isin(GROUPS_PT) | df_["Category"].eq("Avg years of Work Exp.")
                styles.loc[gray_rows, ["Category"] + keys] = 'background-color:#f2f2f2; font-weight:700;'
                return styles

            def style_last_col(df_):
                styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
                if keys:
                    _sel = st.session_state.get("sel_tf_label", "")
                    target = _sel if _sel in keys else keys[-1]
                    mask = df_["Category"].isin(GROUPS_PT)
                    styles.loc[mask, target] = f'background-color:{blue_light}; color:{blue_dark}; font-weight:800;'
                return styles

            def style_tbd(df_):
                styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
                for i, row in df_.iterrows():
                    if "TBD" in str(row["Category"]):
                        for c in keys:
                            v = pd.to_numeric(table_df.loc[i, c], errors="coerce")
                            if pd.notna(v) and v > 0:
                                styles.at[i, c] = f'background-color:{red_light}; color:{red_dark}; font-weight:800;'
                return styles

            styled_table = (
                display_df.style
                .apply(style_gray, axis=None)
                .apply(style_last_col, axis=None)
                .apply(style_tbd, axis=None)
                .hide(axis="index")
            )
            with st.container(key="tct_demo_parttime"):
                st.dataframe(styled_table, use_container_width=True, height=48 + 33 * (len(display_df) + 1), hide_index=True)
            _scroll_table_right_once("tct_demo_parttime")


    # Side KPI charts
    with col_side:
        IDCOL = col_id(df)
        if not IDCOL:
            st.stop()

        sel_lbl = st.session_state.get("sel_tf_label")
        tmode = st.session_state.get("time_mode_side", "Semestral")

        if not sel_lbl:
            st.info("Select a period.")
        else:
            st.markdown(f"<div class='period-label'>{sel_lbl}</div>", unsafe_allow_html=True)

            if tmode == "Semestral":
                active_side = filter_for_timeframe(df, "Semestral", sel_sem=sel_lbl)
            elif tmode == "Anual":
                active_side = filter_for_timeframe(df, "Anual", sel_year=int(sel_lbl))
            else:
                active_side = filter_for_timeframe(df, "Intersemestral", sel_inter_label=sel_lbl)

            total_act = int(active_side[IDCOL].nunique()) if not active_side.empty else 0
            mint, mint_dark, gauge_bg = "#00A896", "#004d47", "#E8FAF7"

            def make_gauge(pct: float, label: str):
                fig = go.Figure(go.Indicator(
                    mode="gauge", value=pct,
                    gauge={'axis': {'range': [0, 100]}, 'bar': {'color': mint}, 'bgcolor': gauge_bg},
                ))
                fig.add_annotation(x=0.5, y=0.40, xref="paper", yref="paper", text=label,
                                    showarrow=False, font=dict(color=mint_dark, size=13))
                fig.add_annotation(x=0.5, y=0.0, xref="paper", yref="paper", text=f"{pct:.1f}%",
                                    showarrow=False, font=dict(color=mint_dark, size=18))
                fig.update_layout(height=110, margin=dict(l=10, r=10, t=10, b=6))
                return fig

            dcol = col_degree(active_side)
            phds = 0
            if total_act and dcol:
                active_side["Degree_norm"] = normalize_degree(active_side[dcol])
                phds = int(active_side[active_side["Degree_norm"] == "PhD"][IDCOL].nunique())
            pct_phd = round(100 * phds / total_act, 1) if total_act else 0.0
            st.plotly_chart(make_gauge(pct_phd, "PhD%:"), use_container_width=True)

            ncol = col_nationality(active_side)
            pct_int = 0.0
            if total_act and ncol:
                nat = active_side[ncol].astype(str).str.strip()
                is_int = ~nat.eq("Colombian") & ~nat.str.upper().eq("TBD") & ~nat.eq("")
                pct_int = round(100 * int(active_side[is_int][IDCOL].nunique()) / total_act, 1)
            st.plotly_chart(make_gauge(pct_int, "International%:"), use_container_width=True)

            gcol = col_gender(active_side)
            male = int(active_side[active_side[gcol].astype(str) == "Male"][IDCOL].nunique()) if gcol and total_act else 0
            female = int(active_side[active_side[gcol].astype(str) == "Female"][IDCOL].nunique()) if gcol and total_act else 0
            pct_m = round(100 * male / total_act, 1) if total_act else 0.0
            pct_f = round(100 * female / total_act, 1) if total_act else 0.0

            fig_gender = go.Figure()
            fig_gender.add_trace(go.Bar(x=[pct_m], y=[" "], orientation="h", name="Male",
                                         text=[f"Male {pct_m}%"], textposition="inside", insidetextanchor="middle"))
            fig_gender.add_trace(go.Bar(x=[pct_f], y=[" "], orientation="h", name="Female",
                                         text=[f"Female {pct_f}%"], textposition="inside", insidetextanchor="middle"))
            fig_gender.update_layout(barmode="stack", showlegend=False,
                                      xaxis=dict(range=[0, 100], visible=False), yaxis=dict(visible=False),
                                      height=100, margin=dict(l=10, r=10, t=18, b=12))
            st.plotly_chart(fig_gender, use_container_width=True)

            if not active_side.empty and "Age" in active_side.columns:
                active_side = active_side.assign(Age_bucket=age_buckets(active_side["Age"]))
                age_counts = (active_side.groupby("Age_bucket")[IDCOL].nunique()
                              .reindex(AGE_LABELS, fill_value=0).reset_index(name="Count"))
            else:
                age_counts = pd.DataFrame({"Age_bucket": AGE_LABELS, "Count": [0] * len(AGE_LABELS)})

            fig_age = px.bar(age_counts, x="Count", y="Age_bucket", orientation="h", text="Count")
            fig_age.update_traces(marker_color="#00A896", textposition="outside", texttemplate="%{text}")
            fig_age.update_xaxes(range=[0, 35], title=None)
            fig_age.update_yaxes(title=None, autorange="reversed")
            fig_age.update_layout(height=200, margin=dict(l=10, r=10, t=0, b=12))
            st.plotly_chart(fig_age, use_container_width=True)


    # Row 1: % PhD over time + PhD by region
    st.markdown("---")

    tmode_ts = st.session_state.get("time_mode_side", "Semestral")
    IDCOL = col_id(df)
    labels_ts, phd_ts, intl_ts = build_time_series(df, tmode_ts, IDCOL, col_degree(df), col_nationality(df))

    sel_lbl = st.session_state.get("sel_tf_label")
    if labels_ts:
        if tmode_ts == "Anual":
            period_current = sel_lbl if sel_lbl in labels_ts else labels_ts[-1]
        elif tmode_ts == "Intersemestral":
            period_current = sel_lbl if (sel_lbl in labels_ts and _is_inter_label(sel_lbl)) else labels_ts[-1]
        else:
            period_current = sel_lbl if (sel_lbl in labels_ts and is_semester_label(sel_lbl)) else labels_ts[-1]
    else:
        period_current = None

    row1_left, row1_right = st.columns([2, 1])

    if mode_now == "Part-time":
        y_min_phd, y_max_phd, bar_h = 0, 30, 220
    else:
        y_min_phd, y_max_phd, bar_h = 70, 100, 220
    if tmode_ts == "Intersemestral":
        y_min_phd, y_max_phd = 0, 100

    line_h = bar_h + 380 + 140  # iguala la altura combinada de la barra de región + el mapa, y un poco más

    with row1_left:
        title_combo = (
            "% PhD & % International — Full-time Faculty" if mode_now == "Full-time"
            else "% PhD & % International — Part-time Faculty"
        )
        fig_combo = go.Figure()
        fig_combo.add_trace(go.Scatter(
            x=labels_ts, y=phd_ts, name="% PhD", mode="lines+markers+text",
            line=dict(color="#00A896", width=3), marker=dict(size=7, color="#00A896"),
            text=[f"{v:.1f}%" for v in phd_ts], textposition="top center",
        ))
        fig_combo.add_trace(go.Scatter(
            x=labels_ts, y=intl_ts, name="% International", mode="lines+markers+text",
            line=dict(color="#2E6FC4", width=3), marker=dict(size=7, color="#2E6FC4"),
            text=[f"{v:.1f}%" for v in intl_ts], textposition="top center",
        ))
        fig_combo.update_layout(
            title=title_combo,
            xaxis=dict(type="category", categoryorder="array", categoryarray=labels_ts,
                       range=[-0.5, max(len(labels_ts) - 0.5, 0.5)], tickangle=-45),
            yaxis=dict(range=[0, 100], title="%"),
        )
        _highlight_band(fig_combo, period_current, labels_ts, color=COLORS["highlight"])
        # Anotación al final de cada línea con su valor: "XX.X% ... with PhD" / "XX.X% ... International Faculty"
        if labels_ts:
            last_phd = phd_ts[-1] if len(phd_ts) else 0
            last_int = intl_ts[-1] if len(intl_ts) else 0
            fig_combo.add_annotation(x=labels_ts[-1], y=last_phd, yref="y", text=f"{last_phd:.1f}% ... with PhD",
                                      showarrow=False, xanchor="left", xshift=8, font=dict(color="#00A896", size=12))
            fig_combo.add_annotation(x=labels_ts[-1], y=last_int, yref="y", text=f"{last_int:.1f}% ... International Faculty",
                                      showarrow=False, xanchor="left", xshift=8, yshift=-14, font=dict(color="#2E6FC4", size=12))
        fig_combo.update_layout(height=line_h + 80, margin=dict(l=10, r=90, t=40, b=60),
                                 legend=dict(orientation="h", yanchor="top", y=-0.15, x=0.5, xanchor="center", title=None))
        st.plotly_chart(fig_combo, use_container_width=True)

    with row1_right:
        if period_current is None:
            active_p = df.iloc[0:0].copy()
        elif tmode_ts == "Anual":
            active_p = filter_for_timeframe(df, "Anual", sel_year=int(period_current))
        else:
            active_p = df[df["Periodo"].astype(str).eq(str(period_current))].copy()

        dcol_here = col_degree(df)
        phd_now_all = pd.DataFrame(columns=df.columns)
        if dcol_here is not None and not active_p.empty:
            active_p["Degree_norm"] = normalize_degree(active_p[dcol_here])
            phd_now_all = active_p[active_p["Degree_norm"].eq("PhD")].copy()

        # Total de PhD: cuenta TODOS los PhD, sin importar si tienen la región
        # diligenciada (antes dependía de eso y por eso mostraba 0 cuando la
        # mayoría de los profesores nuevos tenían región en TBD).
        total_phd_valid = int(phd_now_all[IDCOL].nunique()) if not phd_now_all.empty else 0

        # planta usa el nombre viejo con el error de tipeo original ("were" en
        # vez de "Where"); Info. Profesores ya lo tiene bien escrito — se
        # aceptan ambos para que funcione igual en Full-time y Part-time.
        region_col = next(
            (c for c in ["Region Where it was obtained", "Region were degree was obtained"]
             if c in phd_now_all.columns), None
        )
        if region_col:
            reg = phd_now_all[region_col].astype(str).str.strip()
            mask_valid_region = ~reg.eq("") & ~reg.str.upper().eq("TBD")
            phd_for_regions = phd_now_all[mask_valid_region].copy()
            phd_for_regions[region_col] = phd_for_regions[region_col].astype(str).str.strip()
        else:
            phd_for_regions = pd.DataFrame(columns=df.columns)

        phd_int = 0
        intl_col = next((c for c in ["International Degree?", "International Degree"]
                          if c in phd_now_all.columns), None)
        if intl_col:
            phd_int = int(phd_now_all[phd_now_all[intl_col].astype(str).str.strip().str.lower().eq("yes")][IDCOL].nunique())

        if not phd_for_regions.empty and region_col:
            reg_counts = (phd_for_regions.groupby(region_col)[IDCOL].nunique()
                          .sort_values(ascending=False).reset_index()
                          .rename(columns={region_col: "Region", IDCOL: "Count"}))
        else:
            reg_counts = pd.DataFrame({"Region": [], "Count": []})

        title_phd_bar = f"{total_phd_valid} professors with a PhD, {phd_int} obtained it abroad" if phd_int else f"{total_phd_valid} professors with a PhD"
        fig_phd_reg = px.bar(reg_counts, x="Count", y="Region", orientation="h", title=title_phd_bar, text="Count")
        fig_phd_reg.update_traces(marker_color="#00A896", textposition="outside", texttemplate="%{text}")
        fig_phd_reg.update_xaxes(title=None, dtick=1)
        fig_phd_reg.update_yaxes(title=None, autorange="reversed")
        fig_phd_reg.update_layout(height=bar_h, margin=dict(l=10, r=10, t=50, b=6))
        st.plotly_chart(fig_phd_reg, use_container_width=True)

        def pick_cols(df_, mapping):
            out = {}
            for new, opts in mapping.items():
                match = next((c for c in opts if c in df_.columns), None)
                out[new] = df_[match] if match else pd.Series([""] * len(df_), index=df_.index)
            return pd.DataFrame(out)

        detalle_source = phd_now_all if not phd_now_all.empty else phd_for_regions
        if not detalle_source.empty:
            detalle_phd = pick_cols(detalle_source, {
                "Full Name": ["Full Name", "Full-Name", "Full_Name", "Profesor", "First Name"],
                "Highest Earned Degree": ["Highest Earned Degree", "Highest Degree", "TÍTULO"],
                "University": ["University", "University Name", "University2"],
                "Region Where it was obtained": ["Region Where it was obtained", "Region were degree was obtained", "Region"],
                "Year": ["Year", "Year Earned ", "Year Degree", "Year Earned", "Highest Degree, Year Earned"],
            })
            open_phd_detail = st.button("Show PhDs", key="open_phd_detail", use_container_width=True)
            if open_phd_detail:
                if hasattr(st, "dialog"):
                    @st.dialog("Profesores con PhD", width="large")
                    def _dlg_phd():
                        st.dataframe(detalle_phd.reset_index(drop=True), use_container_width=True, hide_index=True)
                    _dlg_phd()
                else:
                    with st.expander("Profesores con PhD", expanded=True):
                        st.dataframe(detalle_phd.reset_index(drop=True), use_container_width=True, hide_index=True)

        # Nacionalidades del periodo actual (para el mapa de burbujas de abajo)
        nat_col = col_nationality(df)
        intl_now = pd.DataFrame(columns=df.columns)
        if nat_col and period_current:
            if tmode_ts == "Anual":
                active_p2 = filter_for_timeframe(df, "Anual", sel_year=int(period_current))
            else:
                active_p2 = df[df["Periodo"].astype(str).eq(str(period_current))].copy()

            nat = active_p2[nat_col].astype(str).str.strip()
            is_valid = ~nat.eq("Colombian") & ~nat.str.upper().eq("TBD") & ~nat.eq("")
            intl_now = active_p2[is_valid].copy()

            nat_counts = (intl_now.groupby(nat_col)[IDCOL].nunique().sort_values(ascending=False)
                          .reset_index().rename(columns={nat_col: "Nationality", IDCOL: "Count"}))
            total_intl = int(intl_now[IDCOL].nunique()) if not intl_now.empty else 0
            n_nats = int(nat_counts["Nationality"].nunique()) if not nat_counts.empty else 0
        else:
            nat_counts = pd.DataFrame({"Nationality": [], "Count": []})
            total_intl = n_nats = 0

        # Nacionalidad (gentilicio) -> país, para poder ubicar la burbuja en el mapa
        _NATIONALITY_TO_COUNTRY = {
            "American": "United States", "Argentinian": "Argentina", "Australian": "Australia",
            "Brazilian": "Brazil", "British": "United Kingdom", "Bulgarian": "Bulgaria",
            "Canadian": "Canada", "Chilean": "Chile", "Dominican": "Dominican Republic",
            "Egyptian": "Egypt", "French": "France", "German": "Germany", "Indian": "India",
            "Italian": "Italy", "Kenyan": "Kenya", "New Zealander": "New Zealand",
            "Peruvian": "Peru", "Philippine": "Philippines", "Portuguese": "Portugal",
            "Russian": "Russia", "South African": "South Africa", "Spanish": "Spain",
            "Turkish": "Turkey", "Venezuelan": "Venezuela", "Dutch": "Netherlands",
            "Belgian": "Belgium", "Finnish": "Finland", "Mexican": "Mexico",
        }
        nat_counts["Country"] = nat_counts["Nationality"].map(_NATIONALITY_TO_COUNTRY)
        map_df = nat_counts.dropna(subset=["Country"])

        title_nat = f"{total_intl} international Faculty. {n_nats} different nationalities"
        fig_nat = px.scatter_geo(
            map_df, locations="Country", locationmode="country names", size="Count",
            text="Nationality", hover_name="Nationality", hover_data={"Count": True, "Country": False},
            title=title_nat, projection="natural earth", color_discrete_sequence=["#2E6FC4"],
        )
        fig_nat.update_traces(marker=dict(color="#2E6FC4", opacity=0.75, line=dict(width=1, color="#FFFFFF")),
                               mode="markers+text", textposition="top center",
                               textfont=dict(size=10, color="#374151"))
        fig_nat.update_geos(showcountries=True, countrycolor="#E5E7EB", showland=True, landcolor="#F8FFFE",
                             showocean=True, oceancolor="#EAF6F4", bgcolor="rgba(0,0,0,0)")
        fig_nat.update_layout(height=380, margin=dict(l=10, r=10, t=50, b=6))
        st.plotly_chart(fig_nat, use_container_width=True)

        if not intl_now.empty:
            detalle_nat = pick_cols(intl_now, {
                "Full Name": ["Full Name", "Full-Name", "Full_Name", "Profesor", "First Name"],
                "Nationality": ["Nationality", "Country of Birth"],
            })
            open_nat_detail = st.button("Show Nationalities", key="open_nat_detail", use_container_width=True)
            if open_nat_detail:
                if hasattr(st, "dialog"):
                    @st.dialog("Nacionalidad de profesores", width="small")
                    def _dlg_nat():
                        st.dataframe(detalle_nat.reset_index(drop=True), use_container_width=True, hide_index=True)
                    _dlg_nat()
                else:
                    with st.expander("Nacionalidad de profesores", expanded=True):
                        st.dataframe(detalle_nat.reset_index(drop=True), use_container_width=True, hide_index=True)


# 9) PÁGINA 5 — Full-time Faculty Activities
def page_activities():
    # Helper functions (from original)
    def resolve_column(df: pd.DataFrame, target: str) -> Optional[str]:
        t = target.strip().casefold()
        for c in df.columns:
            if c.strip().casefold() == t:
                return c
        return None

    @st.cache_data(ttl=0)
    def load_fulltime():
        df = pd.read_excel(io.BytesIO(_download_drive_file_bytes(PROFESORES_FILE_ID)), sheet_name="planta")
        raw = df.iloc[:, 0].astype(str)
        df["Periodo"] = raw.str.slice(0, 4) + "-" + raw.str.slice(4, 6)
        if "ID Nr." in df.columns and "ID" not in df.columns:
            df = df.rename(columns={"ID Nr.": "ID"})
        df.columns = df.columns.str.strip()
        return df

    @st.cache_data(ttl=0)
    def load_questionnaire():
        df = pd.read_excel(io.BytesIO(_download_drive_file_bytes(QUESTIONNAIRE_FILE_ID)), sheet_name="Faculty_questionnaire")
        df.columns = df.columns.str.strip()
        ycol = resolve_column(df, "Year")
        if ycol:
            df[ycol] = pd.to_numeric(df[ycol], errors="coerce").astype("Int64")
        if "ID Nr." in df.columns and "ID" not in df.columns:
            df = df.rename(columns={"ID Nr.": "ID"})
        return df

    @st.cache_data(ttl=0)
    def load_courses_sheets():
        """Load sheets for: Credit granted courses / Non-credit granted courses (name tolerant).
        Estas hojas no existen en el reparto actual de archivos (BD_cartelera.xlsx
        tiene cartelera/programas/cursos/qualifications) — se deja la búsqueda
        tolerante por si se agregan más adelante; si no aparecen, retorna vacío."""
        xls = pd.ExcelFile(io.BytesIO(_download_drive_file_bytes(CARTELERA_FILE_ID)))
        sheets = xls.sheet_names

        def pick_sheet(candidates: List[str]) -> Optional[str]:
            lowmap = {s.lower(): s for s in sheets}
            for cand in candidates:
                if cand.lower() in lowmap: return lowmap[cand.lower()]
            for s in sheets:
                s_low = s.lower()
                if any(cand.lower() in s_low for cand in candidates):
                    return s
            return None

        credit_candidates = [
            "Creditd granted courses", "Credited granted courses", "Credit granted courses",
            "Credit granted course", "Credit granted", "Credit courses", "Creditd courses"
        ]
        noncredit_candidates = [
            "Non-credit granted courses", "Non credit granted courses", "Non-credit courses",
            "Non credit courses", "Noncredit granted courses"
        ]

        sh_credit = pick_sheet(credit_candidates)
        sh_noncr  = pick_sheet(noncredit_candidates)

        df_credit = pd.read_excel(xls, sheet_name=sh_credit) if sh_credit else pd.DataFrame()
        df_noncr  = pd.read_excel(xls, sheet_name=sh_noncr)  if sh_noncr  else pd.DataFrame()
        if not df_credit.empty: df_credit.columns = df_credit.columns.str.strip()
        if not df_noncr.empty:  df_noncr.columns  = df_noncr.columns.str.strip()
        return df_credit, df_noncr, sh_credit, sh_noncr

    # Header
    _render_header("Full-time Faculty Activities", "Questionnaire-based engagement summary 2020–2025")

    df_full = load_fulltime()
    df_q    = load_questionnaire()
    df_credit_sheet, df_noncredit_sheet, credit_sheet_name, noncredit_sheet_name = load_courses_sheets()

    # ================= SIDEBAR: NAVIGATION (selector + Open) =================

    # CONSTANTS
    TOT_PROFESSORS = 64           # denominator for % (donuts)
    MINT      = "#56D6C9"          # mint for "YES"
    MINT_DARK = "#1FA89B"          # darker mint (center text)
    GREY      = "#C7C7C7"          # grey for "NO"
    DONUT_H   = 160                # height of each donut
    # YEARS (fixed 2020–2025)
    YEARS = [2020, 2021, 2022, 2023, 2024, 2025]
    def _norm(s: pd.Series) -> pd.Series:
        return s.astype(str).str.strip().str.lower()

    def _year_filter(df: pd.DataFrame, year: int) -> Tuple[pd.DataFrame, Optional[str]]:
        ycol = resolve_column(df, "Year")
        if not ycol:
            return pd.DataFrame(), None
        d = df[df[ycol].astype("Int64") == year]
        return d, ycol

    def ft_second_sem_count(full_df: pd.DataFrame, year: int) -> Optional[int]:
        """Full-time total for the 2nd term; fallback to the last term of that year."""
        y = str(year)
        period = f"{y}-20"
        dfy = full_df[full_df["Periodo"] == period]
        if dfy.empty:
            by_year = full_df[full_df["Periodo"].str.startswith(y + "-")]
            if by_year.empty:
                return None
            last_p = sorted(by_year["Periodo"].unique())[-1]
            dfy = full_df[full_df["Periodo"] == last_p]
        idcol = "ID" if "ID" in dfy.columns else ("ID Nr." if "ID Nr." in dfy.columns else None)
        return int(dfy[idcol].nunique()) if idcol else int(len(dfy))

    def count_yes(df: pd.DataFrame, col: str, year: int) -> Optional[int]:
        d, ycol = _year_filter(df, year)
        if ycol is None or d.empty:
            return None
        c = resolve_column(df, col)
        if not c:
            return None
        return int(_norm(d[c]).eq("yes").sum())

    def count_contains(df: pd.DataFrame, col: str, patt: str, year: int, unique_by_id: bool = True) -> Optional[int]:
        d, ycol = _year_filter(df, year)
        if ycol is None or d.empty:
            return None
        c = resolve_column(df, col)
        if not c:
            return None
        mask = _norm(d[c]).str.contains(patt, regex=False, na=False)
        sub = d[mask]
        if sub.empty:
            return 0
        if unique_by_id and "ID" in sub.columns:
            return int(sub["ID"].nunique())
        return int(len(sub))

    # MANUAL OVERRIDES (table + donuts <=2024)
    # Keys map to row indicators for consistency.
    KEY1 = "total_ft"
    KEY2 = "postdoc"
    KEY3 = "editorial_boards"
    KEY4 = "reviewers"
    KEY5 = "boards_directors"
    KEY6 = "teaching_credit_abroad"
    KEY7 = "teaching_nondegree_abroad"
    KEY8 = "admin_positions"
    KEY9 = "execed"

    MANUAL_OVERRIDE = {
        2022: { KEY3: 12, KEY4: 37, KEY5: 15, KEY6: 13, KEY7: 8 },
        2024: { KEY2: 5,  KEY3: 24, KEY4: 15, KEY5: 13, KEY6: 12, KEY7: 5 }
    }

    # From 2025 on, only DB-derived values (no override) for these keys:
    AUTO_KEYS_2025_ON = {KEY1, KEY2, KEY3, KEY4, KEY5, KEY6, KEY7}

    def apply_override(year: int, key: str, computed: Optional[int]) -> Optional[int]:
        if year >= 2025 and key in AUTO_KEYS_2025_ON:
            return computed
        v = MANUAL_OVERRIDE.get(year, {}).get(key, None)
        return v if v is not None else computed

    # Persistent manual entries for rows 8 & 9 (admin / execed)
    if "manual_admin" not in st.session_state:
        st.session_state.manual_admin = {y: None for y in YEARS}
        st.session_state.manual_admin[2024] = 8
        st.session_state.manual_admin[2025] = 17
    if "manual_execed" not in st.session_state:
        st.session_state.manual_execed = {y: None for y in YEARS}
        st.session_state.manual_execed[2025] = 46

    with st.sidebar:
        y_edit = st.selectbox("Year to edit (rows 8 and 9)", YEARS, index=len(YEARS)-1)
        edit_admin = st.checkbox("Edit Administrative Positions", value=(st.session_state.manual_admin[y_edit] is not None))
        if edit_admin:
            st.session_state.manual_admin[y_edit] = st.number_input(
                "Administrative Positions value", min_value=0, step=1,
                value=st.session_state.manual_admin[y_edit] if st.session_state.manual_admin[y_edit] is not None else 0
            )
        edit_exec = st.checkbox("Edit ExecEd", value=(st.session_state.manual_execed[y_edit] is not None))
        if edit_exec:
            st.session_state.manual_execed[y_edit] = st.number_input(
                "ExecEd value", min_value=0, step=1,
                value=st.session_state.manual_execed[y_edit] if st.session_state.manual_execed[y_edit] is not None else 0
            )

    # TABLE BUILD (2020–2025)
    ROWS = [
        "Total Full-time Faculty",
        "Number of Full-time Faculty with Postdoc",
        "Number of Faculty in Editorial Boards",
        "Number of Reviewers in Academic Journals",
        "Number of Faculty in Boards of Directors",
        "Number Teaching Credit Granting Courses Abroad",
        "Number Teaching Non-degree Credit Granting Courses Abroad",
        "Number of Faculty in Administrative Positions",
        "Number of Full-time Faculty Teaching in Executive Education"
    ]

    def compute_metrics_for_year(y: int) -> Dict[str, Optional[int]]:
        """Single source of truth for both the table and the donuts."""
        c1 = apply_override(y, KEY1, ft_second_sem_count(df_full, y))
        c2 = apply_override(y, KEY2, count_yes(df_q, "Q6", y))
        c3a = count_contains(df_q, "Q36", "editorial", y, True)
        c3b = count_contains(df_q, "Q36", "editioral", y, True)
        c3  = None if (c3a is None and c3b is None) else (int(c3a or 0) + int(c3b or 0))
        c3  = apply_override(y, KEY3, c3)
        c4  = apply_override(y, KEY4, count_contains(df_q, "Q36", "journal", y, True))
        c5  = apply_override(y, KEY5, count_contains(df_q, "Q5",  "director", y, True))
        c6  = apply_override(y, KEY6, count_yes(df_q, "Q16", y))
        c7  = apply_override(y, KEY7, count_yes(df_q, "Q18", y))
        c8  = st.session_state.manual_admin.get(y)
        c9  = st.session_state.manual_execed.get(y)
        return {
            KEY1: c1, KEY2: c2, KEY3: c3, KEY4: c4, KEY5: c5, KEY6: c6, KEY7: c7, KEY8: c8, KEY9: c9
        }

    data = { "Indicator": ROWS }
    for y in YEARS:
        m = compute_metrics_for_year(y)
        def show(v):
            return "" if (v is None or (isinstance(v, (int, float, np.integer, np.floating)) and float(v) == 0.0)) else int(v)
        data[str(y)] = [
            show(m[KEY1]), show(m[KEY2]), show(m[KEY3]), show(m[KEY4]), show(m[KEY5]),
            show(m[KEY6]), show(m[KEY7]), show(m[KEY8]), show(m[KEY9])
        ]
    table = pd.DataFrame(data)

    # LAYOUT: LEFT (TABLE) / RIGHT (DONUTS)
    colL, colR = st.columns([7,5], gap="large")

    # LEFT: TABLE
    with colL:
        st.subheader("Summary 2020–2025")

        def _style(df_):
            styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
            year_cols = [c for c in df_.columns if c.isdigit()]
            last_year = str(max(int(c) for c in year_cols)) if year_cols else None
            if len(df_.index) > 0 and year_cols:
                styles.loc[df_.index[0], year_cols] += 'font-weight:800;'
            if last_year is not None and len(df_.index) > 0:
                styles.loc[df_.index[0], last_year] += 'background-color:#E8FAF7; color:#21877D; font-weight:800;'
            return styles

        styled = table.style.apply(_style, axis=None).hide(axis="index")
        st.dataframe(styled, use_container_width=True, height=48 + 33*(len(table)+1), hide_index=True)
        _download_link(
            "Descargar Excel Summary 2020–2025",
            table,  # la tabla base (no el Styler)
            f"summary_{YEARS[0]}_{YEARS[-1]}.xlsx"
        )

    # RIGHT: RESPONSE KPI + DONUTS
    with colR:
        # Year nav
        if "year_idx" not in st.session_state:
            st.session_state.year_idx = len(YEARS) - 1
        cL, cC, cR = st.columns([1, 3, 1])
        with cL:
            if st.button("◀", key="yr_prev"):
                if st.session_state.year_idx > 0:
                    st.session_state.year_idx -= 1
        with cR:
            if st.button("▶", key="yr_next"):
                if st.session_state.year_idx < len(YEARS) - 1:
                    st.session_state.year_idx += 1
        y_sel = YEARS[st.session_state.year_idx]
        with cC:
            st.markdown(f"<div style='text-align:center;font-weight:800'>Year: {y_sel}</div>", unsafe_allow_html=True)

        # Response rate (simple % of respondents over total professors)
        d_y, ycol = _year_filter(df_q, y_sel)
        n_resp = int(d_y["ID"].nunique()) if (ycol and "ID" in d_y.columns) else int(len(d_y))
        rate = (n_resp / TOT_PROFESSORS * 100.0) if TOT_PROFESSORS else 0.0
        st.markdown(f"### Response rate: {rate:.1f}%")

        # Shared legend (YES/NO)
        st.markdown(
            f"<div class='legend-center'>"
            f"<div class='legend-item'><span class='legend-swatch' style='background:{MINT}'></span> YES</div>"
            f"<div class='legend-item'><span class='legend-swatch' style='background:{GREY}'></span> NO</div>"
            f"</div>",
            unsafe_allow_html=True
        )

        # Donut helper
        def donut_fig(title: str, yes_count: Optional[int], total: int = TOT_PROFESSORS, height: int = DONUT_H):
            yv = int(yes_count or 0)
            nv = max(total - yv, 0)
            dfp = pd.DataFrame({"Status": ["YES", "NO"], "Value": [yv, nv]})
            fig = px.pie(
                dfp, names="Status", values="Value", hole=0.65,
                title=title, color="Status",
                color_discrete_map={"YES": MINT, "NO": GREY}
            )
            fig.update_traces(textinfo="none", hovertemplate="%{label}: %{value} of " + str(total))
            pct = (yv / total * 100.0) if total else 0.0
            fig.add_annotation(x=0.5, y=0.5, text=f"{pct:.0f}%", showarrow=False,
                               font=dict(size=18, color=MINT_DARK))
            fig.update_layout(margin=dict(l=6,r=6,t=26,b=6), showlegend=False,
                              height=height, title_font_size=12)
            return fig

        # >>>>>> Donut values come from the SAME source as the table <<<<<<
        metrics_sel = compute_metrics_for_year(y_sel)
        postdoc   = metrics_sel[KEY2] or 0
        editorial = metrics_sel[KEY3] or 0
        reviewers = metrics_sel[KEY4] or 0
        boards    = metrics_sel[KEY5] or 0
        admin     = metrics_sel[KEY8] or 0
        execed    = metrics_sel[KEY9] or 0

        c1, c2, c3 = st.columns(3)
        with c1:
            st.plotly_chart(donut_fig("Faculty with Postdoc", postdoc), use_container_width=True)
            st.plotly_chart(donut_fig("Reviewers in Journals", reviewers), use_container_width=True)
        with c2:
            st.plotly_chart(donut_fig("Editorial Boards", editorial), use_container_width=True)
            st.plotly_chart(donut_fig("Boards of Directors", boards), use_container_width=True)
        with c3:
            st.plotly_chart(donut_fig("Faculty in Administrative Positions", admin), use_container_width=True)
            st.plotly_chart(donut_fig("Faculty Teaching in ExecEd",       execed), use_container_width=True)

    # COURSE TABLES (from dedicated sheets)
    st.markdown("---")
    st.subheader("Courses taught by Full-time Faculty Abroad")

    # Reuse selected year
    y_sel = YEARS[st.session_state.year_idx]

    def _extract_year_from_text(v) -> Optional[int]:
        if pd.isna(v): return None
        if isinstance(v, (int, np.integer)): return int(v)
        if isinstance(v, (float, np.floating)) and np.isfinite(v): return int(v)
        m = re.search(r'(19|20)\d{2}', str(v))
        return int(m.group(0)) if m else None

    def _find_fullname(df: pd.DataFrame) -> Optional[str]:
        for cand in ["Full name", "Full Name", "Fullname", "Name", "Faculty", "Professor", "Profesor", "Nombre"]:
            col = resolve_column(df, cand)
            if col: return col
        f = resolve_column(df, "First Name")
        l = resolve_column(df, "Last Name")
        if f and l:
            df["__FULLNAME__"] = df[f].fillna("").astype(str).str.strip() + " " + df[l].fillna("").astype(str).str.strip()
            return "__FULLNAME__"
        return None

    def flatten_granted_courses_precise(df_src: pd.DataFrame, y_sel: int) -> pd.DataFrame:
        """Return rows (Professor, Course, University, Year delivered) filtered by y_sel, using 'Please specify - ...' columns."""
        if df_src.empty:
            return pd.DataFrame(columns=["Professor", "Course", "University", "Year delivered"])
        df = df_src.copy()
        df.columns = df.columns.str.strip()

        name_col = _find_fullname(df)
        year_col = resolve_column(df, "Year")  # not globally filtering by this; prefer Year:N when present

        low = {c: c.lower().strip() for c in df.columns}
        rx_course = re.compile(r'^please\s*specify\s*-\s*course\s*(?:name|title)\s*(?::\s*|\s+)?(\d+)?\s*:?\s*$', re.I)
        rx_uni    = re.compile(r'^please\s*specify\s*-\s*university\s*(?::\s*|\s+)?(\d+)?\s*:?\s*$', re.I)
        rx_year_i = re.compile(r'^please\s*specify\s*-\s*year\s*(?::\s*|\s+)?(\d+)?\s*:?\s*$', re.I)

        course_cols, uni_cols, year_item_cols = {}, {}, {}
        next_course_idx = 1
        next_uni_idx    = 1

        for col, lw in low.items():
            mc = rx_course.match(lw)
            if mc:
                n = int(mc.group(1)) if mc.group(1) else next_course_idx
                course_cols[n] = col
                if not mc.group(1): next_course_idx += 1
                continue
            mu = rx_uni.match(lw)
            if mu:
                n = int(mu.group(1)) if mu.group(1) else next_uni_idx
                uni_cols[n] = col
                if not mu.group(1): next_uni_idx += 1
                continue
            my = rx_year_i.match(lw)
            if my:
                n = int(my.group(1)) if my.group(1) else 1
                year_item_cols[n] = col
                continue

        idxs = sorted(set(course_cols.keys()) | set(uni_cols.keys()) | set(year_item_cols.keys()))
        if not idxs:
            return pd.DataFrame(columns=["Professor", "Course", "University", "Year delivered"])

        rows = []
        for _, r in df.iterrows():
            prof = str(r.get(name_col, "")).strip() if (name_col and name_col in r.index and pd.notna(r.get(name_col))) else ""
            year_numeric = None
            if year_col and year_col in r.index and pd.notna(r.get(year_col)):
                yv = pd.to_numeric(r.get(year_col), errors="coerce")
                if pd.notna(yv): year_numeric = int(yv)

            for n in idxs:
                ccol = course_cols.get(n); ucol = uni_cols.get(n); ycol = year_item_cols.get(n) or year_item_cols.get(1)
                course = str(r.get(ccol, "")).strip() if (ccol and ccol in r.index and pd.notna(r.get(ccol))) else ""
                uni    = str(r.get(ucol, "")).strip() if (ucol and ucol in r.index and pd.notna(r.get(ucol))) else ""

                year_text_display = ""
                if ycol and ycol in r.index and pd.notna(r.get(ycol)):
                    year_text_display = str(r.get(ycol)).strip()

                year_text_numeric = _extract_year_from_text(r.get(ycol)) if (ycol and ycol in r.index) else None
                year_val = year_numeric if year_numeric is not None else year_text_numeric
                if year_val is None or int(year_val) != int(y_sel):
                    continue
                if not (course or uni):
                    continue

                rows.append({
                    "Professor": prof,
                    "Course": course,
                    "University": uni,
                    "Year delivered": year_text_display
                })
        return pd.DataFrame(rows).drop_duplicates().reset_index(drop=True)

    # Two tables side by side
    cTa, cTb = st.columns(2)

    def _counts(df: pd.DataFrame) -> tuple[int, int]:
        if df.empty:
            return 0, 0
        profs = df["Professor"].astype(str).str.strip()
        profs = profs[profs != ""]
        return int(profs.nunique()), int(len(df))

    with cTa:
        title_credit = credit_sheet_name if credit_sheet_name else "Credit granted courses"
        if df_credit_sheet.empty:
            st.markdown(f"#### {title_credit} — 0 Faculty members taught courses abroad")
            st.info(f"Sheet '{title_credit}' was not found.")
        else:
            df_credit_flat = flatten_granted_courses_precise(df_credit_sheet, y_sel)
            n_prof, n_courses = _counts(df_credit_flat)
            st.markdown(f"#### {n_prof} Faculty members taught {n_courses} {title_credit} abroad")
            if df_credit_flat.empty:
                st.info(f"No records for {y_sel}.")
            else:
                st.dataframe(df_credit_flat, use_container_width=True, hide_index=True)

                # 1) hoja completa (sin filtrar) que alimenta la tabla:
                safe_name_credit = (credit_sheet_name or "Credit granted courses").lower().replace(" ", "_")
                _download_link("Descargar Excel Credit granted",
                            df_credit_sheet,
                            f"{safe_name_credit}_full.xlsx")

    with cTb:
        title_noncr = noncredit_sheet_name if noncredit_sheet_name else "Non-credit granted courses"
        if df_noncredit_sheet.empty:
            st.markdown(f"#### {title_noncr} — 0 Faculty members taught courses abroad")
            st.info(f"Sheet '{title_noncr}' was not found.")
        else:
            df_noncr_flat = flatten_granted_courses_precise(df_noncredit_sheet, y_sel)
            n_prof_nc, n_courses_nc = _counts(df_noncr_flat)
            st.markdown(f"#### {n_prof_nc} Faculty members taught {n_courses_nc} {title_noncr} abroad")
            if df_noncr_flat.empty:
                st.info(f"No records for {y_sel}.")
            else:
                st.dataframe(df_noncr_flat, use_container_width=True, hide_index=True)

                # hoja completa (sin filtrar) que alimenta la tabla:
                safe_name_noncr = (noncredit_sheet_name or "Non-credit granted courses").lower().replace(" ", "_")
                _download_link("Descargar Excel Non-credit granted",
                            df_noncredit_sheet,

                            f"{safe_name_noncr}_full.xlsx")


# 10) PÁGINA 6 — Faculty Qualifications
def page_qualifications():
    _render_header("Full-time Faculty Qualifications", "P/S and qualification type analysis with sensitivity mode")

    df_planta = qual_load_planta()
    df_fd = qual_load_faculty_distribution()
    df_car = qual_load_cartelera()

    # ------------------------ CONSTANTS & HELPERS ------------------------
    MINT = "#1FA89B"
    SUPPORTING = "#7FD3FF"
    TOTAL_SERIES_COLOR = "#D09E33"

    def _resolve(df: pd.DataFrame, target: str):
        t = target.strip().casefold()
        for c in df.columns:
            if c.strip().casefold() == t:
                return c
        return None

    def _norm_str(s: pd.Series) -> pd.Series:
        return s.astype(str).str.strip().str.lower()

    def normalize_ps(val: str) -> str:
        v = str(val).strip().lower()
        if v in {"p","participating","participante","participating faculty"}:
            return "P"
        if v in {"s","supporting","soporte","supporting faculty"}:
            return "S"
        return ""

    def normalize_tipo(val: str) -> str:
        v = str(val).strip().lower()
        if v in {"sa","scholarly academics","scholarly academic"}:
            return "SA"
        if v in {"pa","practice academics","practice academic"}:
            return "PA"
        if v in {"sp","scholarly practitioners","scholarly practitioner"}:
            return "SP"
        if v in {"ip","instructional practitioners","instructional practitioner"}:
            return "IP"
        if v in {"o","other","others","otro","otros"}:
            return "OTHER"
        m = re.search(r"\b(sa|pa|sp|ip|o|other)\b", v)
        if m:
            code = m.group(1).upper()
            return "OTHER" if code in {"O","OTHER"} else code
        return "OTHER"

    def _get_any(df: pd.DataFrame, *cands) -> str | None:
        for c in cands:
            got = _resolve(df, c)
            if got:
                return c
        return None

    def extract_year_from_period(p: str) -> int | None:
        if p is None:
            return None
        m = re.search(r"(19|20)\d{2}", str(p))
        return int(m.group(0)) if m else None

    def period_suffix(p: str) -> str | None:
        m = re.search(r"(?:19|20)\d{2}[-_/ ]?(\d+)", str(p))
        return m.group(1) if m else None

    def is_regular_period(p) -> bool:
        s = str(p).strip().lower()
        if "inter" in s:
            return False
        suf = period_suffix(s)
        return (suf in {"10", "20"}) or (suf is None)

    def list_periods_semestral():
        sem_col = _get_any(df_car, "Semestre", "Periodo", "Periodo Académico", "Periodo academico")
        vals = []
        if sem_col:
            vals = df_car[sem_col].dropna().astype(str).str.strip().tolist()
        regs = [v for v in vals if is_regular_period(v) and period_suffix(v) in {"10","20"}]
        def sort_key(p):
            y = extract_year_from_period(p) or -1
            suf = int(period_suffix(p) or 0)
            return (y, suf)
        return sorted(sorted(set(regs)), key=sort_key, reverse=True)

    def list_years_from_sem():
        sem_col = _get_any(df_car, "Semestre", "Periodo", "Periodo Académico", "Periodo academico")
        years = set()
        if sem_col:
            for s in df_car[sem_col].dropna().astype(str):
                y = extract_year_from_period(s)
                if y:
                    years.add(y)
        ycol_fd = _get_any(df_fd, "Year", "Año")
        if ycol_fd:
            for y in pd.to_numeric(df_fd[ycol_fd], errors="coerce").dropna().astype(int):
                years.add(int(y))
        return sorted(years, reverse=True)

    def years_with_inter():
        sem_col = _get_any(df_car, "Semestre", "Periodo", "Periodo Académico", "Periodo academico")
        inter = set()
        if sem_col:
            for s in df_car[sem_col].dropna().astype(str):
                if "inter" in s.lower():
                    y = extract_year_from_period(s)
                    if y:
                        inter.add(y)
        return sorted(inter, reverse=True)

    def _slugify(s: str) -> str:
        return re.sub(r'[^A-Za-z0-9]+', '_', str(s)).strip('_')

    # —— utilidades de descarga ——
    def _sanitize_for_export(df: pd.DataFrame) -> pd.DataFrame:
        return df[[c for c in df.columns if not str(c).startswith("_")]].copy()

    def _download_xlsx_button(df: pd.DataFrame, fname: str, key: str, label: str = "Download Excel"):
        safe = _sanitize_for_export(df)
        clean = re.sub(r"[^\w\sÁÉÍÓÚÜÑáéíóúüñ().%/-]+", "", label).strip()
        st.download_button(
            clean,
            data=_xlsx_bytes(safe),
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=key,
            use_container_width=False
        )

    # ================== SENSITIVITY HELPERS ==================
    def build_member_list_for_view(df_period: pd.DataFrame, view_mode: str, col_areaCourse, col_field, program_col) -> list[str]:
        if view_mode == "By Academic Area" and col_areaCourse:
            items = sorted(df_period[col_areaCourse].astype(str).str.strip().dropna().unique().tolist())
        elif view_mode == "By Field" and col_field:
            items = sorted(df_period[col_field].astype(str).str.strip().dropna().unique().tolist())
        elif view_mode == "By Program" and program_col:
            items = sorted(df_period[program_col].astype(str).str.strip().dropna().unique().tolist())
        else:
            items = []
        return ["All"] + items

    def apply_ops_to_aggs(agg_ps: pd.DataFrame, agg_tipo: pd.DataFrame, ops: list, member_all_label="All") -> tuple[pd.DataFrame, pd.DataFrame]:
        mod_ps = agg_ps.copy()
        mod_tipo = agg_tipo.copy()
        for op in ops or []:
            scope = op.get("scope")
            cat = op.get("cat")
            member = op.get("member", member_all_label)
            delta = float(op.get("credits", 0.0)) * int(op.get("count", 0))
            if delta == 0:
                continue
            if scope == "PS":
                if cat not in ["P", "S"]:
                    continue
                if cat not in mod_ps.columns:
                    mod_ps[cat] = 0.0
                if member == member_all_label:
                    mod_ps[cat] = (mod_ps[cat] + delta).clip(lower=0.0)
                else:
                    if member in mod_ps.index:
                        mod_ps.at[member, cat] = max(0.0, float(mod_ps.at[member, cat]) + delta)
            elif scope == "QUAL":
                cats = ["SA","SP","IP","PA","OTHER"]
                if cat not in cats:
                    continue
                if cat not in mod_tipo.columns:
                    mod_tipo[cat] = 0.0
                if member == member_all_label:
                    mod_tipo[cat] = (mod_tipo[cat] + delta).clip(lower=0.0)
                else:
                    if member in mod_tipo.index:
                        mod_tipo.at[member, cat] = max(0.0, float(mod_tipo.at[member, cat]) + delta)
        return mod_ps, mod_tipo

    # ===== Cálculo de “profesores necesarios” (3 créditos) por fila =====
    def _needed_for_pctP(p: float, s: float, target_pct: float, credits_each: float = 3.0) -> int:
        # (p + c*n)/(p + s + c*n) >= t  ->  n >= (t*s - (1-t)*p) / (c*(1-t))
        t = target_pct / 100.0
        denom = credits_each * (1 - t)
        if denom <= 0:
            return 0
        rhs = (t * s - (1 - t) * p) / denom
        return max(0, math.ceil(rhs))

    def _needed_for_pctSA(sa: float, rest: float, target_pct: float, credits_each: float = 3.0) -> int:
        # (sa + c*n)/(sa + rest + c*n) >= t -> n >= (t*rest - (1-t)*sa) / (c*(1-t))
        t = target_pct / 100.0
        denom = credits_each * (1 - t)
        if denom <= 0:
            return 0
        rhs = (t * rest - (1 - t) * sa) / denom
        return max(0, math.ceil(rhs))

    def _needed_for_other_leq10(other: float, rest: float, credits_each: float = 3.0) -> int:
        # other/(other + rest + c*n) <= 0.10  ->  n >= (0.9*other - 0.1*rest) / (0.3*c) = (9*other - rest)/(3*c)
        num = 9*other - rest
        denom = 3 * credits_each
        if denom <= 0:
            return 0
        rhs = num / denom
        return max(0, math.ceil(rhs))

    def _objective_targets(obj: str) -> tuple[str, float, float]:
        # devuelve etiqueta y targets por scope (by_area, overall)
        if obj == "%P":   return ("%P", 60.0, 75.0)
        if obj == "%SA":  return ("%SA", 40.0, 40.0)
        return ("%OTHER", 10.0, 10.0)

    # ====== NUEVOS helpers para Overall/Impacto y secundarios ======
    def _needed_for_overall_if_only_this_area_changes(obj: str, totals: dict[str, float], area_vals: dict[str, float], target_overall: float, credits_each: float = 3.0) -> int | None:
        eps = 1e-9
        t = target_overall / 100.0
        Ptot = totals.get("P",0.0);  Stot = totals.get("S",0.0)
        SA   = totals.get("SA",0.0); PA  = totals.get("PA",0.0)
        SP   = totals.get("SP",0.0); IP  = totals.get("IP",0.0)
        OT   = totals.get("OTHER",0.0)
        TQ   = SA + PA + SP + IP + OT
        if obj == "%P":
            den = Ptot + Stot
            if den <= eps: return 0
            rhs = (t*den - Ptot) / (credits_each*(1 - t))
            return max(0, math.ceil(rhs))
        if obj == "%SA":
            if TQ <= eps: return 0
            rhs = (t*TQ - SA) / (credits_each*(1 - t))
            return max(0, math.ceil(rhs))
        if TQ <= eps: return 0
        need_credits = (OT - 0.10*TQ) / 0.90
        need_n = 0 if need_credits <= 0 else math.ceil(need_credits / credits_each)
        OT_a = area_vals.get("OTHER", 0.0)
        max_remove_n = math.floor(OT_a / credits_each)
        return max(0, need_n) if need_n <= max_remove_n else None

    def _impact_pp_area(obj: str, area_vals: dict[str,float], credits_each: float = 3.0) -> tuple[float,float]:
        eps = 1e-9
        P = area_vals.get("P",0.0); S = area_vals.get("S",0.0)
        SA = area_vals.get("SA",0.0); PA = area_vals.get("PA",0.0)
        SP = area_vals.get("SP",0.0); IP = area_vals.get("IP",0.0)
        OT = area_vals.get("OTHER",0.0)
        denPS = P + S
        denQ  = SA + PA + SP + IP + OT
        if obj == "%P":
            if denPS <= eps: return (0.0, 0.0)
            up   = ((P + credits_each) / (denPS + credits_each) - (P / denPS)) * 100.0
            down = ((max(0.0, P - credits_each)) / max(eps, denPS - credits_each) - (P / denPS)) * 100.0 if denPS > credits_each else 0.0
            return (round(up,2), round(down,2))
        if obj == "%SA":
            if denQ <= eps: return (0.0, 0.0)
            up   = ((SA + credits_each) / (denQ + credits_each) - (SA / denQ)) * 100.0
            down = ((max(0.0, SA - credits_each)) / max(eps, denQ - credits_each) - (SA / denQ)) * 100.0 if denQ > credits_each else 0.0
            return (round(up,2), round(down,2))
        if denQ <= eps: return (0.0, 0.0)
        up   = ((OT + credits_each) / (denQ + credits_each) - (OT / denQ)) * 100.0
        down = ((max(0.0, OT - credits_each)) / max(eps, denQ - credits_each) - (OT / denQ)) * 100.0 if denQ > credits_each else 0.0
        return (round(up,2), round(down,2))

    def _impact_pp_overall_if_area_changes(obj: str, totals: dict[str,float], credits_each: float = 3.0) -> tuple[float,float]:
        eps = 1e-9
        P = totals.get("P",0.0); S = totals.get("S",0.0)
        SA = totals.get("SA",0.0); PA = totals.get("PA",0.0)
        SP = totals.get("SP",0.0); IP = totals.get("IP",0.0)
        OT = totals.get("OTHER",0.0)
        denPS = P + S
        denQ  = SA + PA + SP + IP + OT
        if obj == "%P":
            if denPS <= eps: return (0.0, 0.0)
            up   = ((P + credits_each) / (denPS + credits_each) - (P / denPS)) * 100.0
            down = ((max(0.0, P - credits_each)) / max(eps, denPS - credits_each) - (P / denPS)) * 100.0 if denPS > credits_each else 0.0
            return (round(up,2), round(down,2))
        if obj == "%SA":
            if denQ <= eps: return (0.0, 0.0)
            up   = ((SA + credits_each) / (denQ + credits_each) - (SA / denQ)) * 100.0
            down = ((max(0.0, SA - credits_each)) / max(eps, denQ - credits_each) - (SA / denQ)) * 100.0 if denQ > credits_each else 0.0
            return (round(up,2), round(down,2))
        if denQ <= eps: return (0.0, 0.0)
        up   = ((OT + credits_each) / (denQ + credits_each) - (OT / denQ)) * 100.0
        down = ((max(0.0, OT - credits_each)) / max(eps, denQ - credits_each) - (OT / denQ)) * 100.0 if denQ > credits_each else 0.0
        return (round(up,2), round(down,2))

    # Secundarios para tablas "Needed"
    def _needed_S_less_for_pctP_area(p: float, s: float, target_pct: float, credits_each=3.0) -> int:
        # P/(P + S - c*n) >= t  ->  n >= (t*(P+S) - P)/(t*c)
        t = target_pct/100.0
        if t <= 0: return 0
        den = t*credits_each
        rhs = (t*(p+s) - p)/den
        return max(0, math.ceil(rhs))

    def _needed_S_less_for_pctP_overall(totals, area_vals, target_overall: float, credits_each=3.0) -> int | None:
        t = target_overall/100.0
        Ptot = totals.get("P",0.0); Stot = totals.get("S",0.0)
        if t <= 0: return 0
        need = (t*(Ptot+Stot) - Ptot) / (t*credits_each)
        need_n = 0 if need <= 0 else math.ceil(need)
        S_a = area_vals.get("S",0.0)
        max_remove = math.floor(S_a/credits_each)
        return need_n if need_n <= max_remove else None

    def _needed_OTHERS_less_for_SA_area(sa, rest, target_pct, credits_each=3.0) -> int:
        # SA/(SA + rest - c*n) >= t -> n >= (t*(SA+rest) - SA)/(t*c)
        t = target_pct/100.0
        if t <= 0: return 0
        rhs = (t*(sa+rest) - sa)/(t*credits_each)
        return max(0, math.ceil(rhs))

    def _needed_OTHERS_less_for_SA_overall(totals, area_vals, target_overall, credits_each=3.0) -> int | None:
        SA = totals.get("SA",0.0); PA=totals.get("PA",0.0); SP=totals.get("SP",0.0); IP=totals.get("IP",0.0); OT=totals.get("OTHER",0.0)
        TQ = SA+PA+SP+IP+OT; rest = TQ - SA
        t = target_overall/100.0
        if t <= 0: return 0
        need = (t*(SA+rest) - SA)/(t*credits_each)
        need_n = 0 if need <= 0 else math.ceil(need)
        rest_a = max(0.0, area_vals.get("PA",0.0)+area_vals.get("SP",0.0)+area_vals.get("IP",0.0)+area_vals.get("OTHER",0.0))
        max_remove = math.floor(rest_a/credits_each)
        return need_n if need_n <= max_remove else None

    def _needed_OTHERS_more_for_OTHER_area(other, rest, target_pct, credits_each=3.0) -> int:
        # OTHER/(OTHER + rest + c*n) <= t -> c*n >= (OTHER - t*(OTHER+rest))/t
        t = target_pct/100.0
        if t <= 0: return 0
        need_credits = (other - t*(other+rest))/t
        need = 0 if need_credits <= 0 else math.ceil(need_credits/credits_each)
        return max(0, need)

    def _needed_OTHERS_more_for_OTHER_overall(totals, target_overall, credits_each=3.0) -> int:
        OT = totals.get("OTHER",0.0); SA=totals.get("SA",0.0); PA=totals.get("PA",0.0); SP=totals.get("SP",0.0); IP=totals.get("IP",0.0)
        TQ = SA+PA+SP+IP+OT
        t = target_overall/100.0
        if t <= 0: return 0
        need_credits = (OT - t*TQ)/t
        return 0 if need_credits <= 0 else math.ceil(need_credits/credits_each)

    # ================== HISTORY (timeframe-aware) ==================
    def _period_sort_key(p: str) -> tuple[int,int]:
        y = extract_year_from_period(p) or -1
        suf = period_suffix(p)
        try:
            suf_i = int(suf) if suf is not None else 0
        except Exception:
            suf_i = 0
        return (y, suf_i)

    def build_time_axis_for_history(df_hist: pd.DataFrame):
        time_mode = st.session_state.get("time_mode", "Semestral")
        if "_SEM" not in df_hist.columns:
            sc = _get_any(df_hist, "Semestre","Periodo","Periodo Académico","Periodo academico")
            sem = df_hist[sc].astype(str).str.strip() if sc else pd.Series([], dtype=str)
        else:
            sem = df_hist["_SEM"].astype(str).str.strip()
        if time_mode == "Semestral":
            regs = sorted(
                {s for s in sem.dropna().unique() if period_suffix(s) in {"10","20"}},
                key=_period_sort_key
            )
            x_labels = regs
        elif time_mode == "Anual":
            years = sorted({extract_year_from_period(s) for s in sem if extract_year_from_period(s)}, key=int)
            x_labels = years
        else:  # Intersemestral
            inter = sorted(
                {f"{extract_year_from_period(s)} Intersemestral" for s in sem if "inter" in str(s).lower() and extract_year_from_period(s)},
                key=lambda x: int(str(x).split()[0])
            )
            x_labels = inter
        x_map = {lab: i for i, lab in enumerate(x_labels)}
        return "_SEM", x_labels, x_map

    def transform_for_time_mode_ps(df_ps: pd.DataFrame):
        time_mode = st.session_state.get("time_mode", "Semestral")
        base = df_ps.copy()
        base["_YEAR"] = base["_SEM"].map(extract_year_from_period)
        base["_INTER_LABEL"] = base["_SEM"].map(lambda s: f"{extract_year_from_period(s)} Intersemestral" if "inter" in str(s).lower() else None)
        if time_mode == "Semestral":
            return base
        if time_mode == "Anual":
            need_cols = [c for c in base.columns if c not in {"P_share"}]
            g = base[need_cols].groupby(["_YEAR"] + [c for c in base.columns if c.startswith("_") and c not in {"_SEM","_YEAR","_INTER_LABEL"}], dropna=False).sum(numeric_only=True).reset_index()
            if "P" in g and "S" in g:
                g["P_share"] = (g["P"] / (g["P"] + g["S"]).replace(0, pd.NA)) * 100
            return g.rename(columns={"_YEAR":"_SEM"})
        # Intersemestral
        base = base[~base["_INTER_LABEL"].isna()].copy()
        g = base.groupby(["_INTER_LABEL"] + [c for c in base.columns if c.startswith("_") and c not in {"_SEM","_YEAR","_INTER_LABEL"}], dropna=False).sum(numeric_only=True).reset_index()
        if "P" in g and "S" in g:
            g["P_share"] = (g["P"] / (g["P"] + g["S"]).replace(0, pd.NA)) * 100
        return g.rename(columns={"_INTER_LABEL":"_SEM"})

    def transform_for_time_mode_tipo(df_tipo: pd.DataFrame, share_col_name: str):
        time_mode = st.session_state.get("time_mode", "Semestral")
        base = df_tipo.copy()
        base["_YEAR"] = base["_SEM"].map(extract_year_from_period)
        base["_INTER_LABEL"] = base["_SEM"].map(lambda s: f"{extract_year_from_period(s)} Intersemestral" if "inter" in str(s).lower() else None)
        cats = ["SA","PA","SP","IP","OTHER"]
        if time_mode == "Semestral":
            return base
        if time_mode == "Anual":
            keys = ["_YEAR"] + [c for c in base.columns if c.startswith("_") and c not in {"_SEM","_YEAR","_INTER_LABEL"}]
            g = base.groupby(keys, dropna=False)[cats].sum().reset_index()
            den = (g[cats].sum(axis=1)).replace(0, pd.NA)
            if share_col_name == "SA_share":
                g["SA_share"] = (g["SA"] / den) * 100
            else:
                g["OTHER_share"] = (g["OTHER"] / den) * 100
            return g.rename(columns={"_YEAR":"_SEM"})
        # Intersemestral
        base = base[~base["_INTER_LABEL"].isna()].copy()
        keys = ["_INTER_LABEL"] + [c for c in base.columns if c.startswith("_") and c not in {"_SEM","_YEAR","_INTER_LABEL"}]
        g = base.groupby(keys, dropna=False)[cats].sum().reset_index()
        den = (g[cats].sum(axis=1)).replace(0, pd.NA)
        if share_col_name == "SA_share":
            g["SA_share"] = (g["SA"] / den) * 100
        else:
            g["OTHER_share"] = (g["OTHER"] / den) * 100
        return g.rename(columns={"_INTER_LABEL":"_SEM"})

    # === aplicar sensibilidad sobre series históricas SOLO en el período seleccionado ===
    def apply_sensitivity_to_history(
        agg_ps_tm: pd.DataFrame,
        agg_tipo_tm: pd.DataFrame,
        tot_ps_tm: pd.DataFrame,
        tot_tipo_tm: pd.DataFrame,
        level_name: str,
        sel_label_value,  # etiqueta exacta seleccionada en el eje X
        ops: list,
        member_all_label="All"
    ):
        if not ops or sel_label_value is None:
            return agg_ps_tm, agg_tipo_tm, tot_ps_tm, tot_tipo_tm

        ps = agg_ps_tm.copy()
        tq = agg_tipo_tm.copy()
        tps = tot_ps_tm.copy()
        ttq = tot_tipo_tm.copy()

        for k in ["P","S"]:
            if k not in ps.columns:  ps[k] = 0.0
            if k not in tps.columns: tps[k] = 0.0
        for k in ["SA","PA","SP","IP","OTHER"]:
            if k not in tq.columns:  tq[k] = 0.0
            if k not in ttq.columns: ttq[k] = 0.0

        mask_period_ps  = ps["_SEM"].eq(sel_label_value)
        mask_period_tq  = tq["_SEM"].eq(sel_label_value)
        mask_period_tps = tps["_SEM"].eq(sel_label_value)
        mask_period_ttq = ttq["_SEM"].eq(sel_label_value)

        for op in ops:
            scope = op.get("scope")
            cat   = op.get("cat")
            member= op.get("member", member_all_label)
            delta = float(op.get("credits", 0.0)) * int(op.get("count", 0))
            if delta == 0:
                continue

            if scope == "PS" and cat in ["P","S"]:
                if member == member_all_label:
                    ps.loc[mask_period_ps, cat]  = (ps.loc[mask_period_ps, cat].astype(float)  + delta).clip(lower=0.0)
                else:
                    m = mask_period_ps & ps[level_name].eq(member)
                    ps.loc[m, cat] = (ps.loc[m, cat].astype(float) + delta).clip(lower=0.0)
                tps.loc[mask_period_tps, cat] = (tps.loc[mask_period_tps, cat].astype(float) + delta).clip(lower=0.0)

            if scope == "QUAL" and cat in ["SA","PA","SP","IP","OTHER"]:
                if member == member_all_label:
                    tq.loc[mask_period_tq, cat]  = (tq.loc[mask_period_tq, cat].astype(float)  + delta).clip(lower=0.0)
                else:
                    m = mask_period_tq & tq[level_name].eq(member)
                    tq.loc[m, cat] = (tq.loc[m, cat].astype(float) + delta).clip(lower=0.0)
                ttq.loc[mask_period_ttq, cat] = (ttq.loc[mask_period_ttq, cat].astype(float) + delta).clip(lower=0.0)

        den_ps = (ps.loc[mask_period_ps, "P"].astype(float) + ps.loc[mask_period_ps, "S"].astype(float)).replace(0, pd.NA)
        ps.loc[mask_period_ps, "P_share"] = (ps.loc[mask_period_ps, "P"] / den_ps * 100).fillna(0.0)

        cats = ["SA","PA","SP","IP","OTHER"]
        den_q = tq.loc[mask_period_tq, cats].sum(axis=1).replace(0, pd.NA)
        if "SA_share" in tq.columns:
            tq.loc[mask_period_tq, "SA_share"]    = (tq.loc[mask_period_tq, "SA"]    / den_q * 100).fillna(0.0)
        if "OTHER_share" in tq.columns:
            tq.loc[mask_period_tq, "OTHER_share"] = (tq.loc[mask_period_tq, "OTHER"] / den_q * 100).fillna(0.0)

        den_tps = (tps.loc[mask_period_tps, "P"].astype(float) + tps.loc[mask_period_tps, "S"].astype(float)).replace(0, pd.NA)
        tps.loc[mask_period_tps, "P_share"] = (tps.loc[mask_period_tps, "P"] / den_tps * 100).fillna(0.0)

        den_ttq = ttq.loc[mask_period_ttq, cats].sum(axis=1).replace(0, pd.NA)
        if "SA_share" in ttq.columns:
            ttq.loc[mask_period_ttq, "SA_share"]    = (ttq.loc[mask_period_ttq, "SA"]    / den_ttq * 100).fillna(0.0)
        if "OTHER_share" in ttq.columns:
            ttq.loc[mask_period_ttq, "OTHER_share"] = (ttq.loc[mask_period_ttq, "OTHER"] / den_ttq * 100).fillna(0.0)

        return ps, tq, tps, ttq

    # === extracción segura de valores únicos (evita TypeError por columnas duplicadas o NaN mezclado con str) ===
    def _safe_unique_labels(df: pd.DataFrame, col: str) -> list[str]:
        if col not in df.columns:
            return []
        ser = df[col]
        if isinstance(ser, pd.DataFrame):  # columna duplicada -> pandas devuelve un DataFrame
            ser = ser.iloc[:, 0]
        return ser.dropna().astype(str).unique().tolist()

    # --------- Gráfica histórica ---------
    def draw_history(fig_title, level_name, level_values, metric_kind, total_series_builders, agg_ps_all, agg_tipo_all, x_labels, x_map, sel_x):
        palette = px.colors.qualitative.Safe + px.colors.qualitative.Bold + px.colors.qualitative.Pastel
        color_map = {a: palette[i % len(palette)] for i, a in enumerate(level_values)}
        st.markdown(f"<h4 style='margin:0 0 6px 0; font-weight:500;'>{fig_title}</h4>", unsafe_allow_html=True)
        sel_col, radio_col = st.columns([6,4])
        options = ["(All)", "(TOTAL)"] + level_values
        with sel_col:
            opt = st.selectbox("", options, index=0, key=f"{level_name}_filter", label_visibility="collapsed")
        with radio_col:
            metric_choice = st.radio("", ["%P", "%SA", "%OTHER"], index={"%P":0, "%SA":1, "%OTHER":2}[metric_kind], horizontal=True, key=f"metric_{level_name}", label_visibility="collapsed")

        fig = go.Figure()

        if metric_choice == "%P":
            thr = 75 if opt == "(TOTAL)" else 60
            if opt == "(All)":
                for a in level_values:
                    sub = agg_ps_all[(agg_ps_all[level_name] == a)].copy()
                    sub["x"] = sub["_SEM"].map(x_map)
                    sub = sub.sort_values("x")
                    if sub.empty: continue
                    fig.add_trace(go.Scatter(
                        x=sub["x"], y=sub["P_share"], mode="lines+markers", name=a,
                        marker=dict(size=6, color=color_map[a]), line=dict(width=2, color=color_map[a]),
                        hovertemplate=a + "<br>%{y:.1f}%<extra></extra>"
                    ))
            elif opt == "(TOTAL)":
                sub = total_series_builders["P"].copy()
                sub["x"] = sub["_SEM"].map(x_map)
                sub = sub.sort_values("x")
                fig.add_trace(go.Scatter(
                    x=sub["x"], y=sub["P_share"], mode="lines+markers", name="TOTAL",
                    marker=dict(size=6, color=TOTAL_SERIES_COLOR), line=dict(width=2, color=TOTAL_SERIES_COLOR),
                    hovertemplate="TOTAL<br>%{y:.1f}%<extra></extra>"
                ))
            else:
                sub = agg_ps_all[(agg_ps_all[level_name] == opt)].copy()
                sub["x"] = sub["_SEM"].map(x_map)
                sub = sub.sort_values("x")
                fig.add_trace(go.Scatter(
                    x=sub["x"], y=sub["P_share"], mode="lines+markers", name=opt,
                    marker=dict(size=6, color=MINT), line=dict(width=2, color=MINT),
                    hovertemplate=opt + "<br>%{y:.1f}%<extra></extra>"
                ))
            y_min, bad_high = 40, False

        elif metric_choice == "%SA":
            thr = 40
            share_col = "SA_share"
            if opt == "(All)":
                for a in level_values:
                    sub = agg_tipo_all[(agg_tipo_all[level_name] == a)].copy()
                    sub["x"] = sub["_SEM"].map(x_map)
                    sub = sub.sort_values("x")
                    if sub.empty: continue
                    fig.add_trace(go.Scatter(
                        x=sub["x"], y=sub[share_col], mode="lines+markers", name=a,
                        marker=dict(size=6, color=color_map[a]), line=dict(width=2, color=color_map[a]),
                        hovertemplate=a + "<br>%{y:.1f}%<extra></extra>"
                    ))
            elif opt == "(TOTAL)":
                sub = total_series_builders["SA"].copy()
                sub["x"] = sub["_SEM"].map(x_map)
                sub = sub.sort_values("x")
                fig.add_trace(go.Scatter(
                    x=sub["x"], y=sub[share_col], mode="lines+markers", name="TOTAL",
                    marker=dict(size=6, color=TOTAL_SERIES_COLOR), line=dict(width=2, color=TOTAL_SERIES_COLOR),
                    hovertemplate="TOTAL<br>%{y:.1f}%<extra></extra>"
                ))
            else:
                sub = agg_tipo_all[(agg_tipo_all[level_name] == opt)].copy()
                sub["x"] = sub["_SEM"].map(x_map)
                sub = sub.sort_values("x")
                fig.add_trace(go.Scatter(
                    x=sub["x"], y=sub[share_col], mode="lines+markers", name=opt,
                    marker=dict(size=6, color=MINT), line=dict(width=2, color=MINT),
                    hovertemplate=opt + "<br>%{y:.1f}%<extra></extra>"
                ))
            y_min, bad_high = 20, False

        else:  # "%OTHER"
            thr = 10
            share_col = "OTHER_share"
            if opt == "(All)":
                for a in level_values:
                    sub = agg_tipo_all[(agg_tipo_all[level_name] == a)].copy()
                    sub["x"] = sub["_SEM"].map(x_map)
                    sub = sub.sort_values("x")
                    if sub.empty: continue
                    fig.add_trace(go.Scatter(
                        x=sub["x"], y=sub[share_col], mode="lines+markers", name=a,
                        marker=dict(size=6, color=color_map[a]), line=dict(width=2, color=color_map[a]),
                        hovertemplate=a + "<br>%{y:.1f}%<extra></extra>"
                    ))
            elif opt == "(TOTAL)":
                sub = total_series_builders["OTHER"].copy()
                sub["x"] = sub["_SEM"].map(x_map)
                sub = sub.sort_values("x")
                fig.add_trace(go.Scatter(
                    x=sub["x"], y=sub[share_col], mode="lines+markers", name="TOTAL",
                    marker=dict(size=6, color=TOTAL_SERIES_COLOR), line=dict(width=2, color=TOTAL_SERIES_COLOR),
                    hovertemplate="TOTAL<br>%{y:.1f}%<extra></extra>"
                ))
            else:
                sub = agg_tipo_all[(agg_tipo_all[level_name] == opt)].copy()
                sub["x"] = sub["_SEM"].map(x_map)
                sub = sub.sort_values("x")
                fig.add_trace(go.Scatter(
                    x=sub["x"], y=sub[share_col], mode="lines+markers", name=opt,
                    marker=dict(size=6, color=MINT), line=dict(width=2, color=MINT),
                    hovertemplate=opt + "<br>%{y:.1f}%<extra></extra>"
                ))
            y_min, bad_high = 0, True
            y_max = 40

        # Zonas de referencia
        if bad_high:
            fig.update_layout(shapes=[dict(type="rect", xref="paper", yref="y", x0=0, x1=1, y0=thr, y1=100, fillcolor="#FDE2E2", opacity=0.35, layer="below", line_width=0)])
            fig.add_hline(y=thr, line_color="#F5A3A3", line_dash="dash")
        else:
            fig.update_layout(shapes=[dict(type="rect", xref="paper", yref="y", x0=0, x1=1, y0=0, y1=thr, fillcolor="#FDE2E2", opacity=0.35, layer="below", line_width=0)])
            fig.add_hline(y=thr, line_color="red", line_dash="dash")

        if sel_x is not None:
            fig.add_vrect(x0=sel_x-0.5, x1=sel_x+0.5, fillcolor="#E8FAF7", opacity=0.5, layer="below", line_width=0)

        tickvals = list(range(len(x_labels)))
        ticktext = [str(x) for x in x_labels]
        if metric_choice == "%OTHER":
            fig.update_layout(xaxis=dict(tickmode="array", tickvals=tickvals, ticktext=ticktext), yaxis=dict(range=[y_min, y_max]))
        else:
            fig.update_layout(xaxis=dict(tickmode="array", tickvals=tickvals, ticktext=ticktext), yaxis=dict(range=[y_min, 100]))
        fig.update_xaxes(title=None)
        fig.update_yaxes(title=None)
        st.plotly_chart(fig, use_container_width=True)

        # ===== Datos para descargar (lo visible) =====
        def _series_for(level_val: str, ycol: str):
            if ycol == "P_share":
                sub = agg_ps_all[(agg_ps_all[level_name] == level_val)]
            else:
                sub = agg_tipo_all[(agg_tipo_all[level_name] == level_val)]
            m = sub.set_index("_SEM")[ycol].to_dict()
            return [m.get(x, None) for x in x_labels]

        if metric_choice == "%P":
            ycol = "P_share"
            base_cols = {}
            if opt == "(All)":
                for a in level_values:
                    base_cols[a] = _series_for(a, ycol)
            elif opt == "(TOTAL)":
                sub = total_series_builders["P"].set_index("_SEM")["P_share"].to_dict()
                base_cols["TOTAL"] = [sub.get(x, None) for x in x_labels]
            else:
                base_cols[opt] = _series_for(opt, ycol)
        elif metric_choice == "%SA":
            ycol = "SA_share"
            base_cols = {}
            if opt == "(All)":
                for a in level_values:
                    base_cols[a] = _series_for(a, ycol)
            elif opt == "(TOTAL)":
                sub = total_series_builders["SA"].set_index("_SEM")[ycol].to_dict()
                base_cols["TOTAL"] = [sub.get(x, None) for x in x_labels]
            else:
                base_cols[opt] = _series_for(opt, ycol)
        else:
            ycol = "OTHER_share"
            base_cols = {}
            if opt == "(All)":
                for a in level_values:
                    base_cols[a] = _series_for(a, ycol)
            elif opt == "(TOTAL)":
                sub = total_series_builders["OTHER"].set_index("_SEM")[ycol].to_dict()
                base_cols["TOTAL"] = [sub.get(x, None) for x in x_labels]
            else:
                base_cols[opt] = _series_for(opt, ycol)

        export_df = pd.DataFrame({"Period": x_labels, **base_cols})
        fname = f"chart_{_slugify(fig_title)}_{_slugify(metric_choice)}_{_slugify(opt)}_{_slugify(st.session_state.get('sel_label','sel'))}.xlsx"
        _download_xlsx_button(export_df, fname, key=f"dl_hist_{_slugify(fig_title)}_{metric_choice}_{_slugify(opt)}_{_slugify(st.session_state.get('sel_label','sel'))}", label="⬇️ Datos de la gráfica (Excel)")

    # ============== NORMALIZACIÓN BÁSICA EN CARTELERA ==============
    col_sem = _get_any(df_car, "Semestre","Periodo","Periodo Académico","Periodo academico")
    if "_SEM" not in df_car.columns and col_sem:
        df_car["_SEM"] = df_car[col_sem].astype(str).str.strip()
    else:
        df_car["_SEM"] = df_car.get("_SEM", pd.Series(dtype=str))
    df_car["_YEAR"] = df_car["_SEM"].map(extract_year_from_period)
    df_car["_IS_INTER"] = df_car["_SEM"].str.lower().str.contains("inter", na=False)

    # ================== TIMEFRAME FILTERS ==================
    def mask_timeframe(series_sem: pd.Series, mode: str, selected_year: int | None, selected_sem: str | None) -> pd.Series:
        s = series_sem.astype(str)
        if mode == "Semestral" and selected_sem:
            return s.str.strip().eq(str(selected_sem))
        if mode == "Anual" and selected_year is not None:
            return s.str.startswith(str(selected_year))
        if mode == "Intersemestral" and selected_year is not None:
            return s.str.startswith(str(selected_year)) & s.str.lower().str.contains("inter")
        # <<< antes: devolvías todo True (muestra TODO el histórico)
        return pd.Series([False]*len(s), index=series_sem.index)

    def filter_df_car(df: pd.DataFrame, mode: str, selected_year: int | None, selected_sem: str | None) -> pd.DataFrame:
        if "_SEM" not in df.columns:
            sc = _get_any(df, "Semestre","Periodo","Periodo Académico","Periodo academico")
            if sc:
                df = df.assign(_SEM=df[sc].astype(str).str.strip())
            else:
                return df
        m = mask_timeframe(df["_SEM"], mode, selected_year, selected_sem)
        return df[m].copy()

    def filter_df_fd(df: pd.DataFrame, mode: str, selected_year: int | None, selected_sem: str | None) -> pd.DataFrame:
        semc = _get_any(df, "Semestre","Periodo","Periodo Académico","Periodo academico")
        ycol = _get_any(df, "Year","Año")
        out = df.copy()
        if semc:
            sem_series = out[semc].astype(str).str.strip()
            m = mask_timeframe(sem_series, mode, selected_year, selected_sem)
            out = out[m].copy()
        elif ycol and selected_year is not None:
            out = out[pd.to_numeric(out[ycol], errors="coerce").astype("Int64") == int(selected_year)].copy()
        return out

    # ================== SIDEBAR ==================
    SEMESTRAL_PERIODS = list_periods_semestral()
    YEARS_ALL = list_years_from_sem()
    INTER_YEARS = years_with_inter()

    with st.sidebar:
        st.markdown("#### Sensitivity analysis")

        sens_mode = st.toggle(
            "Enable sensitivity mode",
            value=st.session_state.get("sens_mode", False),
            key="sens_mode",
            help=(
                "Esta vista permite hacer un **análisis de sensibilidad** SIN modificar la data original.\n"
                "\n"
                "Qué puedes hacer:\n"
                "\n"
                "• Agregar o eliminar cursos por área o a nivel global.\n"
                "\n"
                "• Ver cambios reflejados en tablas y gráficas en tiempo real.\n"
                "\n"
                "• Calcular cuántos cursos se necesitan para alcanzar los objetivos.\n"
                "\n"
                "• Ver el impacto en puntos porcentuales de agregar o eliminar 1 curso de 3 créditos."
            )
        )

        sens_member_placeholder = st.empty()

        if "sens_ops" not in st.session_state:
            st.session_state.sens_ops = []

        if sens_mode:
            st.session_state.setdefault("sens_cat_ps", "None")
            st.session_state.setdefault("sens_cat_qual", "None")
            st.selectbox("P/S Faculty category", ["None", "P", "S"], key="sens_cat_ps")
            st.selectbox("Faculty Qualification", ["None", "SA", "PA", "SP", "IP", "OTHER"], key="sens_cat_qual")
            st.number_input("# N° of courses", min_value=1, step=1, value=1, key="sens_count")
            st.number_input("Course credits", min_value=0.0, step=0.5, value=3.0, key="sens_credits")

            # ADD (suma)
            if st.button("Add", use_container_width=True, key="sens_add"):
                ops_to_add = []
                member_val = st.session_state.get("sens_member", "All")
                cnt  = int(st.session_state.get("sens_count", 1))
                cred = float(st.session_state.get("sens_credits", 3.0))
                if st.session_state.get("sens_cat_ps") and st.session_state["sens_cat_ps"] != "None":
                    ops_to_add.append({"scope": "PS", "cat": st.session_state["sens_cat_ps"], "member": member_val, "credits": cred, "count": cnt})
                if st.session_state.get("sens_cat_qual") and st.session_state["sens_cat_qual"] != "None":
                    ops_to_add.append({"scope": "QUAL", "cat": st.session_state["sens_cat_qual"], "member": member_val, "credits": cred, "count": cnt})
                if ops_to_add:
                    st.session_state.sens_ops.extend(ops_to_add)
                    st.success("Added.")

            # REMOVE (resta)
            if st.button("Remove", use_container_width=True, key="sens_remove_btn"):
                ops_to_add = []
                member_val = st.session_state.get("sens_member", "All")
                cnt  = -abs(int(st.session_state.get("sens_count", 1)))
                cred = float(st.session_state.get("sens_credits", 3.0))
                if st.session_state.get("sens_cat_ps") and st.session_state["sens_cat_ps"] != "None":
                    ops_to_add.append({"scope": "PS", "cat": st.session_state["sens_cat_ps"], "member": member_val, "credits": cred, "count": cnt})
                if st.session_state.get("sens_cat_qual") and st.session_state["sens_cat_qual"] != "None":
                    ops_to_add.append({"scope": "QUAL", "cat": st.session_state["sens_cat_qual"], "member": member_val, "credits": cred, "count": cnt})
                if ops_to_add:
                    st.session_state.sens_ops.extend(ops_to_add)
                    st.success("Removed.")

            if st.button("Reset to original", use_container_width=True, key="sens_reset"):
                st.session_state.sens_ops = []
                st.success("Reset.")

        if not sens_mode:
            st.markdown("<hr style='margin:10px 0;opacity:.4'>", unsafe_allow_html=True)

        st.markdown("#### Timeframe")
        st.session_state.setdefault("time_mode", "Semestral")
        time_mode = st.radio("Timeframe", ["Semestral", "Anual", "Intersemestral"], key="time_mode", label_visibility="collapsed", horizontal=False)

        if time_mode == "Semestral":
            default_sem = SEMESTRAL_PERIODS[0] if SEMESTRAL_PERIODS else "202510"
            st.session_state.setdefault("sel_sem", default_sem)
            sel_sem = st.selectbox("Semester", SEMESTRAL_PERIODS or [default_sem], key="sel_sem")
            sel_year = extract_year_from_period(sel_sem) or (YEARS_ALL[0] if YEARS_ALL else 2025)
            sel_label = str(sel_sem)
        elif time_mode == "Anual":
            default_year = YEARS_ALL[0] if YEARS_ALL else 2025
            st.session_state.setdefault("sel_year", default_year)
            sel_year = st.selectbox("Year", YEARS_ALL or [default_year], key="sel_year")
            sel_sem = None
            sel_label = f"{sel_year} (Annual)"
        else:
            default_i = INTER_YEARS[0] if INTER_YEARS else (YEARS_ALL[0] if YEARS_ALL else 2025)
            # usa SIEMPRE "sel_year" como fuente de verdad
            st.session_state.setdefault("sel_year", default_i)
            sel_year = st.selectbox("Year (Intersemestral)", INTER_YEARS or YEARS_ALL or [default_i], key="sel_year")
            sel_sem = None
            sel_label = f"{sel_year} Intersemestral"

        st.session_state["sel_label"] = sel_label
        st.session_state.setdefault("view_mode", "By Academic Area")
        view_mode = st.selectbox("View", ["By Program", "By Academic Area", "By Field"], key="view_mode")

    # ================== FILTROS BASE ==================
    df_car_base = df_car.copy()
    base = df_fd.copy()
    df_car_filt_all = filter_df_car(df_car_base, time_mode, sel_year, sel_sem)
    f = filter_df_fd(df_fd, time_mode, sel_year, sel_sem)

    # --------- Sensitivity: “Apply to” ---------
    if st.session_state.get("sens_mode", False):
        col_areaCourse = _get_any(df_car_filt_all, "Area del curso","Área del curso","Area del Curso","AREA DEL CURSO")
        col_field = _get_any(df_car_filt_all, "Field","FIELD","Campo","Área de conocimiento")
        program_col = _get_any(df_car_filt_all, "Program","PROGRAM","program")
        members = build_member_list_for_view(df_car_filt_all, view_mode, col_areaCourse, col_field, program_col)
        with st.sidebar:
            sens_member_placeholder.selectbox("Apply to", members, key="sens_member")

    SENS = {"on": bool(st.session_state.get("sens_mode", False)), "ops": st.session_state.get("sens_ops", [])}

    # ================== RELEVANT COLUMNS ==================
    col_ps_fd   = _get_any(df_fd, "P/S", "P - S", "Participating/Supporting")
    col_area_fd = _get_any(df_fd, "AREA_PROFESOR", "Area_Profesor", "Area Profesor", "Área", "Area")
    col_tipo_fd = _get_any(df_fd, "TIPO", "Tipo", "Ranking", "Tipo Ranking")

    col_cred       = _get_any(df_car, "Créditos", "Creditos", "Credits")
    col_tipoC      = _get_any(df_car, "TIPO", "Tipo", "Tipo Ranking")
    col_areaCourse = _get_any(df_car, "Area del curso","Área del curso","Area del Curso","AREA DEL CURSO")
    col_prof       = _get_any(df_car, "Profesor(es)","Profesor","PROFESOR","Docente")
    col_code       = _get_any(df_car, "Materia","Código Materia","Codigo Materia","CODIGO MATERIA","Código","Codigo","Course Code")
    col_name       = _get_any(df_car, "Nombre largo curso","Nombre Curso","Nombre del curso","Course Name")
    col_field      = _get_any(df_car, "Field","FIELD","Campo","Área de conocimiento")
    col_prog       = _get_any(df_car, "Program","PROGRAM","program")
    col_ps_C       = _get_any(df_car, "P/S","P - S","Participating/Supporting")

    # ---------- Stylers ----------
    def style_percent_tables(df_, id_col):
        sty = pd.DataFrame('', index=df_.index, columns=df_.columns)
        colP = "%P"; colSA = "%SA"; colOTHER = "%OTHER"
        p_vals = pd.to_numeric(df_[colP], errors="coerce")
        sa_vals = pd.to_numeric(df_[colSA], errors="coerce")
        other_vals = pd.to_numeric(df_[colOTHER], errors="coerce")
        is_total = df_[id_col].astype(str).str.upper().eq("TOTAL")
        sty.loc[(~is_total) & (p_vals < 60), colP] = 'background-color:#FDE2E2;'
        sty.loc[is_total & (p_vals < 75), colP] = 'background-color:#FDE2E2; font-weight:700;'
        sty.loc[sa_vals < 40, colSA] = 'background-color:#FDE2E2;'
        sty.loc[other_vals > 10, colOTHER] = 'background-color:#FDE2E2;'
        for c in sty.columns:
            sty.loc[is_total, c] = (sty.loc[is_total, c].astype(str) + 'font-weight:700;').str.replace(';;',';', regex=False)
        return sty

    # ---------- util de estilo para la tabla de "Needed + Impact" ----------
    def _style_needed_impact(df_, id_col):
        sty = pd.DataFrame('', index=df_.index, columns=df_.columns)
        numeric_cols = [c for c in df_.columns if c != id_col]
        # rojo claro para todo valor != 0
        for c in numeric_cols:
            vals = pd.to_numeric(df_[c], errors="coerce").fillna(0)
            sty.loc[vals != 0, c] = 'background-color:#FDE2E2;'
        return sty

    # ---------- helpers de necesidades por objetivo (dos columnas) ----------
    def _needed_pairs_for_obj(
        objective: str,
        scope_label: str,
        P: float, S: float, SA_: float, PA_: float, SP_: float, IP_: float, OT_: float,
        totals: dict[str,float],
        credits_each: float = 3.0
    ) -> tuple[int, int]:
        """
        Devuelve dos números (enteros >= 0) según el objetivo:
          - %P   -> (Need_P_more, Need_S_less)
          - %SA  -> (Need_SA_more, Need_NonSA_less)
          - %OTHER -> (Need_OTHER_less, Need_NonOTHER_more)

        Si scope="Overall", calcula con TOT (global) y limita por factibilidad del renglón cuando es "quitar".
        Nunca devuelve None; si no alcanza, devuelve el máximo posible (capped).
        """
        t_map = {"%P": (60.0, 75.0), "%SA": (40.0, 40.0), "%OTHER": (10.0, 10.0)}
        tgt_area, tgt_overall = t_map[objective]
        t = (tgt_area if scope_label == "By area" else tgt_overall) / 100.0

        # valores por fila
        TQ = SA_ + PA_ + SP_ + IP_ + OT_
        nonSA = PA_ + SP_ + IP_ + OT_
        nonOTHER = SA_ + PA_ + SP_ + IP_

        # totales
        Ptot = totals.get("P",0.0); Stot = totals.get("S",0.0)
        SAt  = totals.get("SA",0.0); PAt = totals.get("PA",0.0)
        SPt  = totals.get("SP",0.0); IPt = totals.get("IP",0.0)
        OTt  = totals.get("OTHER",0.0)
        TQt  = SAt + PAt + SPt + IPt + OTt
        nonSAt = PAt + SPt + IPt + OTt
        nonOTHERt = SAt + PAt + SPt + IPt

        # --- %P ---
        if objective == "%P":
            # Aumentar P (+3cr cada profesor)
            if scope_label == "By area":
                nP = _needed_for_pctP(P, S, tgt_area, credits_each)
            else:
                # (Ptot + c*n)/(Ptot + Stot + c*n) >= t  ->  n >= (t*(P+S) - P)/( (1-t)*c )
                den = credits_each * (1 - t)
                rhs = 0 if den <= 0 else (t*(Ptot+Stot) - Ptot) / den
                nP  = max(0, math.ceil(rhs))

            # Quitar S (-3cr)
            if scope_label == "By area":
                #  P/(P + S - c*n) >= t  ->  n >= (t*(P+S) - P)/(t*c)
                den = credits_each * t if t > 0 else float('inf')
                rhs = 0 if den == float('inf') else (t*(P+S) - P) / den
                nS_less = max(0, math.ceil(rhs))
                # factibilidad
                nmax = math.floor(S / credits_each) if credits_each > 0 else 0
                nS_less = min(nS_less, max(0, nmax))
            else:
                # overall: Ptot/(Ptot + Stot - c*n) >= t
                den = credits_each * t if t > 0 else float('inf')
                rhs = 0 if den == float('inf') else (t*(Ptot+Stot) - Ptot) / den
                nS_less = max(0, math.ceil(rhs))
                # factibilidad: solo puedo quitar del renglón actual
                nmax = math.floor(S / credits_each) if credits_each > 0 else 0
                nS_less = min(nS_less, max(0, nmax))

            return (nP, nS_less)

        # --- %SA ---
        if objective == "%SA":
            # Aumentar SA (+3cr)
            if scope_label == "By area":
                nSA = _needed_for_pctSA(SA_, nonSA, tgt_area, credits_each)
            else:
                den = credits_each * (1 - t)
                rhs = 0 if den <= 0 else (t*TQt - SAt) / den
                nSA = max(0, math.ceil(rhs))

            # Quitar No-SA (PA+SP+IP+OTHER) (-3cr)
            if scope_label == "By area":
                # SA/(SA + nonSA - c*n) >= t -> n >= (t*(SA+nonSA) - SA)/(t*c)
                den = credits_each * t if t > 0 else float('inf')
                rhs = 0 if den == float('inf') else (t*(SA_+nonSA) - SA_) / den
                nNonSA_less = max(0, math.ceil(rhs))
                nmax = math.floor(nonSA / credits_each) if credits_each > 0 else 0
                nNonSA_less = min(nNonSA_less, max(0, nmax))
            else:
                den = credits_each * t if t > 0 else float('inf')
                rhs = 0 if den == float('inf') else (t*TQt - SAt) / den
                nNonSA_less = max(0, math.ceil(rhs))
                nmax = math.floor(nonSA / credits_each) if credits_each > 0 else 0
                nNonSA_less = min(nNonSA_less, max(0, nmax))

            return (nSA, nNonSA_less)

        # --- %OTHER ---
        # Quitar OTHER (-3cr)
        if scope_label == "By area":
            # (OT - c*n)/(TQ - c*n) <= 0.10  ->  c*n >= (OT - 0.10*TQ)/0.90
            need_credits = (OT_ - 0.10*TQ) / 0.90
            nOT_less = 0 if need_credits <= 0 else math.ceil(need_credits / credits_each)
            nmax = math.floor(OT_ / credits_each) if credits_each > 0 else 0
            nOT_less = min(nOT_less, max(0, nmax))
        else:
            # overall
            need_credits = (OTt - 0.10*TQt) / 0.90
            nOT_less = 0 if need_credits <= 0 else math.ceil(need_credits / credits_each)
            nmax = math.floor(OT_ / credits_each) if credits_each > 0 else 0
            nOT_less = min(nOT_less, max(0, nmax))

        # Aumentar No-OTHER (+3cr) -> OT/(OT + nonOTHER + c*n) <= 0.10
        if scope_label == "By area":
            # n >= (0.90*OT - 0.10*nonOTHER)/(0.10*c) = (9*OT - nonOTHER)/c
            num = (9*OT_ - nonOTHER)
            den = credits_each
            nNonOT_more = 0 if num <= 0 else math.ceil(num / den)
        else:
            num = (9*OTt - nonOTHERt)
            den = credits_each
            nNonOT_more = 0 if num <= 0 else math.ceil(num / den)

        return (nOT_less, nNonOT_more)

    # ---------- impacto (siempre visible) ----------
    def _impact_pair(obj: str, area_vals: dict[str,float], totals: dict[str,float], scope_label: str, credits_each: float = 3.0):
        if scope_label == "By area":
            up_pp, down_pp = _impact_pp_area(obj, area_vals, credits_each)
        else:
            up_pp, down_pp = _impact_pp_overall_if_area_changes(obj, totals, credits_each)
        # devolver números (no strings)
        return round(up_pp, 2), round(down_pp, 2)

    # === HEATMAP + ROJO-CLARO PARA "Needed" ===
    def _style_impact_heatmap(df: pd.DataFrame, id_col: str):
        """
        - Heatmap (verde→amarillo→naranja→rojo) para columnas 'Impact +3cr (pp)' y 'Impact -3cr (pp)'.
          Se colorea por magnitud absoluta (mayor impacto = más rojo).
        - Fondo rojo claro en columnas 'Needed ...' cuando el valor != 0.
        - No toca la columna del identificador (id_col) ni otras columnas.
        """
        # DataFrame de estilos vacío
        sty = pd.DataFrame('', index=df.index, columns=df.columns)

        # --- 1) Rojo claro para "Needed ..." cuando != 0 ---
        needed_cols = [c for c in df.columns if c.startswith("Needed ")]
        for c in needed_cols:
            if c in df:
                vals = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
                sty.loc[vals != 0, c] = sty.loc[vals != 0, c].astype(str) + 'background-color:#FDE2E2;'

        # --- 2) Heatmap para columnas de Impact ---
        impact_cols = [c for c in df.columns if c.startswith("Impact ")]
        if impact_cols:
            # Usamos por defecto la magnitud del "+3cr" si existe; si no, el promedio abs de todas
            if "Impact +3cr (pp)" in df.columns:
                base_vals = pd.to_numeric(df["Impact +3cr (pp)"], errors="coerce").abs()
            else:
                base_vals = (
                    df[impact_cols]
                    .apply(pd.to_numeric, errors="coerce")
                    .abs()
                    .mean(axis=1)
                )
            base_vals = base_vals.fillna(0.0)
            vmin = float(np.nanmin(base_vals.values)) if base_vals.size else 0.0
            vmax = float(np.nanmax(base_vals.values)) if base_vals.size else 0.0
            rng = (vmax - vmin) if (vmax - vmin) > 1e-12 else 1.0  # evita división por cero

            # Paleta suave: verde → amarillo → naranja → rojo
            # (cuanto mayor el impacto, más "caliente")
            def color_for(val_abs: float) -> str:
                z = (val_abs - vmin) / rng  # 0..1  (0 = menor impacto, 1 = mayor impacto)
                if z >= 0.60:
                    return "#D9F2D9"  # verde claro
                elif z >= 0.40:
                    return "#FFF6B3"  # amarillo
                elif z >= 0.20:
                    return "#FFD6A6"  # naranja
                else:
                    return "#F5B5B5"  # rojo

            # Aplica la paleta a TODAS las columnas de impacto (según la misma escala)
            for c in impact_cols:
                col_vals = pd.to_numeric(df[c], errors="coerce").abs().fillna(0.0)
                for i, v in col_vals.items():
                    sty.at[i, c] = sty.at[i, c] + f'background-color:{color_for(float(v))};'

        # Asegura que el id_col no reciba estilo accidental
        if id_col in sty.columns:
            sty[id_col] = ''

        return sty

    # ================== PRINCIPAL ==================
    st.markdown("---")

    # --- helpers específicos para el cabezote ---
    def _guess_prof_cols(df: pd.DataFrame) -> list[str]:
        """
        Devuelve columnas candidatas para identificar un profesor.
        Prioridad: Documento/ID/Email -> nombre/profesor.
        """
        pri = []
        # IDs / correos
        for c in df.columns:
            cl = str(c).strip().lower()
            if any(k in cl for k in ["documento", "identific", "id", "correo", "email", "mail"]):
                pri.append(c)
        # Nombres / profesor
        for c in df.columns:
            cl = str(c).strip().lower()
            if any(k in cl for k in ["prof", "docent", "nombre", "name"]):
                pri.append(c)
        # Quitar duplicados preservando orden
        seen, out = set(), []
        for c in pri:
            if c not in seen:
                out.append(c); seen.add(c)
        # Fallback simple si nada matchea
        if not out:
            for cand in ["Profesor(es)", "Profesor", "PROFESOR", "Docente", "Nombre", "Name", "Profesor(a)"]:
                if cand in df.columns:
                    out.append(cand)
                    break
        return out

    def _unique_prof_count(df: pd.DataFrame, cols: list[str]) -> int:
        if df is None or df.empty:
            return 0
        # Construir una UID robusta a partir de las columnas disponibles
        use = [c for c in cols if c in df.columns]
        if not use:
            # último recurso: filas únicas por todas las columnas visibles (puede sobre-contar)
            return int(df.astype(str).drop_duplicates().shape[0])
        uid = df[use].astype(str).apply(lambda s: s.str.strip()).fillna("").agg(" | ".join, axis=1)
        return int(uid.nunique())

    def _filter_fd_by_timeframe(df_fd: pd.DataFrame, time_mode: str, sel_year, sel_sem) -> pd.DataFrame:
        """
        Filtra Faculty Distribution por Semestral / Anual / Intersemestral.

        - Semestral:        == sel_sem (p.ej. '202520')
        - Anual:            empieza por sel_year (incluye 10, 20 e intersemestral)
        - Intersemestral:   contiene el año (en cualquier posición) y 'inter' en el texto
                            (soporta '2025 Intersemestral', 'Intersemestral 2025', '2025-Inter', etc.)
        """
        if df_fd is None or df_fd.empty:
            return df_fd.iloc[0:0]

        sem_col = _get_any(df_fd, "Semestre", "Periodo", "Periodo Académico", "Periodo academico")
        if not sem_col:
            return df_fd.iloc[0:0]

        s = df_fd[sem_col].astype(str).str.strip()
        tm = (time_mode or "Semestral").strip()

        if tm == "Semestral" and sel_sem:
            m = s.eq(str(sel_sem))
            return df_fd[m].copy()

        if tm == "Anual" and sel_year is not None:
            m = s.str.startswith(str(sel_year))
            return df_fd[m].copy()

        if tm == "Intersemestral" and sel_year is not None:
            y = str(sel_year)
            has_year  = s.str.contains(rf"(?:^|[^0-9]){re.escape(y)}(?:[^0-9]|$)", case=False, regex=True)
            has_inter = s.str.contains("inter", case=False, na=False)
            m = has_year & has_inter
            return df_fd[m].copy()

        return df_fd.copy()

    def _count_teaching_from_fd_timeaware(df_fd: pd.DataFrame, time_mode: str, sel_year, sel_sem) -> dict[str,int]:
        """
        Cuenta profesores ÚNICOS en Faculty Distribution según timeframe:
          - Full-time (FT):   PLANTA_CATEDRA == 'PLANTA'
          - Part-time (PT):   PLANTA_CATEDRA == 'CÁTEDRA' / 'CATEDRA'
          - Participating P:  P/S == 'P'
          - Supporting   S:   P/S == 'S'
        """
        if df_fd is None or df_fd.empty:
            return {"FT":0, "PT":0, "P":0, "S":0}

        dff = _filter_fd_by_timeframe(df_fd, time_mode, sel_year, sel_sem)
        if dff is None or dff.empty:
            return {"FT":0, "PT":0, "P":0, "S":0}

        prof_cols = _guess_prof_cols(dff)

        # columnas de clasificación
        pc_col = _get_any(dff, "PLANTA_CATEDRA", "Planta_Catedra", "Planta/Cátedra", "PLANTA CATEDRA", "Planta/Catedra")
        ps_col = _get_any(dff, "P/S", "P - S", "Participating/Supporting", "P S")

        # Full-time / Part-time
        ft = pt = 0
        if pc_col:
            tag = _norm_str(dff[pc_col])
            ft_df = dff[tag.eq("planta")]
            pt_df = dff[tag.isin({"catedra", "cátedra"})]
            ft = _unique_prof_count(ft_df, prof_cols)
            pt = _unique_prof_count(pt_df, prof_cols)

        # Participating / Supporting
        p_cnt = s_cnt = 0
        if ps_col:
            tps = _norm_str(dff[ps_col])
            p_df = dff[tps.eq("p")]
            s_df = dff[tps.eq("s")]
            p_cnt = _unique_prof_count(p_df, prof_cols)
            s_cnt = _unique_prof_count(s_df, prof_cols)

        return {"FT":ft, "PT":pt, "P":p_cnt, "S":s_cnt}

    def compute_header_counts_teaching(df_fd: pd.DataFrame, time_mode: str, sel_year, sel_sem, sens: dict) -> dict:
        base = _count_teaching_from_fd_timeaware(df_fd, time_mode, sel_year, sel_sem)

        # Sensibilidad: +P suma a Full-time y Participating; +S suma a Part-time y Supporting
        dP = dS = 0
        if sens.get("on") and sens.get("ops"):
            for op in sens["ops"]:
                if op.get("scope") == "PS":
                    if op.get("cat") == "P":
                        dP += int(op.get("count", 0))
                    elif op.get("cat") == "S":
                        dS += int(op.get("count", 0))

        return {
            "Full-time":     max(0, base["FT"] + dP),
            "Part-time":     max(0, base["PT"] + dS),
            "Participating": max(0, base["P"]  + dP),
            "Supporting":    max(0, base["S"]  + dS),
        }

    # === Subheader ===

    st.subheader(f"Faculty Sufficiency and Qualifications — {st.session_state.get('sel_label','Selected')}")

    # ====== NORMALIZACIÓN BASE PARA CARTELERA + EXCLUSIONES ======
    if not all([col_cred, col_tipoC, col_areaCourse]):
        st.error("Missing columns in 'BD_Cartelera': 'Credits', 'TIPO', and/or 'Academic Area (course)'.")
    else:
        df_car_n = df_car.copy()
        df_car_n["_CRED"] = pd.to_numeric(df_car_n[col_cred], errors="coerce").fillna(0.0)
        df_car_n["_TIPO"] = _norm_str(df_car_n[col_tipoC]).map(normalize_tipo)
        if "_SEM" not in df_car_n.columns:
            sc = _get_any(df_car_n, "Semestre","Periodo","Periodo Académico","Periodo academico")
            df_car_n["_SEM"] = df_car_n[sc].astype(str).str.strip() if sc else ""
        df_car_n["_YEAR"] = df_car_n["_SEM"].map(extract_year_from_period)
        df_car_n["_AREA"] = df_car_n[col_areaCourse].astype(str).str.strip()
        col_ps_C_local = _get_any(df_car_n, "P/S","P - S","Participating/Supporting")
        df_car_n["_PS"] = _norm_str(df_car_n[col_ps_C_local]).map(normalize_ps) if col_ps_C_local else ""

        # excluir programas
        program_col0 = _get_any(df_car_n, "Program","PROGRAM","program")
        EXCLUDE_SUBJ = {"CONT", "E-IMER", "E-ENEG", "E-AFIN"}
        if program_col0:
            mask_ok = ~df_car_n[program_col0].astype(str).str.strip().str.upper().isin(EXCLUDE_SUBJ)
            df_car_global = df_car_n[mask_ok].copy()
        else:
            df_car_global = df_car_n.copy()

        # ---------- Filtro por timeframe seleccionado ----------
        sel_label = st.session_state.get("sel_label")
        time_mode = st.session_state.get("time_mode", "Semestral")
        sel_year  = st.session_state.get("sel_year")
        sel_sem   = st.session_state.get("sel_sem")

        fil = filter_df_car(df_car_global, time_mode, sel_year, sel_sem)
        df_car_filt_all = fil.copy()  # usar en expander/tabla/dona

        # ============================ VISTAS ============================
        def build_percent_table(base_idx_name, agg_tipo, agg_ps):
            den_ps = (agg_ps["P"] + agg_ps["S"]).replace(0, pd.NA)
            p_share = (agg_ps["P"] / den_ps) * 100
            s_share = 100 - p_share
            denom_q = (agg_tipo.sum(axis=1)).replace(0, pd.NA)
            dfm = pd.DataFrame({
                base_idx_name: agg_tipo.index,
                "%P": p_share,
                "%S": s_share,
                "%SA": (agg_tipo["SA"] / denom_q) * 100,
                "%OTHER": (agg_tipo["OTHER"] / denom_q) * 100,
            }).fillna(0.0)

            # TOTAL
            tot_P, tot_S = agg_ps["P"].sum(), agg_ps["S"].sum()
            tot_den_ps = tot_P + tot_S
            p_tot = (tot_P / tot_den_ps * 100) if tot_den_ps else 0.0
            s_tot = 100 - p_tot
            tipo_sums = agg_tipo[["SA","PA","SP","IP","OTHER"]].sum(axis=0)
            denom_q_tot = float(tipo_sums.sum())

            total_row = {
                base_idx_name: "TOTAL",
                "%P": round(p_tot, 1),
                "%S": round(s_tot, 1),
                "%SA": round((tipo_sums["SA"] / denom_q_tot * 100) if denom_q_tot else 0.0, 1),
                "%OTHER": round((tipo_sums["OTHER"] / denom_q_tot * 100) if denom_q_tot else 0.0, 1),
            }
            dfm[["%P","%S","%SA","%OTHER"]] = dfm[["%P","%S","%SA","%OTHER"]].round(1)
            dfm = pd.concat([dfm, pd.DataFrame([total_row])], ignore_index=True)
            return dfm[[f"{base_idx_name}", "%P", "%S", "%SA", "%OTHER"]]

        if fil.empty:
            st.info(f"No records for the selected timeframe: {sel_label}.")
        else:
            # ========== BY ACADEMIC AREA ==========
            if view_mode == "By Academic Area":
                colT, colG = st.columns([6,6], gap="large")

                # Agregaciones
                agg_tipo = (fil.groupby(["_AREA","_TIPO"], dropna=False)["_CRED"].sum().unstack(fill_value=0.0))
                for k in ["SA","PA","SP","IP","OTHER"]:
                    if k not in agg_tipo.columns: agg_tipo[k] = 0.0
                agg_tipo = agg_tipo[["SA","PA","SP","IP","OTHER"]]

                agg_ps = (fil.groupby(["_AREA","_PS"], dropna=False)["_CRED"].sum().unstack(fill_value=0.0))
                for k in ["P","S"]:
                    if k not in agg_ps.columns: agg_ps[k] = 0.0
                agg_ps = agg_ps[["P","S"]]

                # Sensibilidad
                base_agg_ps = agg_ps.copy()
                base_agg_tipo = agg_tipo.copy()
                if SENS["on"] and SENS["ops"]:
                    mod_agg_ps, mod_agg_tipo = apply_ops_to_aggs(base_agg_ps, base_agg_tipo, SENS["ops"])
                else:
                    mod_agg_ps, mod_agg_tipo = base_agg_ps, base_agg_tipo

                with colT:
                    # Controles: solo si Sensitivity ON y toggle ON se muestra el selector; el IMPACTO ya es siempre visible
                    needed_mode = False
                    if SENS["on"]:
                        r1c1, r1c2, r1c3 = st.columns([1.8, 1.1, 1.6])
                        with r1c1:
                            needed_mode = st.toggle("# N° of courses needed for…", value=False, key="area_needed_mode", help="La tabla muestra la cantidad de cursos de 3 créditos que se necesitan para llegar al objetivo y el impacto en puntos porcentuales de agregar o eliminar un curso de 3 cr.")
                        if needed_mode:
                            with r1c2:
                                objective = st.selectbox("Objective", ["%P", "%SA", "%OTHER"], key="area_objective")
                            with r1c3:
                                scope_label = st.radio(
                                    "Target scope",
                                    ["By area", "Overall"],
                                    horizontal=True,
                                    key="area_scope",
                                    help=(
                                        "**Objective by Area**\n"
                                        "- %P > 60%\n"
                                        "- %SA > 40%\n"
                                        "- %OTHER < 10%\n\n"
                                        "**Overall Objective**\n"
                                        "- %P > 75%\n"
                                        "- %SA > 40%\n"
                                        "- %OTHER < 10%"
                                    )
                                )

                        else:
                            objective = st.session_state.get("area_objective", "%P")
                            scope_label = st.session_state.get("area_scope", "By area")
                    else:
                        objective = st.session_state.get("area_objective", "%P")
                        scope_label = st.session_state.get("area_scope", "By area")

                    if not needed_mode:
                        metrics_tbl = build_percent_table("Academic Area", mod_agg_tipo, mod_agg_ps)
                        _download_xlsx_button(metrics_tbl, f"table_ByArea_{_slugify(sel_label)}.xlsx",
                                              key=f"dl_tbl_area_{_slugify(sel_label)}", label="⬇️ Download table (Excel)")
                        styled_tbl = (
                            metrics_tbl.style
                            .format({"%P": "{:.1f}%", "%S": "{:.1f}%", "%SA": "{:.1f}%", "%OTHER": "{:.1f}%"})
                            .apply(style_percent_tables, id_col="Academic Area", axis=None)
                            .hide(axis="index")
                        )
                        st.dataframe(styled_tbl, use_container_width=True, hide_index=True)
                    else:
                        # ===== Tabla: Needed (dos columnas) + Impact (siempre) SIN TOTAL =====
                        # union de índices para no perder filas
                        idx_all = sorted(set(mod_agg_ps.index.tolist()) | set(mod_agg_tipo.index.tolist()))
                        p   = mod_agg_ps["P"].reindex(idx_all, fill_value=0.0)
                        s   = mod_agg_ps["S"].reindex(idx_all, fill_value=0.0)
                        sa  = mod_agg_tipo["SA"].reindex(idx_all, fill_value=0.0)
                        pa  = mod_agg_tipo["PA"].reindex(idx_all, fill_value=0.0)
                        sp  = mod_agg_tipo["SP"].reindex(idx_all, fill_value=0.0)
                        ip  = mod_agg_tipo["IP"].reindex(idx_all, fill_value=0.0)
                        oth = mod_agg_tipo["OTHER"].reindex(idx_all, fill_value=0.0)

                        totals = {
                            "P": float(p.sum()), "S": float(s.sum()),
                            "SA": float(sa.sum()), "PA": float(pa.sum()),
                            "SP": float(sp.sum()), "IP": float(ip.sum()),
                            "OTHER": float(oth.sum())
                        }

                        # nombres de columnas según objetivo
                        if objective == "%P":
                            main_col, aux_col = "P courses needed (3cr)", "Less S courses needed (3cr)"
                        elif objective == "%SA":
                            main_col, aux_col = "SA courses needed (3cr)", "Less other Qualific. courses needed (3cr)"
                        else:
                            main_col, aux_col = "Less OTHER Courses needed (3cr)", "More other Qualific. courses needed (3cr)"

                        rows = []
                        for label in idx_all:
                            Pv, Sv = float(p.get(label,0.0)), float(s.get(label,0.0))
                            SAv, PAv = float(sa.get(label,0.0)), float(pa.get(label,0.0))
                            SPv, IPv = float(sp.get(label,0.0)), float(ip.get(label,0.0))
                            OTv      = float(oth.get(label,0.0))

                            need1, need2 = _needed_pairs_for_obj(
                                objective, scope_label,
                                Pv, Sv, SAv, PAv, SPv, IPv, OTv,
                                totals, credits_each=3.0
                            )

                            area_vals = {"P":Pv,"S":Sv,"SA":SAv,"PA":PAv,"SP":SPv,"IP":IPv,"OTHER":OTv}
                            up_pp, down_pp = _impact_pair(objective, area_vals, totals, scope_label, credits_each=3.0)

                            rows.append({
                                "Academic Area": label,
                                main_col: int(need1),
                                aux_col:  int(need2),
                                "Impact increasing 1 course in %p.p.": up_pp,
                                "Impact decreasing 1 course %p.p.": down_pp
                            })

                        need_tbl = pd.DataFrame(rows)
                        # formateo + heatmap (verde→amarillo→naranja→rojo) SOLO en columnas "Impact ..."
                        fmt_map = {}
                        if 'Academic Area' in need_tbl.columns:
                            fmt_map['Academic Area'] = '{}'
                        for col in ['Needed P (3cr)', 'Needed S less (3cr)', 'Needed SA (3cr)', 'Needed OTHER less (3cr)', 'Needed OTHER more (3cr)']:
                            if col in need_tbl.columns:
                                fmt_map[col] = '{:.0f}'
                        for col in ['Impact increasing 1 course in %p.p.', 'Impact decreasing 1 course %p.p.']:
                            if col in need_tbl.columns:
                                fmt_map[col] = '{:+.2f}'

                        styled = (
                            need_tbl.style
                            .format(fmt_map)
                            .apply(_style_impact_heatmap, id_col="Academic Area", axis=None)  # HEATMAP aplicado aquí
                            .hide(axis="index")
                        )
                        _download_xlsx_button(
                            need_tbl,
                            f"needed_ByArea_{_slugify(sel_label)}_{_slugify(objective)}_{_slugify(scope_label)}.xlsx",
                            key=f"dl_need_area_{_slugify(sel_label)}_{_slugify(objective)}_{_slugify(scope_label)}",
                            label="⬇️ Download (Excel)"
                        )
                        st.markdown(styled.to_html(escape=False), unsafe_allow_html=True)

                # ========== Series históricas ==========
                df_hist = df_car_global.copy()
                agg_ps_all = (df_hist.groupby(["_SEM","_AREA","_PS"], dropna=False)["_CRED"].sum().unstack(fill_value=0.0))
                for k in ["P","S"]:
                    if k not in agg_ps_all.columns: agg_ps_all[k] = 0.0
                agg_ps_all["P_share"] = (agg_ps_all["P"] / (agg_ps_all["P"] + agg_ps_all["S"]).replace(0, pd.NA)) * 100
                agg_ps_all = agg_ps_all.reset_index()

                agg_tipo_all = (df_hist.groupby(["_SEM","_AREA","_TIPO"], dropna=False)["_CRED"].sum().unstack(fill_value=0.0))
                for k in ["SA","PA","SP","IP","OTHER"]:
                    if k not in agg_tipo_all.columns: agg_tipo_all[k] = 0.0
                den_all = (agg_tipo_all[["SA","PA","SP","IP","OTHER"]].sum(axis=1)).replace(0, pd.NA)
                agg_tipo_all["SA_share"] = (agg_tipo_all["SA"] / den_all) * 100
                agg_tipo_all["OTHER_share"] = (agg_tipo_all["OTHER"] / den_all) * 100
                agg_tipo_all = agg_tipo_all.reset_index()

                tot_by_sem_P = (df_hist.groupby(["_SEM","_PS"])["_CRED"].sum().unstack(fill_value=0.0))
                for k in ["P","S"]:
                    if k not in tot_by_sem_P.columns: tot_by_sem_P[k] = 0.0
                tot_by_sem_P["P_share"] = (tot_by_sem_P["P"] / (tot_by_sem_P["P"] + tot_by_sem_P["S"]).replace(0, pd.NA)) * 100
                tot_by_sem_P = tot_by_sem_P.reset_index()

                tot_by_sem_tipo = (df_hist.groupby(["_SEM","_TIPO"])["_CRED"].sum().unstack(fill_value=0.0))
                for k in ["SA","PA","SP","IP","OTHER"]:
                    if k not in tot_by_sem_tipo.columns: tot_by_sem_tipo[k] = 0.0
                den_tot = (tot_by_sem_tipo[["SA","PA","SP","IP","OTHER"]].sum(axis=1)).replace(0, pd.NA)
                tot_by_sem_tipo["SA_share"] = (tot_by_sem_tipo["SA"] / den_tot) * 100
                tot_by_sem_tipo["OTHER_share"] = (tot_by_sem_tipo["OTHER"] / den_tot) * 100
                tot_by_sem_tipo = tot_by_sem_tipo.reset_index()

                # Adaptación a modo temporal
                agg_ps_all_tm  = transform_for_time_mode_ps(agg_ps_all.rename(columns={"_AREA":"__LEVEL__"})).rename(columns={"__LEVEL__":"_AREA"})
                agg_tipo_sa_tm = transform_for_time_mode_tipo(agg_tipo_all.rename(columns={"_AREA":"__LEVEL__"}), "SA_share").rename(columns={"__LEVEL__":"_AREA"})
                agg_tipo_ot_tm = transform_for_time_mode_tipo(agg_tipo_all.rename(columns={"_AREA":"__LEVEL__"}), "OTHER_share").rename(columns={"__LEVEL__":"_AREA"})
                agg_tipo_all_tm = (
                    agg_tipo_sa_tm.drop(columns=[c for c in ["OTHER_share"] if c in agg_tipo_sa_tm], errors="ignore")
                    .merge(agg_tipo_ot_tm[["_SEM","_AREA","OTHER","SA","PA","SP","IP","OTHER_share"]],
                           on=["_SEM","_AREA","SA","PA","SP","IP","OTHER"], how="outer")
                )
                tot_by_sem_P_tm = transform_for_time_mode_ps(tot_by_sem_P.copy())
                tot_tipo_sa_tm  = transform_for_time_mode_tipo(tot_by_sem_tipo.copy(), "SA_share")
                tot_tipo_ot_tm  = transform_for_time_mode_tipo(tot_by_sem_tipo.copy(), "OTHER_share")
                tot_by_sem_tipo_tm = (
                    tot_tipo_sa_tm.drop(columns=[c for c in ["OTHER_share"] if c in tot_tipo_sa_tm], errors="ignore")
                    .merge(tot_tipo_ot_tm[["_SEM","SA","PA","SP","IP","OTHER","OTHER_share"]],
                           on=["_SEM","SA","PA","SP","IP","OTHER"], how="outer")
                )

                key_col, x_labels, x_map = build_time_axis_for_history(df_hist)
                if time_mode == "Semestral":
                    sel_x = x_map.get(str(sel_sem)) if sel_sem else None
                    sel_label_exact = str(sel_sem) if sel_sem else None
                elif time_mode == "Anual":
                    sel_x = x_map.get(sel_year) if sel_year is not None else None
                    sel_label_exact = sel_year
                else:
                    inter_label = f"{sel_year} Intersemestral" if sel_year else None
                    sel_x = x_map.get(inter_label) if inter_label else None
                    sel_label_exact = inter_label

                if SENS["on"] and SENS["ops"] and sel_label_exact is not None:
                    agg_ps_all_tm, agg_tipo_all_tm, tot_by_sem_P_tm, tot_by_sem_tipo_tm = apply_sensitivity_to_history(
                        agg_ps_all_tm, agg_tipo_all_tm, tot_by_sem_P_tm, tot_by_sem_tipo_tm,
                        level_name="_AREA",
                        sel_label_value=sel_label_exact,
                        ops=SENS["ops"],
                        member_all_label="All"
                    )

                areas_all = sorted(set(_safe_unique_labels(agg_ps_all_tm, "_AREA")) | set(_safe_unique_labels(agg_tipo_all_tm, "_AREA")))
                with colG:
                    draw_history(
                        "Evolution by Academic Area",
                        level_name="_AREA",
                        level_values=areas_all,
                        metric_kind="%P",
                        total_series_builders={"P": tot_by_sem_P_tm, "SA": tot_by_sem_tipo_tm, "OTHER": tot_by_sem_tipo_tm},
                        agg_ps_all=agg_ps_all_tm,
                        agg_tipo_all=agg_tipo_all_tm,
                        x_labels=x_labels, x_map=x_map, sel_x=sel_x
                    )

            # -------------- BY FIELD --------------
            elif view_mode == "By Field" and col_field:
                colF_L, colF_R = st.columns([6,6], gap="large")
                fil_field = fil.copy()
                fil_field["_FIELD"] = fil_field[col_field].astype(str).str.strip()

                agg_tipo_f = (fil_field.groupby(["_FIELD","_TIPO"], dropna=False)["_CRED"].sum().unstack(fill_value=0.0))
                for k in ["SA","PA","SP","IP","OTHER"]:
                    if k not in agg_tipo_f.columns: agg_tipo_f[k] = 0.0
                agg_tipo_f = agg_tipo_f[["SA","PA","SP","IP","OTHER"]]

                agg_ps_f = (fil_field.groupby(["_FIELD","_PS"], dropna=False)["_CRED"].sum().unstack(fill_value=0.0))
                for k in ["P","S"]:
                    if k not in agg_ps_f.columns: agg_ps_f[k] = 0.0
                agg_ps_f = agg_ps_f[["P","S"]]

                base_agg_ps = agg_ps_f.copy()
                base_agg_tipo = agg_tipo_f.copy()
                if SENS["on"] and SENS["ops"]:
                    mod_agg_ps, mod_agg_tipo = apply_ops_to_aggs(base_agg_ps, base_agg_tipo, SENS["ops"])
                else:
                    mod_agg_ps, mod_agg_tipo = base_agg_ps, base_agg_tipo

                with colF_L:
                    needed_mode_f = False
                    if SENS["on"]:
                        r1c1, r1c2, r1c3 = st.columns([1.8, 1.1, 1.6])
                        with r1c1:
                            needed_mode_f = st.toggle("Show necessary # of Faculty for…", value=False, key="field_needed_mode")
                        if needed_mode_f:
                            with r1c2:
                                objective_f = st.selectbox("Objective", ["%P", "%SA", "%OTHER"], key="field_objective")
                            with r1c3:
                                scope_label_f = st.radio("Target scope", ["By area", "Overall"], horizontal=True, key="field_scope")
                        else:
                            objective_f = st.session_state.get("field_objective", "%P")
                            scope_label_f = st.session_state.get("field_scope", "By area")
                    else:
                        objective_f = st.session_state.get("field_objective", "%P")
                        scope_label_f = st.session_state.get("field_scope", "By area")

                    if not needed_mode_f:
                        metrics_tbl_f = build_percent_table("Field", mod_agg_tipo, mod_agg_ps)
                        _download_xlsx_button(metrics_tbl_f, f"table_ByField_{_slugify(sel_label)}.xlsx",
                                              key=f"dl_tbl_field_{_slugify(sel_label)}", label="⬇️ Download table (Excel)")
                        styled_tbl_f = (
                            metrics_tbl_f.style
                            .format({"%P":"{:.1f}%","%S":"{:.1f}%","%SA":"{:.1f}%","%OTHER":"{:.1f}%"})
                            .apply(style_percent_tables, id_col="Field", axis=None)
                            .hide(axis="index")
                        )
                        st.markdown(f"<div class='scroll-wrap-400'>{styled_tbl_f.to_html(escape=False)}</div>", unsafe_allow_html=True)
                    else:
                        idx_all = sorted(set(mod_agg_ps.index.tolist()) | set(mod_agg_tipo.index.tolist()))
                        p   = mod_agg_ps["P"].reindex(idx_all, fill_value=0.0)
                        s   = mod_agg_ps["S"].reindex(idx_all, fill_value=0.0)
                        sa  = mod_agg_tipo["SA"].reindex(idx_all, fill_value=0.0)
                        pa  = mod_agg_tipo["PA"].reindex(idx_all, fill_value=0.0)
                        sp  = mod_agg_tipo["SP"].reindex(idx_all, fill_value=0.0)
                        ip  = mod_agg_tipo["IP"].reindex(idx_all, fill_value=0.0)
                        oth = mod_agg_tipo["OTHER"].reindex(idx_all, fill_value=0.0)

                        totals = {
                            "P": float(p.sum()), "S": float(s.sum()),
                            "SA": float(sa.sum()), "PA": float(pa.sum()),
                            "SP": float(sp.sum()), "IP": float(ip.sum()),
                            "OTHER": float(oth.sum())
                        }

                        if objective_f == "%P":
                            main_col, aux_col = "P courses needed (3cr)", "Less S courses needed (3cr)"
                        elif objective_f == "%SA":
                            main_col, aux_col = "SA courses needed (3cr)", "Less other Qualific. courses needed (3cr)"
                        else:
                            main_col, aux_col = "Less OTHER Courses needed (3cr)", "More other Qualific. courses needed (3cr)"

                        rows = []
                        for label in idx_all:
                            Pv, Sv = float(p.get(label,0.0)), float(s.get(label,0.0))
                            SAv, PAv = float(sa.get(label,0.0)), float(pa.get(label,0.0))
                            SPv, IPv = float(sp.get(label,0.0)), float(ip.get(label,0.0))
                            OTv      = float(oth.get(label,0.0))

                            need1, need2 = _needed_pairs_for_obj(
                                objective_f, scope_label_f,
                                Pv, Sv, SAv, PAv, SPv, IPv, OTv, totals, credits_each=3.0
                            )
                            area_vals = {"P":Pv,"S":Sv,"SA":SAv,"PA":PAv,"SP":SPv,"IP":IPv,"OTHER":OTv}
                            up_pp, down_pp = _impact_pair(objective_f, area_vals, totals, scope_label_f, credits_each=3.0)

                            rows.append({
                                "Field": label,
                                main_col: int(need1),
                                aux_col:  int(need2),
                                "Impact increasing 1 course %p.p.": up_pp,
                                "Impact decreasing 1 course %p.p.": down_pp
                            })

                        need_tbl_f = pd.DataFrame(rows)

                        fmt_map_f = {}
                        if 'Field' in need_tbl_f.columns:
                            fmt_map_f['Field'] = '{}'
                        for col in ['Needed P (3cr)', 'Needed S less (3cr)', 'Needed SA (3cr)', 'Needed OTHER less (3cr)', 'Needed OTHER more (3cr)']:
                            if col in need_tbl_f.columns:
                                fmt_map_f[col] = '{:.0f}'
                        for col in ['Impact increasing 1 course %p.p.', 'Impact decreasing 1 course %p.p.']:
                            if col in need_tbl_f.columns:
                                fmt_map_f[col] = '{:+.2f}'

                        styled_f = (
                            need_tbl_f.style
                            .format(fmt_map_f)
                            .apply(_style_impact_heatmap, id_col="Field", axis=None)  # HEATMAP aplicado aquí
                            .hide(axis="index")
                        )

                        _download_xlsx_button(
                            need_tbl_f,
                            f"needed_ByField_{_slugify(sel_label)}_{_slugify(objective_f)}_{_slugify(scope_label_f)}.xlsx",
                            key=f"dl_need_field_{_slugify(sel_label)}_{_slugify(objective_f)}_{_slugify(scope_label_f)}",
                            label="⬇️ Download (Excel)"
                        )

                        st.markdown(styled_f.to_html(escape=False), unsafe_allow_html=True)

                # Históricos Field
                df_hist_f = df_car_global.copy()
                df_hist_f["_FIELD"] = df_hist_f[col_field].astype(str).str.strip()

                agg_ps_all_f = (df_hist_f.groupby(["_SEM","_FIELD","_PS"], dropna=False)["_CRED"].sum().unstack(fill_value=0.0))
                for k in ["P","S"]:
                    if k not in agg_ps_all_f.columns: agg_ps_all_f[k] = 0.0
                agg_ps_all_f["P_share"] = (agg_ps_all_f["P"] / (agg_ps_all_f["P"] + agg_ps_all_f["S"]).replace(0, pd.NA)) * 100
                agg_ps_all_f = agg_ps_all_f.reset_index()

                agg_tipo_all_f = (df_hist_f.groupby(["_SEM","_FIELD","_TIPO"], dropna=False)["_CRED"].sum().unstack(fill_value=0.0))
                for k in ["SA","PA","SP","IP","OTHER"]:
                    if k not in agg_tipo_all_f.columns: agg_tipo_all_f[k] = 0.0
                den_all_f = (agg_tipo_all_f[["SA","PA","SP","IP","OTHER"]].sum(axis=1)).replace(0, pd.NA)
                agg_tipo_all_f["SA_share"] = (agg_tipo_all_f["SA"] / den_all_f) * 100
                agg_tipo_all_f["OTHER_share"] = (agg_tipo_all_f["OTHER"] / den_all_f) * 100
                agg_tipo_all_f = agg_tipo_all_f.reset_index()

                tot_by_sem_tipo_f = (df_hist_f.groupby(["_SEM","_TIPO"])["_CRED"].sum().unstack(fill_value=0.0))
                for k in ["SA","PA","SP","IP","OTHER"]:
                    if k not in tot_by_sem_tipo_f.columns: tot_by_sem_tipo_f[k] = 0.0
                den_f_tot = (tot_by_sem_tipo_f[["SA","PA","SP","IP","OTHER"]].sum(axis=1)).replace(0, pd.NA)
                tot_by_sem_tipo_f["SA_share"] = (tot_by_sem_tipo_f["SA"] / den_f_tot) * 100
                tot_by_sem_tipo_f["OTHER_share"] = (tot_by_sem_tipo_f["OTHER"] / den_f_tot) * 100
                tot_by_sem_tipo_f = tot_by_sem_tipo_f.reset_index()

                tot_by_sem_f = (df_hist_f.groupby(["_SEM","_PS"])["_CRED"].sum().unstack(fill_value=0.0))
                for k in ["P","S"]:
                    if k not in tot_by_sem_f.columns: tot_by_sem_f[k] = 0.0
                tot_by_sem_f["P_share"] = (tot_by_sem_f["P"] / (tot_by_sem_f["P"] + tot_by_sem_f["S"]).replace(0, pd.NA)) * 100
                tot_by_sem_f = tot_by_sem_f.reset_index()

                agg_ps_all_tm = transform_for_time_mode_ps(agg_ps_all_f.rename(columns={"_FIELD":"__LEVEL__"})).rename(columns={"__LEVEL__":"_FIELD"})
                agg_tipo_sa_tm = transform_for_time_mode_tipo(agg_tipo_all_f.rename(columns={"_FIELD":"__LEVEL__"}), "SA_share").rename(columns={"__LEVEL__":"_FIELD"})
                agg_tipo_ot_tm = transform_for_time_mode_tipo(agg_tipo_all_f.rename(columns={"_FIELD":"__LEVEL__"}), "OTHER_share").rename(columns={"__LEVEL__":"_FIELD"})
                agg_tipo_all_tm = (
                    agg_tipo_sa_tm.drop(columns=[c for c in ["OTHER_share"] if c in agg_tipo_sa_tm], errors="ignore")
                    .merge(
                        agg_tipo_ot_tm[["_SEM","_FIELD","OTHER","SA","PA","SP","IP","OTHER_share"]],
                        on=["_SEM","_FIELD","SA","PA","SP","IP","OTHER"], how="outer"
                    )
                )
                tot_by_sem_P_tm = transform_for_time_mode_ps(tot_by_sem_f.copy())
                tot_tipo_sa_tm  = transform_for_time_mode_tipo(tot_by_sem_tipo_f.copy(), "SA_share")
                tot_tipo_ot_tm  = transform_for_time_mode_tipo(tot_by_sem_tipo_f.copy(), "OTHER_share")
                tot_by_sem_tipo_tm = (
                    tot_tipo_sa_tm.drop(columns=[c for c in ["OTHER_share"] if c in tot_tipo_sa_tm], errors="ignore")
                    .merge(
                        tot_tipo_ot_tm[["_SEM","SA","PA","SP","IP","OTHER","OTHER_share"]],
                        on=["_SEM","SA","PA","SP","IP","OTHER"], how="outer"
                    )
                )

                key_col, x_labels, x_map = build_time_axis_for_history(df_hist_f)
                if time_mode == "Semestral":
                    sel_x = x_map.get(str(sel_sem)) if sel_sem else None
                    sel_label_exact = str(sel_sem) if sel_sem else None
                elif time_mode == "Anual":
                    sel_x = x_map.get(sel_year) if sel_year is not None else None
                    sel_label_exact = sel_year
                else:
                    inter_label = f"{sel_year} Intersemestral" if sel_year else None
                    sel_x = x_map.get(inter_label) if inter_label else None
                    sel_label_exact = inter_label

                if SENS["on"] and SENS["ops"] and sel_label_exact is not None:
                    agg_ps_all_tm, agg_tipo_all_tm, tot_by_sem_P_tm, tot_by_sem_tipo_tm = apply_sensitivity_to_history(
                        agg_ps_all_tm, agg_tipo_all_tm, tot_by_sem_P_tm, tot_by_sem_tipo_tm,
                        level_name="_FIELD",
                        sel_label_value=sel_label_exact,
                        ops=SENS["ops"],
                        member_all_label="All"
                    )

                fields_all = sorted(set(_safe_unique_labels(agg_ps_all_tm, "_FIELD")) | set(_safe_unique_labels(agg_tipo_all_tm, "_FIELD")))
                with colF_R:
                    draw_history(
                        "Evolution by Academic Field",
                        level_name="_FIELD",
                        level_values=fields_all,
                        metric_kind="%P",
                        total_series_builders={"P": tot_by_sem_P_tm, "SA": tot_by_sem_tipo_tm, "OTHER": tot_by_sem_tipo_tm},
                        agg_ps_all=agg_ps_all_tm,
                        agg_tipo_all=agg_tipo_all_tm,
                        x_labels=x_labels, x_map=x_map, sel_x=sel_x
                    )
            # -------------- BY PROGRAM --------------
            elif view_mode == "By Program" and col_prog:
                colP_L, colP_R = st.columns([6,6], gap="large")

                # === Agregado actual (tabla de porcentajes y "needed") ===
                fil_prog = fil.copy()
                fil_prog["_PROG"] = fil_prog[col_prog].astype(str).str.strip()

                # Asegurar columnas base
                if "_TIPO" not in fil_prog.columns and col_tipoC:
                    fil_prog["_TIPO"] = _norm_str(fil_prog[col_tipoC]).map(normalize_tipo)
                if "_PS" not in fil_prog.columns and col_ps_C:
                    fil_prog["_PS"] = _norm_str(fil_prog[col_ps_C]).map(normalize_ps)
                if "_CRED" not in fil_prog.columns and col_cred:
                    fil_prog["_CRED"] = pd.to_numeric(fil_prog[col_cred], errors="coerce").fillna(0.0)

                # Aggregations por Programa
                agg_tipo_p = (
                    fil_prog.groupby(["_PROG","_TIPO"], dropna=False)["_CRED"]
                            .sum().unstack(fill_value=0.0)
                )
                for k in ["SA","PA","SP","IP","OTHER"]:
                    if k not in agg_tipo_p.columns: agg_tipo_p[k] = 0.0
                agg_tipo_p = agg_tipo_p[["SA","PA","SP","IP","OTHER"]]

                agg_ps_p = (
                    fil_prog.groupby(["_PROG","_PS"], dropna=False)["_CRED"]
                            .sum().unstack(fill_value=0.0)
                )
                for k in ["P","S"]:
                    if k not in agg_ps_p.columns: agg_ps_p[k] = 0.0
                agg_ps_p = agg_ps_p[["P","S"]]

                # Sensibilidad
                base_agg_ps_p   = agg_ps_p.copy()
                base_agg_tipo_p = agg_tipo_p.copy()
                if SENS.get("on") and SENS.get("ops"):
                    mod_agg_ps_p, mod_agg_tipo_p = apply_ops_to_aggs(base_agg_ps_p, base_agg_tipo_p, SENS["ops"])
                else:
                    mod_agg_ps_p, mod_agg_tipo_p = base_agg_ps_p, base_agg_tipo_p

                with colP_L:
                    needed_mode_p = False
                    if SENS.get("on"):
                        r1c1, r1c2, r1c3 = st.columns([1.8, 1.1, 1.6])
                        with r1c1:
                            needed_mode_p = st.toggle("Show necessary # of Faculty for…", value=False, key="prog_needed_mode")
                        if needed_mode_p:
                            with r1c2:
                                objective_p = st.selectbox("Objective", ["%P", "%SA", "%OTHER"], key="prog_objective")
                            with r1c3:
                                scope_label_p = st.radio("Target scope", ["By area", "Overall"], horizontal=True, key="prog_scope")
                        else:
                            objective_p   = st.session_state.get("prog_objective", "%P")
                            scope_label_p = st.session_state.get("prog_scope", "By area")
                    else:
                        objective_p   = st.session_state.get("prog_objective", "%P")
                        scope_label_p = st.session_state.get("prog_scope", "By area")

                    if not needed_mode_p:
                        # Tabla de % por Programa (sin botón de impacto y sin total)
                        metrics_tbl_p = build_percent_table("Program", mod_agg_tipo_p, mod_agg_ps_p)
                        _download_xlsx_button(metrics_tbl_p, f"table_ByProgram_{_slugify(sel_label)}.xlsx",
                                              key=f"dl_tbl_prog_{_slugify(sel_label)}", label="⬇️ Download table (Excel)")
                        styled_tbl_p = (
                            metrics_tbl_p.style
                            .format({"%P":"{:.1f}%","%S":"{:.1f}%","%SA":"{:.1f}%","%OTHER":"{:.1f}%"})
                            .apply(style_percent_tables, id_col="Program", axis=None)
                            .hide(axis="index")
                        )
                        st.markdown(f"<div class='scroll-wrap-400'>{styled_tbl_p.to_html(escape=False)}</div>", unsafe_allow_html=True)
                    else:
                        # Tabla "needed" + impacto (heatmap)
                        idx_all = sorted(set(mod_agg_ps_p.index.tolist()) | set(mod_agg_tipo_p.index.tolist()))
                        p   = mod_agg_ps_p["P"].reindex(idx_all, fill_value=0.0)
                        s   = mod_agg_ps_p["S"].reindex(idx_all, fill_value=0.0)
                        sa  = mod_agg_tipo_p["SA"].reindex(idx_all, fill_value=0.0)
                        pa  = mod_agg_tipo_p["PA"].reindex(idx_all, fill_value=0.0)
                        sp  = mod_agg_tipo_p["SP"].reindex(idx_all, fill_value=0.0)
                        ip  = mod_agg_tipo_p["IP"].reindex(idx_all, fill_value=0.0)
                        oth = mod_agg_tipo_p["OTHER"].reindex(idx_all, fill_value=0.0)

                        totals = {
                            "P": float(p.sum()), "S": float(s.sum()),
                            "SA": float(sa.sum()), "PA": float(pa.sum()),
                            "SP": float(sp.sum()), "IP": float(ip.sum()),
                            "OTHER": float(oth.sum())
                        }

                        if objective_p == "%P":
                            main_col, aux_col = "P courses needed (3cr)", "Less S courses needed (3cr)"
                        elif objective_p == "%SA":
                            main_col, aux_col = "SA courses needed (3cr)", "Less other Qualific. courses needed (3cr)"
                        else:
                            main_col, aux_col = "Less OTHER Courses needed (3cr)", "More other Qualific. courses needed (3cr)"

                        rows = []
                        for label in idx_all:
                            Pv, Sv   = float(p.get(label,0.0)),  float(s.get(label,0.0))
                            SAv, PAv = float(sa.get(label,0.0)), float(pa.get(label,0.0))
                            SPv, IPv = float(sp.get(label,0.0)), float(ip.get(label,0.0))
                            OTv      = float(oth.get(label,0.0))

                            need1, need2 = _needed_pairs_for_obj(
                                objective_p, scope_label_p,
                                Pv, Sv, SAv, PAv, SPv, IPv, OTv, totals, credits_each=3.0
                            )
                            area_vals = {"P":Pv,"S":Sv,"SA":SAv,"PA":PAv,"SP":SPv,"IP":IPv,"OTHER":OTv}
                            up_pp, down_pp = _impact_pair(objective_p, area_vals, totals, scope_label_p, credits_each=3.0)

                            rows.append({
                                "Program": label,
                                main_col: int(need1),
                                aux_col:  int(need2),
                                "Impact increasing 1 course %p.p.": up_pp,
                                "Impact decreasing 1 course %p.p.": down_pp
                            })

                        need_tbl_p = pd.DataFrame(rows)

                        # Formato dinámico según columnas presentes
                        fmt_map_p = {
                            "Program": "{}",
                            "Impact increasing 1 course %p.p.": "{:+.2f}",
                            "Impact decreasing 1 course %p.p.": "{:+.2f}"
                        }
                        if main_col in need_tbl_p.columns: fmt_map_p[main_col] = "{:.0f}"
                        if aux_col  in need_tbl_p.columns: fmt_map_p[aux_col]  = "{:.0f}"

                        styled_p = (
                            need_tbl_p.style
                            .format(fmt_map_p)
                            .apply(_style_impact_heatmap, id_col="Program", axis=None)
                            .hide(axis="index")
                        )
                        _download_xlsx_button(
                            need_tbl_p,
                            f"needed_ByProgram_{_slugify(sel_label)}_{_slugify(objective_p)}_{_slugify(scope_label_p)}.xlsx",
                            key=f"dl_need_prog_{_slugify(sel_label)}_{_slugify(objective_p)}_{_slugify(scope_label_p)}",
                            label="⬇️ Download (Excel)"
                        )
                        st.markdown(styled_p.to_html(escape=False), unsafe_allow_html=True)

                # ====== Series históricas por Program ======
                # Normalización previa (por si df_car_global no trae las columnas _SEM/_PROG/_PS/_TIPO/_CRED)
                df_hist = df_car_global.copy()

                semH  = _get_any(df_hist, "Semestre","Periodo","Periodo Académico","Periodo academico")
                psH   = _get_any(df_hist, "P/S","P - S","Participating/Supporting")
                credH = _get_any(df_hist, "Créditos","Creditos","Credits")
                tipoH = _get_any(df_hist, "TIPO","Tipo","Ranking","Tipo Ranking")
                progH = _get_any(df_hist, "Program","PROGRAM","Programa","Program Code","ProgramName")

                if "_SEM" not in df_hist.columns:
                    df_hist["_SEM"] = df_hist[semH].astype(str).str.strip() if semH else ""
                else:
                    df_hist["_SEM"] = df_hist["_SEM"].astype(str).str.strip()

                if "_PS" not in df_hist.columns:
                    df_hist["_PS"] = _norm_str(df_hist[psH]).map(normalize_ps) if psH else ""
                else:
                    df_hist["_PS"] = _norm_str(df_hist["_PS"]).map(normalize_ps)

                if "_CRED" not in df_hist.columns:
                    df_hist["_CRED"] = pd.to_numeric(df_hist[credH], errors="coerce").fillna(0.0) if credH else 0.0
                else:
                    df_hist["_CRED"] = pd.to_numeric(df_hist["_CRED"], errors="coerce").fillna(0.0)

                if "_TIPO" not in df_hist.columns:
                    df_hist["_TIPO"] = _norm_str(df_hist[tipoH]).map(normalize_tipo) if tipoH else "OTHER"
                else:
                    df_hist["_TIPO"] = _norm_str(df_hist["_TIPO"]).map(normalize_tipo)

                if "_PROG" not in df_hist.columns:
                    if progH:
                        df_hist["_PROG"] = df_hist[progH].astype(str).str.strip().replace({"": "N/A"})
                    elif "_MAT" in df_hist.columns:
                        df_hist["_PROG"] = df_hist["_MAT"].astype(str).str.strip().replace({"": "N/A"})
                    else:
                        df_hist["_PROG"] = "N/A"
                else:
                    df_hist["_PROG"] = df_hist["_PROG"].astype(str).str.strip().replace({"": "N/A"})
                df_hist["_PROG"] = df_hist["_PROG"].fillna("N/A")

                # Agregaciones
                agg_ps_all_p = (
                    df_hist.groupby(["_SEM","_PROG","_PS"], dropna=False)["_CRED"]
                           .sum().unstack(fill_value=0.0)
                )
                for k in ["P","S"]:
                    if k not in agg_ps_all_p.columns: agg_ps_all_p[k] = 0.0
                agg_ps_all_p["P_share"] = (agg_ps_all_p["P"] / (agg_ps_all_p["P"] + agg_ps_all_p["S"]).replace(0, pd.NA)) * 100
                agg_ps_all_p = agg_ps_all_p.reset_index()

                agg_tipo_all_p = (
                    df_hist.groupby(["_SEM","_PROG","_TIPO"], dropna=False)["_CRED"]
                           .sum().unstack(fill_value=0.0)
                )
                for k in ["SA","PA","SP","IP","OTHER"]:
                    if k not in agg_tipo_all_p.columns: agg_tipo_all_p[k] = 0.0
                den_all_p = (agg_tipo_all_p[["SA","PA","SP","IP","OTHER"]].sum(axis=1)).replace(0, pd.NA)
                agg_tipo_all_p["SA_share"]    = (agg_tipo_all_p["SA"]    / den_all_p) * 100
                agg_tipo_all_p["OTHER_share"] = (agg_tipo_all_p["OTHER"] / den_all_p) * 100
                agg_tipo_all_p = agg_tipo_all_p.reset_index()

                tot_by_sem_P_p = (
                    df_hist.groupby(["_SEM","_PS"])["_CRED"]
                           .sum().unstack(fill_value=0.0)
                )
                for k in ["P","S"]:
                    if k not in tot_by_sem_P_p.columns: tot_by_sem_P_p[k] = 0.0
                tot_by_sem_P_p["P_share"] = (tot_by_sem_P_p["P"] / (tot_by_sem_P_p["P"] + tot_by_sem_P_p["S"]).replace(0, pd.NA)) * 100
                tot_by_sem_P_p = tot_by_sem_P_p.reset_index()

                tot_by_sem_tipo_p = (
                    df_hist.groupby(["_SEM","_TIPO"])["_CRED"]
                           .sum().unstack(fill_value=0.0)
                )
                for k in ["SA","PA","SP","IP","OTHER"]:
                    if k not in tot_by_sem_tipo_p.columns: tot_by_sem_tipo_p[k] = 0.0
                den_tot_p = (tot_by_sem_tipo_p[["SA","PA","SP","IP","OTHER"]].sum(axis=1)).replace(0, pd.NA)
                tot_by_sem_tipo_p["SA_share"]    = (tot_by_sem_tipo_p["SA"]    / den_tot_p) * 100
                tot_by_sem_tipo_p["OTHER_share"] = (tot_by_sem_tipo_p["OTHER"] / den_tot_p) * 100
                tot_by_sem_tipo_p = tot_by_sem_tipo_p.reset_index()

                # Adaptación a modo temporal
                agg_ps_all_p_tm  = transform_for_time_mode_ps(agg_ps_all_p.rename(columns={"_PROG":"__LEVEL__"})).rename(columns={"__LEVEL__":"_PROG"})
                agg_tipo_sa_p_tm = transform_for_time_mode_tipo(agg_tipo_all_p.rename(columns={"_PROG":"__LEVEL__"}), "SA_share").rename(columns={"__LEVEL__":"_PROG"})
                agg_tipo_ot_p_tm = transform_for_time_mode_tipo(agg_tipo_all_p.rename(columns={"_PROG":"__LEVEL__"}), "OTHER_share").rename(columns={"__LEVEL__":"_PROG"})
                agg_tipo_all_p_tm = (
                    agg_tipo_sa_p_tm.drop(columns=[c for c in ["OTHER_share"] if c in agg_tipo_sa_p_tm], errors="ignore")
                    .merge(agg_tipo_ot_p_tm[["_SEM","_PROG","OTHER","SA","PA","SP","IP","OTHER_share"]],
                           on=["_SEM","_PROG","SA","PA","SP","IP","OTHER"], how="outer")
                )
                tot_by_sem_P_p_tm = transform_for_time_mode_ps(tot_by_sem_P_p.copy())
                tot_tipo_sa_p_tm  = transform_for_time_mode_tipo(tot_by_sem_tipo_p.copy(), "SA_share")
                tot_tipo_ot_p_tm  = transform_for_time_mode_tipo(tot_by_sem_tipo_p.copy(), "OTHER_share")
                tot_by_sem_tipo_p_tm = (
                    tot_tipo_sa_p_tm.drop(columns=[c for c in ["OTHER_share"] if c in tot_tipo_sa_p_tm], errors="ignore")
                    .merge(tot_tipo_ot_p_tm[["_SEM","SA","PA","SP","IP","OTHER","OTHER_share"]],
                           on=["_SEM","SA","PA","SP","IP","OTHER"], how="outer")
                )

                key_col_p, x_labels_p, x_map_p = build_time_axis_for_history(df_hist)
                if time_mode == "Semestral":
                    sel_x_p = x_map_p.get(str(sel_sem)) if sel_sem else None
                    sel_label_exact_p = str(sel_sem) if sel_sem else None
                elif time_mode == "Anual":
                    sel_x_p = x_map_p.get(sel_year) if sel_year is not None else None
                    sel_label_exact_p = sel_year
                else:
                    inter_label_p = f"{sel_year} Intersemestral" if sel_year else None
                    sel_x_p = x_map_p.get(inter_label_p) if inter_label_p else None
                    sel_label_exact_p = inter_label_p

                if SENS.get("on") and SENS.get("ops") and sel_label_exact_p is not None:
                    agg_ps_all_p_tm, agg_tipo_all_p_tm, tot_by_sem_P_p_tm, tot_by_sem_tipo_p_tm = apply_sensitivity_to_history(
                        agg_ps_all_p_tm, agg_tipo_all_p_tm, tot_by_sem_P_p_tm, tot_by_sem_tipo_p_tm,
                        level_name="_PROG",
                        sel_label_value=sel_label_exact_p,
                        ops=SENS["ops"],
                        member_all_label="All"
                    )

                progs_all = sorted(set(_safe_unique_labels(agg_ps_all_p_tm, "_PROG")) |
                                   set(_safe_unique_labels(agg_tipo_all_p_tm, "_PROG")))
                with colP_R:
                    draw_history(
                        "Evolution by Program",
                        level_name="_PROG",
                        level_values=progs_all,
                        metric_kind="%P",
                        total_series_builders={"P": tot_by_sem_P_p_tm, "SA": tot_by_sem_tipo_p_tm, "OTHER": tot_by_sem_tipo_p_tm},
                        agg_ps_all=agg_ps_all_p_tm,
                        agg_tipo_all=agg_tipo_all_p_tm,
                        x_labels=x_labels_p, x_map=x_map_p, sel_x=sel_x_p
                    )

    # --------------------------
    # CREDIT SUMS (EXPANDER)
    # --------------------------
    try:
        period_df = df_car_filt_all.copy()
        if "_CRED"  not in period_df.columns and col_cred:  period_df["_CRED"]  = pd.to_numeric(period_df[col_cred], errors="coerce").fillna(0.0)
        if "_PS"    not in period_df.columns and col_ps_C:  period_df["_PS"]    = _norm_str(period_df[col_ps_C]).map(normalize_ps)
        if "_TIPO"  not in period_df.columns and col_tipoC: period_df["_TIPO"]  = _norm_str(period_df[col_tipoC]).map(normalize_tipo)
        if "_AREA"  not in period_df.columns and col_areaCourse: period_df["_AREA"] = period_df[col_areaCourse].astype(str).str.strip()
        if "_FIELD" not in period_df.columns and col_field:      period_df["_FIELD"] = period_df[col_field].astype(str).str.strip()
        if "_PROG"  not in period_df.columns and col_prog:       period_df["_PROG"] = period_df[col_prog].astype(str).str.strip()

        view = st.session_state.view_mode if "view_mode" in st.session_state else "By Academic Area"
        if view == "By Academic Area":
            dim_col, dim_label = "_AREA", "Academic Area"
        elif view == "By Field":
            dim_col, dim_label = "_FIELD", "Field"
        else:
            dim_col, dim_label = "_PROG", "Program"

        if dim_col in period_df.columns:
            base_index = period_df.groupby(dim_col)["_CRED"].sum().sort_values(ascending=False)
            idx = base_index.index

            sum_total = base_index.rename("Credit Sum")
            sum_P  = (period_df[period_df["_PS"]   == "P"     ].groupby(dim_col)["_CRED"].sum().reindex(idx, fill_value=0.0)).rename("P Sum")
            sum_S  = (period_df[period_df["_PS"]   == "S"     ].groupby(dim_col)["_CRED"].sum().reindex(idx, fill_value=0.0)).rename("S Sum")
            sum_SA = (period_df[period_df["_TIPO"] == "SA"    ].groupby(dim_col)["_CRED"].sum().reindex(idx, fill_value=0.0)).rename("SA Sum")
            sum_PA = (period_df[period_df["_TIPO"] == "PA"    ].groupby(dim_col)["_CRED"].sum().reindex(idx, fill_value=0.0)).rename("PA Sum")
            sum_SP = (period_df[period_df["_TIPO"] == "SP"    ].groupby(dim_col)["_CRED"].sum().reindex(idx, fill_value=0.0)).rename("SP Sum")
            sum_IP = (period_df[period_df["_TIPO"] == "IP"    ].groupby(dim_col)["_CRED"].sum().reindex(idx, fill_value=0.0)).rename("IP Sum")
            sum_OT = (period_df[period_df["_TIPO"] == "OTHER" ].groupby(dim_col)["_CRED"].sum().reindex(idx, fill_value=0.0)).rename("OTHER Sum")

            tbl = pd.concat([sum_total, sum_P, sum_S, sum_SA, sum_PA, sum_SP, sum_IP, sum_OT], axis=1).fillna(0.0)

            if SENS.get("on"):
                agg_tipo = (period_df.groupby([dim_col,"_TIPO"], dropna=False)["_CRED"].sum().unstack(fill_value=0.0))
                for k in ["SA","PA","SP","IP","OTHER"]:
                    if k not in agg_tipo.columns: agg_tipo[k] = 0.0
                agg_ps = (period_df.groupby([dim_col,"_PS"], dropna=False)["_CRED"].sum().unstack(fill_value=0.0))
                for k in ["P","S"]:
                    if k not in agg_ps.columns: agg_ps[k] = 0.0
                agg_ps = agg_ps[["P","S"]]; agg_tipo = agg_tipo[["SA","PA","SP","IP","OTHER"]]

                mod_ps, mod_tipo = apply_ops_to_aggs(agg_ps, agg_tipo, SENS.get("ops", []), member_all_label="All")
                tbl["P Sum"]     = mod_ps["P"].reindex(tbl.index, fill_value=0.0)
                tbl["S Sum"]     = mod_ps["S"].reindex(tbl.index, fill_value=0.0)
                tbl["SA Sum"]    = mod_tipo["SA"].reindex(tbl.index, fill_value=0.0)
                tbl["PA Sum"]    = mod_tipo["PA"].reindex(tbl.index, fill_value=0.0)
                tbl["SP Sum"]    = mod_tipo["SP"].reindex(tbl.index, fill_value=0.0)
                tbl["IP Sum"]    = mod_tipo["IP"].reindex(tbl.index, fill_value=0.0)
                tbl["OTHER Sum"] = mod_tipo["OTHER"].reindex(tbl.index, fill_value=0.0)
                tbl["Credit Sum"]= tbl[["P Sum","S Sum"]].sum(axis=1)

            total_row = pd.DataFrame(tbl.sum(axis=0)).T
            total_row.index = ["TOTAL"]
            tbl_out = pd.concat([tbl, total_row], axis=0)

            display_label = st.session_state.get('sel_label','Selected Period')
            with st.expander(f"Credit sums by {dim_label}", expanded=False):
                export_tbl = tbl_out.reset_index().rename(columns={"index": dim_label})
                _download_xlsx_button(export_tbl,
                                      f"credit_sums_{_slugify(dim_label)}_{_slugify(display_label)}.xlsx",
                                      key=f"dl_credit_sums_{_slugify(dim_label)}_{_slugify(display_label)}",
                                      label=f"⬇️ Download table {display_label} (Excel)")
                _tbl_out_display = tbl_out.reset_index().rename(columns={"index": dim_label})
                st.dataframe(_tbl_out_display.style.format("{:,.0f}", subset=tbl_out.columns), use_container_width=True, hide_index=True)

                # ===== Selector propio de dimensión (independiente del gráfico superior) =====
                # Construye lista de miembros visibles en esta tabla
                members = [str(x) for x in tbl.index.tolist() if str(x) != "TOTAL"]
                members_sorted = sorted(set(members))
                dim_options = ["(All)", "(TOTAL)"] + members_sorted
                dim_opt = st.selectbox(
                    f"Select {dim_label} for the evolution lines",
                    dim_options,
                    index=1,  # por defecto "(TOTAL)"
                    key=f"credit_dim_selector_{dim_col}"
                )
                # Nota: para evitar 5×N líneas, si eligen "(All)" mostramos TOTAL
                if dim_opt == "(All)":
                    dim_opt_eff = "(TOTAL)"
                else:
                    dim_opt_eff = dim_opt

                # ===== Toggle de series: Qualifications ↔ P/S =====
                mode_line = st.radio(
                    "",
                    ["Qualifications", "P/S"],
                    horizontal=True,
                    key=f"credit_line_mode_{dim_col}"
                )

                # --- histórico base normalizado ---
                df_hist = df_car_global.copy()
                if "_CRED" not in df_hist.columns and col_cred:
                    df_hist["_CRED"] = pd.to_numeric(df_hist[col_cred], errors="coerce").fillna(0.0)
                if "_TIPO" not in df_hist.columns and col_tipoC:
                    df_hist["_TIPO"] = _norm_str(df_hist[col_tipoC]).map(normalize_tipo)
                if "_PS" not in df_hist.columns and col_ps_C:
                    df_hist["_PS"] = _norm_str(df_hist[col_ps_C]).map(normalize_ps)
                if "_SEM" not in df_hist.columns:
                    sc = _get_any(df_hist, "Semestre","Periodo","Periodo Académico","Periodo academico")
                    df_hist["_SEM"] = df_hist[sc].astype(str).str.strip() if sc else ""

                # columna de dimensión si falta
                if dim_col == "_AREA" and "_AREA" not in df_hist.columns and col_areaCourse:
                    df_hist["_AREA"] = df_hist[col_areaCourse].astype(str).str.strip()
                if dim_col == "_FIELD" and "_FIELD" not in df_hist.columns and col_field:
                    df_hist["_FIELD"] = df_hist[col_field].astype(str).str.strip()
                if dim_col == "_PROG" and "_PROG" not in df_hist.columns and col_prog:
                    df_hist["_PROG"] = df_hist[col_prog].astype(str).str.strip()

                # filtro por miembro elegido en este selector
                if dim_opt_eff != "(TOTAL)" and dim_col in df_hist.columns:
                    df_hist = df_hist[df_hist[dim_col].astype(str).str.strip() == str(dim_opt_eff)]

                # --- agregaciones base ---
                cats_qual = ["SA","PA","SP","IP","OTHER"]

                agg_tipo = (
                    df_hist.groupby(["_SEM","_TIPO"], dropna=False)["_CRED"]
                    .sum().unstack(fill_value=0.0)
                )
                for k in cats_qual:
                    if k not in agg_tipo.columns:
                        agg_tipo[k] = 0.0
                agg_tipo = agg_tipo[cats_qual].reset_index()

                agg_ps = (
                    df_hist.groupby(["_SEM","_PS"], dropna=False)["_CRED"]
                    .sum().unstack(fill_value=0.0)
                )
                for k in ["P","S"]:
                    if k not in agg_ps.columns:
                        agg_ps[k] = 0.0
                agg_ps = agg_ps[["P","S"]].reset_index()

                # --- adaptar a modo temporal (sumas) ---
                tm = st.session_state.get("time_mode", "Semestral")
                def adapt_time_sum(df_in: pd.DataFrame, value_cols: list[str]) -> pd.DataFrame:
                    tmp = df_in.copy()
                    tmp["_YEAR"] = tmp["_SEM"].map(extract_year_from_period)
                    tmp["_INTER_LABEL"] = tmp["_SEM"].map(lambda s: f"{extract_year_from_period(s)} Intersemestral" if "inter" in str(s).lower() else None)
                    if tm == "Semestral":
                        out = tmp.rename(columns={"_SEM":"_X"})
                    elif tm == "Anual":
                        out = tmp.groupby("_YEAR", dropna=False)[value_cols].sum().reset_index().rename(columns={"_YEAR":"_X"})
                    else:
                        inter_only = tmp[~tmp["_INTER_LABEL"].isna()].copy()
                        out = inter_only.groupby("_INTER_LABEL", dropna=False)[value_cols].sum().reset_index().rename(columns={"_INTER_LABEL":"_X"})
                    return out

                plot_qual = adapt_time_sum(agg_tipo, cats_qual)
                plot_ps   = adapt_time_sum(agg_ps, ["P","S"])

                # --- eje X consistente ---
                def build_axis(df_x: pd.DataFrame) -> tuple[list, dict]:
                    if tm == "Semestral":
                        x_labels = sorted(
                            {x for x in df_x["_X"].dropna().astype(str) if period_suffix(x) in {"10","20"}},
                            key=_period_sort_key
                        )
                    elif tm == "Anual":
                        x_labels = sorted({int(x) for x in df_x["_X"].dropna()}, key=int)
                    else:
                        x_labels = sorted(
                            df_x["_X"].dropna().astype(str).unique().tolist(),
                            key=lambda s: int(str(s).split()[0]) if str(s).split() else 0
                        )
                    x_map = {lab: i for i, lab in enumerate(x_labels)}
                    return x_labels, x_map

                x_labels_q, x_map_q = build_axis(plot_qual)
                x_labels_ps, x_map_ps = build_axis(plot_ps)
                plot_qual["_xi"] = plot_qual["_X"].map(x_map_q)
                plot_ps["_xi"]   = plot_ps["_X"].map(x_map_ps)
                plot_qual = plot_qual.sort_values("_xi")
                plot_ps   = plot_ps.sort_values("_xi")

                # --- sensibilidad SOLO en el período seleccionado ---
                if tm == "Semestral":
                    sel_label_exact = st.session_state.get("sel_sem")  # 'YYYY10' / 'YYYY20' (str)
                elif tm == "Anual":
                    sel_label_exact = st.session_state.get("sel_year")  # año (int)
                else:  # Intersemestral
                    y = st.session_state.get("sel_year")                # usa el MISMO sel_year
                    sel_label_exact = f"{y} Intersemestral" if y else None  # 'YYYY Intersemestral' (str)

                if SENS.get("on") and SENS.get("ops") and sel_label_exact is not None:
                    # Qualifications
                    sens_tipo = plot_qual[["_X"] + cats_qual].rename(columns={"_X":"_SEM"}).copy()
                    dummy_ps = pd.DataFrame({"_SEM": sens_tipo["_SEM"]})
                    sens_tipo2, _A, _B, _C = apply_sensitivity_to_history(
                        agg_ps_tm=dummy_ps, agg_tipo_tm=sens_tipo,
                        tot_ps_tm=dummy_ps.copy(), tot_tipo_tm=sens_tipo.copy(),
                        level_name="_SEM", sel_label_value=sel_label_exact,
                        ops=SENS["ops"], member_all_label="All"
                    )
                    plot_qual[cats_qual] = sens_tipo2[cats_qual].values

                    # P/S
                    sens_ps = plot_ps[["_X","P","S"]].rename(columns={"_X":"_SEM"}).copy()
                    dummy_tipo = pd.DataFrame({"_SEM": sens_ps["_SEM"], "SA":0.0,"PA":0.0,"SP":0.0,"IP":0.0,"OTHER":0.0})
                    sens_ps2, _tq, _tps, _ttq = apply_sensitivity_to_history(
                        agg_ps_tm=sens_ps, agg_tipo_tm=dummy_tipo,
                        tot_ps_tm=sens_ps.copy(), tot_tipo_tm=dummy_tipo.copy(),
                        level_name="_SEM", sel_label_value=sel_label_exact,
                        ops=SENS["ops"], member_all_label="All"
                    )
                    plot_ps[["P","S"]] = sens_ps2[["P","S"]].values

                # --- dibujar ---
                if mode_line == "Qualifications":
                    COL_SA = "#1FA89B"  # menta verdoso
                    COL_PA = "#232D3C"  # verde apagado
                    COL_SP = "#565656"  # azul grisoso
                    COL_IP = "#8F8F8F"  # gris
                    COL_OT = "#A13B3B"  # rojo claro
                    cmap = {"SA":COL_SA, "PA":COL_PA, "SP":COL_SP, "IP":COL_IP, "OTHER":COL_OT}

                    fig = go.Figure()
                    for k in ["SA","PA","SP","IP","OTHER"]:
                        fig.add_trace(go.Scatter(
                            x=plot_qual["_xi"], y=plot_qual[k],
                            mode="lines+markers",
                            name=k,
                            line=dict(width=2, color=cmap[k]),
                            marker=dict(size=6, color=cmap[k]),
                            hovertemplate=f"{k}<br>%{{y:.0f}} cr<extra></extra>"
                        ))

                    sel_x  = x_map_q.get(sel_label_exact)  if (sel_label_exact is not None) else None
                    if sel_x is not None:
                        fig.add_vrect(x0=sel_x-0.5, x1=sel_x+0.5, fillcolor="#E8FAF7", opacity=0.5, layer="below", line_width=0)

                    tickvals = list(range(len(x_labels_q)))
                    ticktext = [str(x) for x in x_labels_q]
                    fig.update_layout(
                        title=f"Evolution of Credits — Qualifications ({dim_opt_eff})",
                        margin=dict(l=10,r=10,t=40,b=60),
                        legend=dict(orientation="h", y=-0.2, yanchor="top", x=0.5, xanchor="center"),
                    )
                    fig.update_xaxes(title=None, tickmode="array", tickvals=tickvals, ticktext=ticktext)
                    fig.update_yaxes(title="Credits", rangemode="tozero")
                    st.plotly_chart(fig, use_container_width=True)

                else:
                    COL_P = "#1FA89B"  # P (verde menta)
                    COL_S = "#9E9E9E"  # S (gris)

                    fig2 = go.Figure()
                    fig2.add_trace(go.Scatter(
                        x=plot_ps["_xi"], y=plot_ps["P"],
                        mode="lines+markers",
                        name="P",
                        line=dict(width=2, color=COL_P),
                        marker=dict(size=6, color=COL_P),
                        hovertemplate="P<br>%{y:.0f} cr<extra></extra>"
                    ))
                    fig2.add_trace(go.Scatter(
                        x=plot_ps["_xi"], y=plot_ps["S"],
                        mode="lines+markers",
                        name="S",
                        line=dict(width=2, color=COL_S),
                        marker=dict(size=6, color=COL_S),
                        hovertemplate="S<br>%{y:.0f} cr<extra></extra>"
                    ))

                    sel_x2 = x_map_ps.get(sel_label_exact) if (sel_label_exact is not None) else None
                    if sel_x2 is not None:
                        fig2.add_vrect(x0=sel_x2-0.5, x1=sel_x2+0.5, fillcolor="#E8FAF7", opacity=0.5, layer="below", line_width=0)

                    tickvals2 = list(range(len(x_labels_ps)))
                    ticktext2 = [str(x) for x in x_labels_ps]
                    fig2.update_layout(
                        title=f"Evolution of Credits — P/S ({dim_opt_eff})",
                        margin=dict(l=10,r=10,t=40,b=60),
                        legend=dict(orientation="h", y=-0.2, yanchor="top", x=0.5, xanchor="center"),
                    )
                    fig2.update_xaxes(title=None, tickmode="array", tickvals=tickvals2, ticktext=ticktext2)
                    fig2.update_yaxes(title="Credits", rangemode="tozero")
                    st.plotly_chart(fig2, use_container_width=True)

    except Exception:
        # Evita romper la app si algo falla en este bloque
        pass

    # HELPERS (únicos en este módulo, sin duplicados)
    def _extract_year(s):
        m = re.search(r"(19|20)\d{2}", str(s) if s is not None else "")
        return int(m.group(0)) if m else None

    def _normalize_sem_str(x: str) -> str:
        return str(x).strip().replace("\xa0", " ")

    def _ensure_pid(df: pd.DataFrame) -> pd.DataFrame:
        """Agrega _PID (persona) usando ID; si no hay, usa nombre; si no, índice."""
        out = df.copy()
        idc   = _get_any(out, "ID","ID Nr.","Documento")
        namec = _get_any(out, "Profesor","PROFESOR","Docente","Nombre")
        if idc and idc in out:
            out["_PID"] = out[idc].astype(str).str.strip()
        elif namec and namec in out:
            out["_PID"] = out[namec].astype(str).str.strip().str.lower()
        else:
            out["_PID"] = out.index.astype(str)
        return out

    def _norm_gender(x: str) -> str:
        v = str(x).strip().lower()
        if v in {"male","masculino","m","hombre"}:   return "Male"
        if v in {"female","femenino","f","mujer"}:   return "Female"
        return "Other"

    def _is_doctoral(x: str) -> bool:
        v = str(x).strip().lower().replace(".", "")
        return ("phd" in v) or ("doctor" in v)

    def _norm_ftpt(x: str) -> str:
        v = str(x).strip().upper()
        if "PLANTA"  in v: return "PLANTA"
        if "CATEDRA" in v or "CÁTEDRA" in v: return "CÁTEDRA"
        return ""

    def _first_map(df_, key_col, val_col):
        if key_col not in df_ or val_col not in df_:
            return {}
        tmp = df_[[key_col, val_col]].dropna()
        return tmp.drop_duplicates(subset=[key_col]).set_index(key_col)[val_col].to_dict()

    def _pick(df_, *cands):
        c = _get_any(df_, *cands)
        return df_[c] if c else pd.Series([None]*len(df_), index=df_.index)

    # ---------- Parser robusto de periodos ----------
    # Detecta: "YYYY", "YYYY 10", "YYYY 20", "YYYY Intersemestral"
    _SEM_RE = re.compile(r'^\s*(?P<y>(?:19|20)\d{2})\s*(?P<t>10|20|Intersemestral)?\s*$', re.IGNORECASE)

    def _year_term(x):
        s = _normalize_sem_str(x)
        m = _SEM_RE.match(s)
        if not m:
            return None, None, s
        year = int(m.group('y'))
        t = m.group('t')
        if not t:
            term = None
        else:
            term = 'INTER' if t.lower().startswith('inter') else t  # "10", "20", "INTER"
        return year, term, s

    # ---------- Filtros de alcance (Distribution / Cartelera) ----------
    def filter_df_fd(df_fd_base: pd.DataFrame, time_mode: str, sel_year: int | None, sel_sem_code: str | int | None) -> pd.DataFrame:
        """Filtra Faculty Distribution según alcance temporal actual (Semestral/Intersemestral/Anual)."""
        if df_fd_base is None or df_fd_base.empty:
            return pd.DataFrame()
        df = df_fd_base.copy()
        semc_fd = _get_any(df, "Semestre","Periodo","Periodo Académico","Periodo academico")
        if not semc_fd:
            return df

        yt = df[semc_fd].map(_year_term)
        df["_YEARX"]  = yt.map(lambda z: z[0]).astype("Int64")
        df["_TERM"]   = yt.map(lambda z: z[1])   # "10" | "20" | "INTER" | None
        df["_SEM_SRC"]= yt.map(lambda z: z[2])   # normalizado

        if time_mode == "Semestral" and sel_sem_code is not None:
            goal = _normalize_sem_str(sel_sem_code)
            return df[df["_SEM_SRC"].eq(goal)].copy()

        if time_mode == "Intersemestral" and sel_year is not None:
            yr = int(sel_year)
            return df[(df["_YEARX"] == yr) & (df["_TERM"].eq("INTER"))].copy()

        if time_mode == "Anual" and sel_year is not None:
            yr = int(sel_year)
            return df[df["_YEARX"] == yr].copy()

        return df

    def filter_df_car(df_car_base: pd.DataFrame, time_mode: str, sel_year: int | None, sel_sem_code: str | int | None) -> pd.DataFrame:
        """Filtra Cartelera según alcance temporal actual (Semestral/Intersemestral/Anual)."""
        if df_car_base is None or df_car_base.empty:
            return pd.DataFrame()
        df = df_car_base.copy()
        semc = _get_any(df, "Semestre","Periodo","Periodo Académico","Periodo academico")
        if not semc:
            return df

        yt = df[semc].map(_year_term)
        df["_YEARX"]  = yt.map(lambda z: z[0]).astype("Int64")
        df["_TERM"]   = yt.map(lambda z: z[1])
        df["_SEM_SRC"]= yt.map(lambda z: z[2])

        if time_mode == "Semestral" and sel_sem_code is not None:
            goal = _normalize_sem_str(sel_sem_code)
            return df[df["_SEM_SRC"].eq(goal)].copy()

        if time_mode == "Intersemestral" and sel_year is not None:
            yr = int(sel_year)
            return df[(df["_YEARX"] == yr) & (df["_TERM"].eq("INTER"))].copy()

        if time_mode == "Anual" and sel_year is not None:
            yr = int(sel_year)
            return df[df["_YEARX"] == yr].copy()

        return df

    # --------------------------
    # DETAIL TABLE + DONUT  (no search)
    # (Hidden automatically when Sensitivity mode is active)
    # --------------------------
    if not SENS.get("on", False):
        try:
            # ---------- View config ----------
            cfg = {
                "By Academic Area": {"key": "_AREA_filter",  "col": "_AREA",  "label": "area",    "metric_key": "metric__AREA"},
                "By Field":         {"key": "_FIELD_filter", "col": "_FIELD", "label": "field",   "metric_key": "metric__FIELD"},
                "By Program":       {"key": "_PROG_filter",  "col": "_PROG",  "label": "program", "metric_key": "metric__PROG"},
            }
            view = st.session_state.view_mode

            if view in cfg:
                key        = cfg[view]["key"]
                col_tag    = cfg[view]["col"]
                metric_key = cfg[view]["metric_key"]
                metric_choice = st.session_state.get(metric_key, "%P")
                opt_val    = st.session_state.get(key, "(All)")

                # ---------- Minimal enriched Cartelera base (already time-filtered above) ----------
                base = df_car_filt_all.copy()
                if "_AREA"  not in base.columns and col_areaCourse: base["_AREA"]  = base[col_areaCourse].astype(str).str.strip()
                if "_FIELD" not in base.columns and col_field:      base["_FIELD"] = base[col_field].astype(str).str.strip()
                if "_PROG"  not in base.columns and col_prog:       base["_PROG"]  = base[col_prog].astype(str).str.strip()
                if "_TIPO"  not in base.columns and col_tipoC:      base["_TIPO"]  = _norm_str(base[col_tipoC]).map(normalize_tipo)
                if "_PS"    not in base.columns and col_ps_C:       base["_PS"]    = _norm_str(base[col_ps_C]).map(normalize_ps)
                if "_CRED"  not in base.columns and col_cred:       base["_CRED"]  = pd.to_numeric(base[col_cred], errors="coerce").fillna(0.0)

                # Safe default for mint if not globally set
                MINT = globals().get("MINT", "#2DD4BF")
                MINT_DIAMOND = "#D6FFF2"  # lighter mint for highlights

                cL, cR = st.columns([7,5], gap="large")

                # LEFT: Filters + Detail Table (+ popup trigger)
                with cL:
                    # --- Filters (always visible) ---
                    colF1, colF2 = st.columns([1,1])
                    with colF1:
                        table_filter_ps = st.radio(
                            "Filter by P/S",
                            ["All", "Only P", "Only S"],
                            index=0, horizontal=True,
                            key=f"table_filt_ps_{view}_{opt_val}"
                        )
                    with colF2:
                        table_filter_tipo = st.radio(
                            "Filter by Qualification",
                            ["All", "Only SA", "Only OTHER"],
                            index=0, horizontal=True,
                            key=f"table_filt_tipo_{view}_{opt_val}"
                        )

                    # --- Scope by selected tag (Area/Field/Program) for the TABLE ONLY ---
                    base_tbl = base.copy()
                    if opt_val not in {"(All)", "(TOTAL)"} and col_tag in base_tbl.columns:
                        base_tbl = base_tbl[base_tbl[col_tag] == opt_val].copy()

                    # --- Apply P/S filter (TABLE ONLY) ---
                    if table_filter_ps == "Only P":
                        base_tbl = base_tbl[base_tbl["_PS"] == "P"]
                    elif table_filter_ps == "Only S":
                        base_tbl = base_tbl[base_tbl["_PS"] == "S"]

                    # --- Apply Qualification filter (TABLE ONLY) ---
                    if table_filter_tipo == "Only SA":
                        base_tbl = base_tbl[base_tbl["_TIPO"] == "SA"]
                    elif table_filter_tipo == "Only OTHER":
                        base_tbl = base_tbl[base_tbl["_TIPO"] == "OTHER"]

                    # ---------- Popup function (faculty by course count) ----------
                    def _show_faculty_popup(df_in, display_label, opt_val_local):
                        # Wider dialog (white box) so the table fits

                        # Helper: most frequent value per professor (for _PS and _TIPO)
                        def _first_mode(s: pd.Series):
                            try:
                                m = s.mode(dropna=True)
                                return m.iloc[0] if not m.empty else None
                            except Exception:
                                return None

                        col_prof_safe = col_prof if (col_prof and col_prof in df_in.columns) else None
                        if not col_prof_safe:
                            st.info("No 'Professor' column available to compute faculty counts.")
                            return

                        grp = (
                            df_in
                            .groupby(df_in[col_prof_safe].astype(str).str.strip(), dropna=False)
                            .agg(
                                **{
                                    "P/S":      ("_PS",   _first_mode),
                                    "Type":     ("_TIPO", _first_mode),
                                    "#Courses": (col_prof_safe, "count"),
                                    "Credits":  ("_CRED", "sum"),
                                }
                            )
                            .reset_index()
                            .rename(columns={col_prof_safe: "Professor"})
                        ).sort_values(["#Courses", "Credits"], ascending=[False, False]).reset_index(drop=True)

                        # English title (respecting the selected scope)
                        if opt_val_local in {"(TOTAL)", "(All)"}:
                            fac_title = f"Faculty with the most courses in {display_label}"
                        else:
                            fac_title = f"Faculty with the most {cfg[view]['label']} courses in {display_label}: {opt_val_local}"
                        st.markdown(f"### {fac_title}")

                        # Style top 5 (mint diamond)
                        def _style_top5(df_):
                            sty = pd.DataFrame('', index=df_.index, columns=df_.columns)
                            top_mask = df_.index < 5
                            for c in df_.columns:
                                sty.loc[top_mask, c] = f'background-color: {MINT_DIAMOND}; font-weight:600;'
                            return sty

                        _download_xlsx_button(
                            grp,
                            f"faculty_by_courses_{_slugify(opt_val_local)}_{_slugify(display_label)}.xlsx",
                            key=f"dl_fac_by_courses_{_slugify(opt_val_local)}_{_slugify(display_label)}",
                            label="⬇️ Download table (Excel)"
                        )
                        st.dataframe(
                            grp.style.format({"Credits": "{:,.1f}"}).apply(_style_top5, axis=None),
                            use_container_width=True, hide_index=True
                        )

                    # ---------- Title (only) ----------
                    display_label = st.session_state.get('sel_label', 'Selected Period')

                    # Build detail table
                    col_periodo_orig = _get_any(df_car, "Periodo")
                    wanted_map = {
                        "Period": col_periodo_orig or col_sem, "Course Code": col_code, "Credits": col_cred,
                        "Course Name": col_name, "Program": col_prog, "Professor": col_prof,
                        "Course Area": col_areaCourse, "Field": col_field, "Type": col_tipoC, "P/S": col_ps_C,
                    }
                    present = {nice: col for nice, col in wanted_map.items() if col in base_tbl.columns}
                    out = base_tbl[list(present.values())].rename(columns={v: k for k, v in present.items()})
                    n_courses = len(out)

                    # Human-readable filter suffix for the title
                    desc_parts = []
                    if table_filter_ps == "Only P":
                        desc_parts.append("by Participating Faculty")
                    elif table_filter_ps == "Only S":
                        desc_parts.append("by Supporting Faculty")
                    if table_filter_tipo == "Only SA":
                        desc_parts.append("by Scholarly Academics")
                    elif table_filter_tipo == "Only OTHER":
                        desc_parts.append("by Others")
                    desc_suffix = (" " + " and ".join(desc_parts)) if desc_parts else ""

                    if opt_val in {"(TOTAL)", "(All)"}:
                        title = f"{n_courses} courses were taught in {display_label}{desc_suffix}"
                    else:
                        title = f"{n_courses} {cfg[view]['label']} courses were taught in {display_label}: {opt_val}{desc_suffix}"
                    st.markdown(f"### {title}")

                    # ---------- Action row (left: download, right: popup button) ----------
                    act_l, act_r = st.columns([0.70, 0.30])
                    with act_l:
                        _download_xlsx_button(
                            out,
                            f"table_detail_{_slugify(opt_val)}_{_slugify(display_label)}.xlsx",
                            key=f"dl_tbl_detail_{_slugify(opt_val)}_{_slugify(display_label)}",
                            label="⬇️ Download table (Excel)"
                        )
                    with act_r:
                        open_popup = st.button(
                            "Faculty by course count",
                            key=f"open_popup_{view}_{opt_val}",
                            use_container_width=True
                        )

                    # ---------- Popup (modal if available; else expander fallback) ----------
                    if open_popup:
                        if hasattr(st, "dialog"):
                            @st.dialog("Faculty by course count", width="large")
                            def _dlg():
                                _show_faculty_popup(base_tbl, display_label, opt_val)
                                if st.button("Close"):
                                    st.rerun()
                            _dlg()
                        else:
                            with st.expander("Faculty by course count", expanded=True):
                                _show_faculty_popup(base_tbl, display_label, opt_val)

                    # ---------- Final detail table ----------
                    st.dataframe(out, use_container_width=True, hide_index=True)

                # ==================================================
                # RIGHT: Donut %P or %Type + download
                # NOTE: The donut MUST NOT be affected by the P/S or SA/OTHER table filters above.
                #       It only reacts to the selected Area/Field/Program (opt_val) and timeframe.
                # ==================================================
                with cR:
                    st.markdown("<div style='height: 110px'></div>", unsafe_allow_html=True)

                    # Build a SCOPE for the donut that ignores the table's P/S and Type filters:
                    # -> Use the time-filtered base; narrow ONLY by Area/Field/Program selection.
                    if opt_val in {"(TOTAL)", "(All)"} or col_tag not in base.columns:
                        base_scoped = base.copy()
                        title_suffix = "TOTAL"
                    else:
                        base_scoped = base[base[col_tag] == opt_val].copy()
                        title_suffix = opt_val

                    # Aggregates for donut (from base_scoped ONLY)
                    agg_tipo = base_scoped.groupby("_TIPO", dropna=False)["_CRED"].sum() if "_TIPO" in base_scoped else pd.Series(dtype=float)
                    agg_ps   = base_scoped.groupby("_PS",   dropna=False)["_CRED"].sum() if "_PS"   in base_scoped else pd.Series(dtype=float)

                    # Fill missing categories
                    p_val = float(agg_ps.get("P", 0.0))
                    s_val = float(agg_ps.get("S", 0.0))
                    sa = float(agg_tipo.get("SA", 0.0))
                    pa = float(agg_tipo.get("PA", 0.0))
                    sp = float(agg_tipo.get("SP", 0.0))
                    ip = float(agg_tipo.get("IP", 0.0))
                    other = float(agg_tipo.get("OTHER", 0.0))

                    donut_h = 360
                    thrP = 75.0 if title_suffix == "TOTAL" else 60.0  # optional alert threshold

                    if metric_choice == "%P":
                        den = p_val + s_val
                        p_share = (p_val/den*100) if den else 0.0
                        alert = (p_share < thrP)
                        color_map = {"P": ( "#F5A3A3" if alert else MINT ), "S": "#B0B0B0"}
                        fig = px.pie(
                            names=["P","S"],
                            values=[p_val, s_val],
                            color=["P","S"],
                            color_discrete_map=color_map,
                            hole=0.55
                        )
                        fig.update_traces(textinfo="percent+label", hovertemplate="%{label}: %{percent:.1%}<extra></extra>")
                        fig.update_layout(
                            title=f"% Participating Distribution — {title_suffix}",
                            height=donut_h, margin=dict(l=10, r=10, t=40, b=10),
                            legend=dict(orientation="v", yanchor="bottom", y=0.4, xanchor="center", x=0.9)
                        )
                        st.plotly_chart(fig, use_container_width=True)

                        donut_df = pd.DataFrame({"Group": ["P","S"], "Credits": [p_val, s_val]})
                        donut_df["Percent"] = (donut_df["Credits"] / max(1e-9, donut_df["Credits"].sum()))*100
                        _download_xlsx_button(
                            donut_df,
                            f"chart_donut_PS_{_slugify(title_suffix)}_{_slugify(st.session_state.get('sel_label','sel'))}.xlsx",
                            key=f"dl_donut_ps_{_slugify(title_suffix)}_{_slugify(st.session_state.get('sel_label','sel'))}",
                            label="⬇️ Download (Excel)"
                        )
                    else:
                        labels_all = ["SA", "PA", "SP", "IP", "OTHER"]
                        values_all = [sa, pa, sp, ip, other]
                        filtered   = [(l, v) for l, v in zip(labels_all, values_all) if v > 0]

                        if filtered:
                            labels = [l for l, _ in filtered]; values = [v for _, v in filtered]
                            den = sum(values_all) or 1.0
                            sa_share    = sa/den*100
                            other_share = other/den*100
                            cmap = {l: "#B0B0B0" for l in labels}
                            if "SA" in labels:    cmap["SA"]    = ("#F5A3A3" if sa_share   < 40.0 else MINT)
                            if "OTHER" in labels: cmap["OTHER"] = ("#F5A3A3" if other_share > 10.0 else "#6B7280")

                            fig = px.pie(
                                names=labels, values=values, color=labels, color_discrete_map=cmap, hole=0.55
                            )
                            fig.update_traces(textinfo="percent+label", sort=False, hovertemplate="%{label}: %{percent:.1%}<extra></extra>")
                            title_txt = "%SA Distribution" if metric_choice == "%SA" else "%OTHER Distribution"
                            fig.update_layout(
                                title=f"{title_txt} — {title_suffix}",
                                height=donut_h, margin=dict(l=10, r=10, t=40, b=10),
                                legend=dict(orientation="v", yanchor="bottom", y=0.4, xanchor="center", x=0.9)
                            )
                            st.plotly_chart(fig, use_container_width=True)

                            donut_df = pd.DataFrame({"Type": labels_all, "Credits": values_all})
                            donut_df["Percent"] = (donut_df["Credits"] / max(1e-9, donut_df["Credits"].sum()))*100
                            _download_xlsx_button(
                                donut_df,
                                f"chart_donut_TIPO_{_slugify(title_suffix)}_{_slugify(st.session_state.get('sel_label','sel'))}.xlsx",
                                key=f"dl_donut_tipo_{_slugify(title_suffix)}_{_slugify(st.session_state.get('sel_label','sel'))}",
                                label="⬇️ Download (Excel)"
                            )
                        else:
                            st.caption("No type records for this metric in the selected timeframe.")
        except Exception:
            pass

    # --------------------------
    # COUNTS — PIVOT / BSQ (Oculto cuando Sensitivity mode está activo)
    # --------------------------
    if not SENS.get("on", False):
        st.markdown("---")
        st.subheader(f"Participating vs Supporting — {st.session_state.get('sel_label','Selected')}")

        # ------- Base Faculty Distribution filtrada -------
        def _filter_fd_scope(df_fd_raw: pd.DataFrame) -> pd.DataFrame:
            if df_fd_raw.empty:
                return df_fd_raw.copy()
            out = df_fd_raw.copy()
            semc = _get_any(out, "Semestre","Periodo","Periodo Académico","Periodo academico")
            if semc:
                out["_SEM_SRC"]   = out[semc].astype(str).str.strip()
                out["_YEARX"]     = out["_SEM_SRC"].map(_extract_year).astype("Int64")
                out["_IS_INTER"]  = out["_SEM_SRC"].str.lower().str.contains("inter", na=False)
            else:
                out["_SEM_SRC"]  = ""
                out["_YEARX"]    = pd.Series(dtype="Int64")
                out["_IS_INTER"] = False

            time_mode = st.session_state.get("time_mode", "Semestral")
            sel_sem   = st.session_state.get("sel_sem")
            sel_year  = st.session_state.get("sel_year")

            if time_mode == "Semestral" and sel_sem is not None:
                return out[out["_SEM_SRC"].eq(str(sel_sem))].copy()
            if time_mode == "Intersemestral" and sel_year is not None:
                return out[(out["_YEARX"] == int(sel_year)) & (out["_IS_INTER"])].copy()
            if time_mode == "Anual" and sel_year is not None:
                return out[out["_YEARX"] == int(sel_year)].copy()
            return out

        df_fd_scope = _filter_fd_scope(df_fd)
        df_fd_f = df_fd_scope.copy()

        # -------- columnas base y extra --------
        if col_ps_fd:   df_fd_f["_PS"]   = _norm_str(df_fd_f[col_ps_fd]).map(normalize_ps)
        if col_area_fd: df_fd_f["_AREA"] = df_fd_f[col_area_fd].astype(str).str.strip()
        if col_tipo_fd: df_fd_f["_TIPO"] = _norm_str(df_fd_f[col_tipo_fd]).map(normalize_tipo)

        col_genero = _get_any(df_fd_f, "GÉNERO", "GENERO", "Genero", "Gender")
        col_degree = _get_any(df_fd_f, "Highest Degree", "HighestDegree", "DEGREE", "Grado máximo", "Grado")
        col_ftpt   = _get_any(df_fd_f, "PLANTA_CATEDRA", "Planta_Catedra", "Planta/Catedra", "Full/Part")

        # --------- Controles en una sola fila (3 botones) ----------
        pivot_mode = st.radio(
            "View",
            ["BSQ Compensation", "AREA", "Qualification Type"],
            index=0,  # BSQ por defecto y seleccionado
            horizontal=True,
            label_visibility="collapsed",
            key="counts_view_mode"
        )

        # ===================== MODO BSQ =====================
        if pivot_mode == "BSQ Compensation":
            left, right = st.columns([6,6], gap="large")

            if not all([col_genero, col_degree, col_ftpt]):
                st.error("Missing columns in 'Faculty Distribution' for BSQ tables: 'GÉNERO', 'Highest Degree', and/or 'PLANTA_CATEDRA'.")
            else:
                # ---- construir df_bsq inicial y tomar columnas relevantes ----
                df_bsq = _ensure_pid(df_fd_f).assign(
                    Gender     = df_fd_f[col_genero].map(_norm_gender),
                    IsDoctoral = df_fd_f[col_degree].map(_is_doctoral),
                    FTPT_raw   = df_fd_f[col_ftpt],
                    PS_raw     = df_fd_f["_PS"],
                    TIPO_raw   = df_fd_f["_TIPO"]
                )

                # ---- Normalización básica sin librerías nuevas ----
                # 1) pasar a str, fillna, strip y upper
                df_bsq["FTPT"] = df_bsq["FTPT_raw"].fillna("").astype(str).str.strip().str.upper()
                df_bsq["PS"]   = df_bsq["PS_raw"].fillna("").astype(str).str.strip().str.upper()
                df_bsq["TIPO"] = df_bsq["TIPO_raw"].fillna("OTHER").astype(str).str.strip().str.upper()

                # 2) reemplazar acentos comunes manualmente (mayúsculas ya)
                replacements = {
                    "Á":"A", "É":"E", "Í":"I", "Ó":"O", "Ú":"U",
                    "À":"A", "È":"E", "Ì":"I", "Ò":"O", "Ù":"U",
                    "Ñ":"N"
                }
                for old, new in replacements.items():
                    df_bsq["FTPT"] = df_bsq["FTPT"].str.replace(old, new, regex=False)
                    df_bsq["PS"]   = df_bsq["PS"].str.replace(old, new, regex=False)
                    df_bsq["TIPO"] = df_bsq["TIPO"].str.replace(old, new, regex=False)

                # eliminar columnas temporales si quieres
                df_bsq = df_bsq.drop(columns=["FTPT_raw","PS_raw","TIPO_raw"])

                # ----------------- Agrupar por _PID para obtener una fila única por profesor (versión robusta) -----------------
                def _first_non_empty_str_like(series):
                    vals = series.fillna("").astype(str).str.strip().tolist()
                    for v in vals:
                        if v != "":
                            return v
                    return vals[0] if vals else ""

                def _first_not_other_str_like(series):
                    vals = series.fillna("").astype(str).str.strip().tolist()
                    for v in vals:
                        if v != "" and v.upper() != "OTHER":
                            return v
                    return vals[0] if vals else ""

                def _make_fac_row(g):
                    return pd.Series({
                        "Gender":     _first_non_empty_str_like(g["Gender"]),
                        "IsDoctoral": bool(g["IsDoctoral"].any()),
                        "FTPT":       _first_non_empty_str_like(g["FTPT"]),
                        "PS":         _first_non_empty_str_like(g["PS"]),
                        "TIPO":       _first_not_other_str_like(g["TIPO"])
                    })

                # apply devuelve una fila por grupo (por _PID)
                df_fac = df_bsq.groupby("_PID", as_index=True).apply(_make_fac_row).reset_index()

                # ----------------- Cálculos de tabla 7 (género) sobre df_fac (una fila por profesor) ----
                def _count_by_gender_from(df_, mask) -> dict:
                    sub = df_.loc[mask]
                    male   = int((sub["Gender"] == "Male").sum())
                    female = int((sub["Gender"] == "Female").sum())
                    other  = int((sub["Gender"] == "Other").sum())
                    return {"Male": male, "Female": female, "Other": other, "Total": male + female + other}

                row7a = _count_by_gender_from(df_fac, df_fac["PS"] == "P")
                row7b = _count_by_gender_from(df_fac, (df_fac["PS"] == "P") & (df_fac["IsDoctoral"]))
                row7c = _count_by_gender_from(df_fac, df_fac["PS"] == "S")
                row7d = _count_by_gender_from(df_fac, (df_fac["PS"] == "S") & (df_fac["IsDoctoral"]))

                tbl7 = pd.DataFrame([
                    {"Row": "a. Total number of participating faculty members", **row7a},
                    {"Row": "b. Total number of participating faculty members with doctoral degrees", **row7b},
                    {"Row": "c. Total number of supporting faculty members", **row7c},
                    {"Row": "d. Total number of supporting faculty members with doctoral degrees", **row7d},
                ])

                def _bold_rows_7(df_):
                    sty = pd.DataFrame('', index=df_.index, columns=df_.columns)
                    mask = df_["Row"].str.startswith(("b.", "d."))
                    for c in df_.columns:
                        sty.loc[mask, c] = 'font-weight:700;'
                    return sty

                # ----------------- Cálculos de tabla 8 (por TIPO) sobre df_fac -----------------
                cats = ["SA","PA","SP","IP","OTHER"]
                def _row_qual_from(df_, ps_code: str, ftpt_code: str | None):
                    m = (df_["PS"] == ps_code)
                    if ftpt_code is not None:
                        m = m & (df_["FTPT"] == ftpt_code)
                    sub = df_[m]  # df_ ya es por _PID único (df_fac)
                    counts = {c: int((sub["TIPO"] == c).sum()) for c in cats}
                    total = sum(counts.values())
                    return {**counts, "TOTAL": total}

                r8a = _row_qual_from(df_fac, "P", "PLANTA")
                r8b = _row_qual_from(df_fac, "P", "CATEDRA")  # normalizamos acentos arriba -> 'CATEDRA'
                r8c = {k: r8a.get(k,0) + r8b.get(k,0) for k in cats + ["TOTAL"]}
                r8d = _row_qual_from(df_fac, "S", "PLANTA")
                r8e = _row_qual_from(df_fac, "S", "CATEDRA")
                r8f = {k: r8d.get(k,0) + r8e.get(k,0) for k in cats + ["TOTAL"]}

                tbl8 = pd.DataFrame([
                    {"Row": "a. Full-time Participating faculty members", **r8a},
                    {"Row": "b. Part-time Participating faculty members", **r8b},
                    {"Row": "c. Total Participating faculty members", **r8c},
                    {"Row": "d. Full-time Supporting faculty members", **r8d},
                    {"Row": "e. Part-time Supporting faculty members", **r8e},
                    {"Row": "f. Total Supporting faculty members", **r8f},
                ])[["Row"] + cats + ["TOTAL"]]

                def _bold_rows_8(df_):
                    sty = pd.DataFrame('', index=df_.index, columns=df_.columns)
                    mask = df_["Row"].str.startswith(("c.", "f."))
                    for c in df_.columns:
                        sty.loc[mask, c] = 'font-weight:700;'
                    return sty

                # ---- mostrar tablas finales ----
                with left:
                    st.markdown("**7. Participating and Supporting Faculty Counts †**")
                    _download_xlsx_button(
                        tbl7, f"bsq_7_gender_counts_{_slugify(st.session_state.get('sel_label','sel'))}.xlsx",
                        key=f"dl_bsq7_{_slugify(st.session_state.get('sel_label','sel'))}",
                        label="Download table 7 (Excel)"
                    )
                    st.dataframe(
                        tbl7.style.apply(_bold_rows_7, axis=None).format({"Male":"{:,.0f}","Female":"{:,.0f}","Other":"{:,.0f}","Total":"{:,.0f}"}),
                        use_container_width=True, hide_index=True
                    )

                with right:
                    st.markdown("**8. Faculty Counts by Qualification Types †**")
                    _download_xlsx_button(
                        tbl8, f"bsq_8_qual_counts_{_slugify(st.session_state.get('sel_label','sel'))}.xlsx",
                        key=f"dl_bsq8_{_slugify(st.session_state.get('sel_label','sel'))}",
                        label="Download table 8 (Excel)"
                    )
                    st.dataframe(
                        tbl8.style.apply(_bold_rows_8, axis=None).format({c: "{:,.0f}" for c in cats + ["TOTAL"]}),
                        use_container_width=True, hide_index=True
                    )

        # ===================== MODO PIVOT ORIGINAL (AREA / TYPE) =====================
        else:
            # Define filas según modo
            if pivot_mode == "AREA":
                row_name   = "AREA"
                row_series = df_fd_f["_AREA"].astype(str).str.strip().replace({"": "N/A"})
                desired_order = None
            else:  # "Qualification Type"
                row_name   = "Type"
                row_series = df_fd_f["_TIPO"].map(lambda v: str(v).upper())
                desired_order = ["SA", "PA", "SP", "IP", "OTHER"]

            # Persona + variables para deduplicar
            df_cnt = _ensure_pid(df_fd_f)
            df_cnt[row_name] = row_series
            df_cnt["_PS2"]   = df_cnt["_PS"].fillna("")

            # DEDUP: 1 vez por persona y categoría (row_name, _PS2)
            df_cnt = df_cnt.drop_duplicates(subset=["_PID", row_name, "_PS2"])

            base = pd.DataFrame({row_name: df_cnt[row_name], "_PS": df_cnt["_PS2"]})
            table = (base.groupby([row_name, "_PS"], dropna=False)
                          .size()
                          .unstack(fill_value=0)
                          .rename(columns={"P": "Participating", "S": "Supporting"}))
            for k in ["Participating", "Supporting"]:
                if k not in table.columns: table[k] = 0
            table["__Total__"] = table["Participating"] + table["Supporting"]

            # Ajuste por sensibilidad (impacto total) — si hubiera operaciones cargadas
            if SENS["on"] and SENS.get("ops"):
                add_P = sum(op.get("count",0) for op in SENS["ops"] if op.get("scope")=="PS" and op.get("cat")=="P")
                add_S = sum(op.get("count",0) for op in SENS["ops"] if op.get("scope")=="PS" and op.get("cat")=="S")
                incs = {"Participating": int(add_P), "Supporting": int(add_S)}
            else:
                incs = {"Participating": 0, "Supporting": 0}

            df_counts = table[["Participating", "Supporting"]].astype(int).reset_index()
            total_row = pd.DataFrame([{row_name: "TOTAL",
                                       "Participating": int(df_counts["Participating"].sum()) + incs["Participating"],
                                       "Supporting":    int(df_counts["Supporting"].sum())    + incs["Supporting"]}])
            df_counts_out = pd.concat([df_counts, total_row], ignore_index=True)

            def _bold_total(df_):
                sty = pd.DataFrame('', index=df_.index, columns=df_.columns)
                mask = df_[row_name].astype(str).str.upper().eq("TOTAL")
                for c in df_.columns: sty.loc[mask, c] = 'font-weight:700;'
                return sty

            left, right = st.columns([6,6], gap="large")

            # Porcentajes y orden para gráfica
            denom = table["__Total__"].replace(0, pd.NA)
            perc_df = pd.DataFrame({
                row_name: table.index,
                "%Participating": (table["Participating"] / denom * 100).round(1).fillna(0.0),
                "%Supporting":    (table["Supporting"]    / denom * 100).round(1).fillna(0.0),
            })
            if desired_order:
                for code in desired_order:
                    if code not in perc_df[row_name].tolist():
                        perc_df.loc[len(perc_df)] = [code, 0.0, 0.0]
                cat_order = desired_order
            else:
                cat_order = perc_df[row_name].tolist()

            chart_export = perc_df.melt(id_vars=row_name, value_vars=["%Participating", "%Supporting"],
                                        var_name="Group", value_name="Percent")

            with left:
                _download_xlsx_button(
                    df_counts_out,
                    f"ps_counts_{_slugify(row_name)}_{_slugify(st.session_state.get('sel_label','sel'))}.xlsx",
                    key=f"dl_ps_counts_{_slugify(row_name)}_{_slugify(st.session_state.get('sel_label','sel'))}",
                    label="Download table (Excel)"
                )
                styled_counts = (df_counts_out.style
                                 .format({"Participating": "{:,.0f}", "Supporting": "{:,.0f}"})
                                 .apply(_bold_total, axis=None))
                st.dataframe(styled_counts, use_container_width=True, hide_index=True)

            with right:
                fig = px.bar(
                    chart_export, x=row_name, y="Percent", color="Group",
                    barmode="group", text="Percent",
                    color_discrete_map={"%Participating": MINT, "%Supporting": SUPPORTING},
                    category_orders={row_name: cat_order}
                )
                fig.update_traces(texttemplate="%{text:.1f}%")
                fig.update_layout(
                    xaxis_title=None, yaxis_title=None, height=340,
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
                    legend_title_text=None, margin=dict(l=20, r=10, t=10, b=40)
                )
                st.plotly_chart(fig, use_container_width=True)
                _download_xlsx_button(
                    chart_export,
                    f"chart_ps_perc_{_slugify(row_name)}_{_slugify(st.session_state.get('sel_label','sel'))}.xlsx",
                    key=f"dl_chart_ps_perc_{_slugify(row_name)}_{_slugify(st.session_state.get('sel_label','sel'))}",
                    label="Download (Excel)"
                )

    # ==========================================================
    # MÓDULO FINAL: Top 5 (más/menos créditos) y FT sin cursos + Buscador
    # (Oculto cuando Sensitivity mode está activo)
    # ==========================================================
    if not SENS.get("on", False):
        st.markdown("---")

        # ====== Constantes para evitar errores de texto en radios ======
        HL_MOST   = "Top 5 with most credits"
        HL_LEAST  = "Top 5 with least credits"
        HL_ZERO   = "Full-time with 0 courses"

        SM_FAC    = "By Faculty"
        SM_COURSE = "By Course"

        # ====== Label de periodo ======
        display_label = st.session_state.get("sel_label", "Selected")

        # ====== Título pegado a los controles ======
        head_l, head_r = st.columns([7,5], gap="large")
        with head_l:
            st.markdown(f"#### Faculty credit highlights of {display_label}")
        with head_r:
            st.write("")

        df_planta = qual_load_planta()

        # ====== Parámetros de alcance temporal ======
        time_mode    = st.session_state.get("time_mode","Semestral")
        sel_year     = st.session_state.get("sel_year")
        sel_sem_code = st.session_state.get("sel_sem")

        # ====== Aplicar filtros robustos ======
        df_fd_sem    = filter_df_fd(df_fd, time_mode, sel_year, sel_sem_code).copy()
        df_car_scope = filter_df_car(df_car_filt_all, time_mode, sel_year, sel_sem_code)

        # -------- Distribution (para enriquecer con ID/AREA/TIPO/P-S y FT/PT) --------
        col_prof_fd = _get_any(df_fd_sem, "Profesor","PROFESOR","Docente","Nombre")
        col_id_fd   = _get_any(df_fd_sem, "ID","ID Nr.","Documento")
        col_area_fd = _get_any(df_fd_sem, "AREA_PROFESOR","Area_Profesor","Area Profesor","Área","Area")
        col_tipo_fd = _get_any(df_fd_sem, "TIPO","Tipo","Ranking","Tipo Ranking")
        col_ps_fd   = _get_any(df_fd_sem, "P/S","P - S","Participating/Supporting")
        col_ftpt    = _get_any(df_fd_sem, "PLANTA_CATEDRA", "Planta_Catedra", "Planta/Catedra", "Full/Part")

        if col_prof_fd: df_fd_sem["_PROF_N"]    = df_fd_sem[col_prof_fd].astype(str).str.strip()
        if col_id_fd:   df_fd_sem["_ID"]        = df_fd_sem[col_id_fd].astype(str).str.strip()
        if col_area_fd: df_fd_sem["_AREA_PROF"] = df_fd_sem[col_area_fd].astype(str).str.strip()
        if col_tipo_fd: df_fd_sem["_TIPO"]      = df_fd_sem[col_tipo_fd].astype(str).str.strip()
        if col_ps_fd:   df_fd_sem["_PS"]        = _norm_str(df_fd_sem[col_ps_fd]).map(normalize_ps)

        # FT/PT directo o fallback por TIPO
        if col_ftpt:
            df_fd_sem["_FTPT"] = df_fd_sem[col_ftpt].map(_norm_ftpt)
        else:
            df_fd_sem["_FTPT"] = df_fd_sem["_TIPO"].map(_norm_ftpt) if "_TIPO" in df_fd_sem else ""

        # --- Maps por profesor (primera coincidencia) ---
        prof_to_id_map_by_name = _first_map(df_fd_sem, "_PROF_N", "_ID")
        prof_to_area_map       = _first_map(df_fd_sem, "_PROF_N", "_AREA_PROF")
        prof_to_tipo_map       = _first_map(df_fd_sem, "_PROF_N", "_TIPO")
        prof_to_ps_map         = _first_map(df_fd_sem, "_PROF_N", "_PS")

        # Conjunto de IDs PLANTA (cruce por ID)
        planta_ids = set(df_fd_sem.loc[df_fd_sem["_FTPT"] == "PLANTA", "_ID"].dropna().astype(str).unique().tolist())

        # -------- helpers de columnas en Cartelera (sobre df_car_scope) --------
        col_prof_car = _get_any(df_car_scope, "Profesor(es)","Profesor","PROFESOR","Docente")
        col_cred_car = _get_any(df_car_scope, "Créditos","Creditos","Credits")
        col_sem_car  = _get_any(df_car_scope, "Semestre","Periodo","Periodo Académico","Periodo academico")
        col_code_car = _get_any(df_car_scope, "Código Materia","Codigo Materia","CODIGO MATERIA","Código","Codigo","Course Code")
        col_name_car = _get_any(df_car_scope, "Nombre largo curso","Nombre Curso","Nombre del curso","Course Name")
        col_secc_car = _get_any(df_car_scope, "Secc","Sección","Seccion","Section")
        col_acar_car = _get_any(df_car_scope, "Area del curso","Área del curso","Area del Curso","AREA DEL CURSO")
        col_field_car= _get_any(df_car_scope, "Field","FIELD","Campo","Área de conocimiento")
        col_prog_car = _get_any(df_car_scope, "Program","PROGRAM","program")
        col_campus   = _get_any(df_car_scope, "Campus","CAMPUS","Sede")

        if col_cred_car and "_CRED" not in df_car_scope.columns:
            df_car_scope["_CRED"] = pd.to_numeric(df_car_scope[col_cred_car], errors="coerce").fillna(0.0)

        # ========= Controles de “Top / Zero” (izquierda) + Buscador (derecha) =========
        opt_highlight = st.radio(
            "",
            [HL_MOST, HL_LEAST, HL_ZERO],
            index=0, horizontal=True, label_visibility="visible", key="highlight_mode"
        )

        left, right = st.columns([7,5], gap="large")

        # ======================= PANEL IZQUIERDO =======================
        with left:
            if opt_highlight in {HL_MOST, HL_LEAST}:
                # switch PLANTA (por ID)
                only_ft = st.toggle("Only Full-time Faculty", value=False, key="top_only_ft")

                if (not col_prof_car) or ("_CRED" not in df_car_scope.columns):
                    st.info("Missing credits or professor column in Cartelera for this view.")
                else:
                    df_top = (
                        df_car_scope
                        .assign(_PROF=df_car_scope[col_prof_car].astype(str).str.strip())
                        .groupby("_PROF", as_index=False)
                        .agg(Credits=("_CRED","sum"), nCourses=(col_prof_car,"count"))
                    )

                    # ID por nombre (para filtrar PLANTA)
                    df_top["ID"] = df_top["_PROF"].map(prof_to_id_map_by_name)

                    # Filtrar por PLANTA si aplica
                    if only_ft:
                        if planta_ids:
                            df_top = df_top[df_top["ID"].astype(str).isin(planta_ids)]
                        else:
                            df_top = df_top.iloc[0:0]

                    # Orden asc/desc según modo
                    asc = (opt_highlight == HL_LEAST)
                    df_top = df_top.sort_values("Credits", ascending=asc).head(5).copy()

                    # Enriquecer
                    df_top["AREA_PROFESOR"] = df_top["_PROF"].map(prof_to_area_map)
                    df_top["TIPO"]          = df_top["_PROF"].map(prof_to_tipo_map)
                    df_top["P/S"]           = df_top["_PROF"].map(prof_to_ps_map)

                    out = (
                        df_top.rename(columns={"_PROF":"Profesor"})
                              [["Profesor","ID","AREA_PROFESOR","TIPO","P/S","Credits","nCourses"]]
                              .rename(columns={"nCourses":"#Cursos"})
                    )

                    title = "Top 5 professors by credits (most)" if not asc else "Top 5 professors by credits (least)"
                    _download_xlsx_button(
                        out,
                        f"highlight_{_slugify(title)}_{_slugify(display_label)}.xlsx",
                        key=f"dl_highlight_{_slugify(title)}_{_slugify(display_label)}",
                        label="Download (Excel)"
                    )
                    st.dataframe(out.style.format({"Credits":"{:,.1f}"}), use_container_width=True, hide_index=True)

            else:
                # ========= Full-time con 0 cursos =========
                col_period_pl = _get_any(df_planta, "Periodo","PERIODO","Semestre")
                col_id_pl     = _get_any(df_planta, "ID Nr.","ID","Documento")
                if df_planta.empty or not all([col_period_pl, col_id_pl]):
                    st.info("Load 'BD_PLANTA' to compute FT with 0 courses.")
                else:
                    # Alcance temporal para PLANTA y DISTRIBUTION
                    col_sem_fd_all = _get_any(df_fd, "Semestre","Periodo","Periodo Académico","Periodo academico")
                    col_id_fd_all  = _get_any(df_fd, "ID","ID Nr.","Documento")

                    if time_mode == "Semestral" and sel_sem_code is not None:
                        df_ft = df_planta[df_planta[col_period_pl].astype(str).str.strip().eq(str(sel_sem_code))].copy()
                        alcance_txt = str(sel_sem_code)
                        taught_ids = set()
                        if col_sem_fd_all and col_id_fd_all:
                            taught_ids = set(
                                df_fd.loc[df_fd[col_sem_fd_all].astype(str).str.strip().eq(str(sel_sem_code)), col_id_fd_all]
                                     .astype(str).str.strip()
                            )
                    elif time_mode == "Intersemestral" and sel_year is not None:
                        goal = f"{int(sel_year)} Intersemestral"
                        mask_inter = df_planta[col_period_pl].map(_normalize_sem_str).str.fullmatch(re.escape(goal), case=False, na=False)
                        df_ft = df_planta[mask_inter].copy()
                        alcance_txt = goal
                        taught_ids = set()
                        if col_sem_fd_all and col_id_fd_all:
                            taught_ids = set(
                                df_fd.loc[
                                    df_fd[col_sem_fd_all].map(_normalize_sem_str).str.fullmatch(re.escape(goal), case=False, na=False),
                                    col_id_fd_all
                                ].astype(str).str.strip()
                            )
                    elif time_mode == "Anual" and sel_year is not None:
                        df_ft = df_planta[df_planta[col_period_pl].astype(str).str.contains(str(sel_year), na=False)].copy()
                        alcance_txt = f"{sel_year} (annual)"
                        taught_ids = set()
                        if col_sem_fd_all and col_id_fd_all:
                            taught_ids = set(
                                df_fd.loc[df_fd[col_sem_fd_all].astype(str).str.contains(str(sel_year), na=False), col_id_fd_all]
                                     .astype(str).str.strip()
                            )
                    else:
                        df_ft = pd.DataFrame()
                        taught_ids = set()
                        alcance_txt = display_label

                    if df_ft.empty:
                        st.info(f"No full-time data found for {alcance_txt}.")
                    else:
                        df_ft["_ID"] = df_ft[col_id_pl].astype(str).str.strip()
                        ft_ids = set(df_ft["_ID"])
                        ft_total = len(ft_ids)
                        ft_teaching = len(ft_ids & taught_ids)
                        st.markdown(f"**Of the {ft_total} full-time Faculty, {ft_teaching} are teaching in {alcance_txt}.**")

                        missing_ids = sorted(ft_ids - taught_ids)
                        sub = df_ft[df_ft["_ID"].isin(missing_ids)].copy()

                        out = pd.DataFrame({
                            "Semester":      _pick(sub, "Periodo","Period"),
                            "ID Nr.":        sub["_ID"],
                            "First Name":    _pick(sub, "First Name","Nombre","Nombres"),
                            "Last Name":     _pick(sub, "Last Name","Apellidos","Apellido"),
                            "Academic Area": _pick(sub, "Academic Area","Área Académica","Area Académica","AREA_PROFESOR","Área"),
                            "Faculty Ranking": _pick(sub, "Faculty Ranking","Ranking","Rango"),
                            "Faculty Qualific.": _pick(sub, "Faculty Qualific.","Qualification","Qualific.","Qualif.","Tipo Ranking","TIPO"),
                            "P/S": _pick(sub, "P/S","P - S","Participating/Supporting")
                        })

                        _download_xlsx_button(
                            out, f"ft_zero_courses_{_slugify(alcance_txt)}.xlsx",
                            key=f"dl_ft_zero_{_slugify(alcance_txt)}",
                            label="Download (Excel)"
                        )
                        st.dataframe(out, use_container_width=True, hide_index=True)

        # ======================= PANEL DERECHO — BUSCADOR (modo único; control pegado) =======================
        with right:
            # Base: Cartelera YA FILTRADA (alcance de tiempo)
            base = df_car_scope.copy()

            if col_prof_car: base["_PROF"] = base[col_prof_car].astype(str).str.strip()
            if col_sem_car:  base["_SEM"]  = base[col_sem_car].astype(str).str.strip()
            if col_code_car: base["_CODE"] = base[col_code_car].astype(str).str.strip()
            if col_name_car: base["_NAME"] = base[col_name_car].astype(str).str.strip()

            # Enriquecer con ID (para cruzar con PLANTA) y Área (solo visual)
            if "_PROF" in base and prof_to_id_map_by_name:
                base["_ID_FROM_NAME"] = base["_PROF"].map(prof_to_id_map_by_name)
            if "_PROF" in base and prof_to_area_map:
                base["_AREA_PROF"] = base["_PROF"].map(prof_to_area_map)

            # ===== Opciones autocompletar =====
            # Profes (por timeframe). Si "Only Full-time" ON, limitar a PLANTA
            only_ft_toggle = bool(st.session_state.get("top_only_ft", False))
            if "_PROF" in base:
                if only_ft_toggle and planta_ids:
                    prof_opts_core = (
                        base.loc[base["_ID_FROM_NAME"].astype(str).isin(planta_ids), "_PROF"]
                        .dropna().unique().tolist()
                    )
                else:
                    prof_opts_core = base["_PROF"].dropna().unique().tolist()
                prof_opts_core = sorted(prof_opts_core)
            else:
                prof_opts_core = []
            prof_opts = [""] + prof_opts_core  # primer elemento vacío

            # Cursos (por timeframe) — sin filtrar por PLANTA (solo aplica en modo Faculty)
            if "_NAME" in base and base["_NAME"].notna().any():
                course_opts_core = sorted(base["_NAME"].dropna().unique().tolist())
            elif "_CODE" in base:
                course_opts_core = sorted(base["_CODE"].dropna().unique().tolist())
            else:
                course_opts_core = []
            course_opts = [""] + course_opts_core

            # Espaciador para alinear hacia abajo
            st.markdown("<div style='min-height:140px'></div>", unsafe_allow_html=True)

            # Callback: al cambiar el modo, limpiar el otro selector
            def _on_mode_change():
                mode = st.session_state.get("srch_mode_right", SM_FAC)
                if mode == SM_FAC:
                    st.session_state["srch_course"] = ""
                else:
                    st.session_state["srch_prof"] = ""

            # Selector de modo pegado al buscador
            search_mode = st.radio(
                "Search...",
                [SM_FAC, SM_COURSE],
                index=0, horizontal=True, key="srch_mode_right",
                on_change=_on_mode_change
            )

            # Control único a todo el ancho según modo
            if search_mode == SM_FAC:
                st.selectbox(
                    "Faculty Name",
                    options=prof_opts,
                    index=(prof_opts.index(st.session_state.get("srch_prof",""))
                           if st.session_state.get("srch_prof","") in prof_opts else 0),
                    key="srch_prof"
                )
            else:  # SM_COURSE
                st.selectbox(
                    "Course Name",
                    options=course_opts,
                    index=(course_opts.index(st.session_state.get("srch_course",""))
                           if st.session_state.get("srch_course","") in course_opts else 0),
                    key="srch_course"
                )

        # ======================= RESULTADOS BUSQUEDA — FULL WIDTH =======================
        sel_prof    = st.session_state.get("srch_prof", "")
        sel_course  = st.session_state.get("srch_course", "")
        search_mode = st.session_state.get("srch_mode_right", SM_FAC)  # SM_FAC = "By Faculty", SM_COURSE = "By Course"

        has_query = (search_mode == SM_FAC and bool(sel_prof)) or (search_mode == SM_COURSE and bool(sel_course))

        if has_query:
            # Base: Cartelera YA FILTRADA (no sumar periodos fuera del alcance)
            base = df_car_scope.copy()

            # Normalizaciones base
            if col_prof_car: base["_PROF"] = base[col_prof_car].astype(str).str.strip()
            if col_sem_car:  base["_SEM"]  = base[col_sem_car].astype(str).str.strip()
            if col_code_car: base["_CODE"] = base[col_code_car].astype(str).str.strip()
            if col_name_car: base["_NAME"] = base[col_name_car].astype(str).str.strip()

            # Enriquecer con ID/AREA por nombre (para filtrar PLANTA y mostrar)
            if "_PROF" in base and prof_to_id_map_by_name:
                base["_ID_FROM_NAME"] = base["_PROF"].map(prof_to_id_map_by_name)
            if "_PROF" in base and prof_to_area_map:
                base["_AREA_PROF"] = base["_PROF"].map(prof_to_area_map)

            # Filtro según modo
            mask_all = pd.Series(True, index=base.index)

            if search_mode == SM_FAC and sel_prof:
                # Solo por nombre (sin ID)
                m_name = base["_PROF"].str.contains(re.escape(sel_prof), case=False, na=False) if "_PROF" in base else pd.Series(False, index=base.index)
                mask_all &= m_name

                # Si "Only Full-time" está activo, limitar a PLANTA
                only_ft_toggle = bool(st.session_state.get("top_only_ft", False))
                if only_ft_toggle and planta_ids:
                    m_ft = base["_ID_FROM_NAME"].astype(str).isin(planta_ids) if "_ID_FROM_NAME" in base else pd.Series(False, index=base.index)
                    mask_all &= m_ft

            if search_mode == SM_COURSE and sel_course:
                m_name = base["_NAME"].str.contains(re.escape(sel_course), case=False, na=False) if "_NAME" in base else pd.Series(False, index=base.index)
                m_code = base["_CODE"].str.contains(re.escape(sel_course), case=False, na=False) if "_CODE" in base else pd.Series(False, index=base.index)
                mask_all &= (m_name | m_code)

            res = base[mask_all].copy()

            # Resumen
            display_label = st.session_state.get("sel_label", "Selected")
            if search_mode == SM_FAC and sel_prof:
                if "_CRED" not in res.columns and col_cred_car:
                    res["_CRED"] = pd.to_numeric(res[col_cred_car], errors="coerce").fillna(0.0)
                tot_cr = float(res.get("_CRED", pd.Series([0]*len(res))).sum())
                tot_courses = int(res.shape[0])
                st.info(f"**The professor has taught {tot_cr:,.1f} credits in {tot_courses} course{'s' if tot_courses!=1 else ''} in {display_label}.**")

            if search_mode == SM_COURSE and sel_course:
                profs_cnt = res["_PROF"].nunique() if "_PROF" in res else 0
                st.info(f"**The course has been taught by {profs_cnt} professor{'s' if profs_cnt!=1 else ''} in {display_label}.**")

            # Salida
            col_sem_fd_all = _get_any(df_fd, "Semestre","Periodo","Periodo Académico","Periodo academico")
            col_id_fd_all  = _get_any(df_fd, "ID","ID Nr.","Documento")

            show_cols = {
                "Periodo": "_SEM" if "_SEM" in res else (col_sem_car or col_sem_fd_all),
                "Profesor": col_prof_car or col_prof_fd,
                "ID": "_ID_FROM_NAME" if "_ID_FROM_NAME" in res else col_id_fd_all,
                "AREA_PROFESOR": "_AREA_PROF" if "_AREA_PROF" in res else _get_any(df_fd, "AREA_PROFESOR","Area_Profesor","Area Profesor","Área","Area"),
                "Código Materia": col_code_car,
                "Nombre largo curso": col_name_car,
                "Secc": col_secc_car,
                "Area del curso": col_acar_car,
                "Field": col_field_car,
                "Program": col_prog_car,
                "Créditos": col_cred_car,
                "Campus": col_campus
            }

            data = {}
            out_cols = []
            for nice, col in show_cols.items():
                data[nice] = res[col] if (col in res.columns) else None
                out_cols.append(nice)

            res_out = pd.DataFrame(data, columns=out_cols).copy()
            if "Créditos" in res_out.columns:
                res_out["Créditos"] = pd.to_numeric(res_out["Créditos"], errors="coerce").fillna(0.0)

            _download_xlsx_button(
                res_out,
                f"search_results_{_slugify(display_label)}.xlsx",
                key=f"dl_search_{_slugify(display_label)}",
                label="Download Results (Excel)"
            )
            st.dataframe(res_out, use_container_width=True, hide_index=True)


# 6) PÁGINA 7 — Update Data (BD_PLANTA)
# ---------------------------------------------------------------------------
# Reemplaza el antiguo modal "Update data" del HTML de KPIs. Sube la
# Template_BD_PLANTA.xlsx, la transforma fila por fila con el mismo mapeo de
# columnas que usaba ese modal (función buildBDRow original), y la escribe
# directamente en la pestaña BD_PLANTA del Google Sheet maestro.

# Columnas B..W de la Template (índice 0-based tras quitar la columna A,
# que es solo la etiqueta descriptiva) → columna destino en BD_PLANTA.
# F y V de la base son fórmulas (Full Name y Age) y nunca se sobreescriben
# con datos de la template.
_PLANTA_TEMPLATE_HEADER_ROW = 4     # fila 4 = nombres de columna en la template
_PLANTA_TEMPLATE_DATA_ROW = 6       # los datos empiezan en la fila 6 (1-indexado)
_PLANTA_TEMPLATE_START_COL = 1      # columna B (0-indexado) = primera columna de datos

# Todas las templates (BD_PLANTA, BD_Cartelera, Cursos_Nuevos) comparten el
# mismo layout: encabezados en la fila 4, datos desde la fila 6, columna B en adelante.
def _read_generic_template(uploaded_file, header_row: int = 4, data_row: int = 6, start_col: int = 1) -> pd.DataFrame:
    raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)
    headers = raw.iloc[header_row - 1, start_col:].tolist()
    headers = [str(h).replace("\n", " ").strip() if pd.notna(h) else "" for h in headers]
    data = raw.iloc[data_row - 1:, start_col:].copy()
    data.columns = headers
    data = data.dropna(how="all").reset_index(drop=True)
    return data


def _validate_template_columns(df: pd.DataFrame, required_cols: List[str], template_label: str):
    """Revisa que el archivo tenga la forma esperada de esa template — si no,
    lanza un error claro (lo captura el try/except del uploader y lo muestra
    con st.error) en vez de dejar que reviente más adelante con un error
    críptico de pandas/KeyError."""
    missing = [c for c in required_cols if c not in df.columns]
    if df.empty or missing:
        detalle = f" Columnas esperadas que no encontré: {', '.join(missing)}." if missing else " El archivo está vacío."
        raise ValueError(
            f"Este archivo no parece ser la {template_label} — usa la plantilla oficial "
            f"descargada desde esta misma sección.{detalle}"
        )


# Estilo base compartido por las 4 funciones push_*_updates() que escriben en
# los .xlsx de Drive — evita recrear el mismo Font 4 veces.
_BASE_ARIAL_FONT = Font(name="Arial", size=11, color="000000") if _OPENPYXL_OK else None


AREA_OPTIONS = [
    "ORGANIZATIONS", "SUSTAINABILITY", "STRATEGY & ENTREPRENEURSHIP",
    "MANAGEMENT", "MARKETING", "FINANCE", "SCM & IT",
]

# Listas de valores reales encontrados en 'Info. Profesores', para los
# desplegables de la sección "Profesores nuevos" (mismo patrón que AREA_OPTIONS).
GENERO_OPTIONS = ["Male", "Female"]
TIPO_OPTIONS = ["IP", "OTHER", "PA", "SA", "SP"]
PS_OPTIONS = ["P", "S"]
PLANTA_CATEDRA_OPTIONS = ["PLANTA", "CÁTEDRA"]
HIGHEST_DEGREE_OPTIONS = ["Bachelor", "Master", "Ph.D.", "Specialization"]
REGION_OPTIONS = ["Africa", "Asia", "Europe", "Latin America", "North America", "Oceania"]
INTL_DEGREE_OPTIONS = ["Yes", "No"]


def _get_formula_text(cell) -> Tuple[Optional[str], bool]:
    """Devuelve (texto_de_la_fórmula, es_array) de una celda. Algunas fórmulas
    (XLOOKUP entrado con Ctrl+Shift+Enter) se guardan como ArrayFormula en vez
    de string plano — hay que detectar cuál es para copiarlas bien."""
    v = cell.value
    if isinstance(v, ArrayFormula):
        return v.text, True
    if isinstance(v, str) and v.startswith("="):
        return v, False
    return None, False


def _write_translated_formula(ws, col: int, row: int, tpl_text: str, is_array: bool, origin_ref: str, target_ref: str):
    """Copia una fórmula existente (tpl_text, tomada de origin_ref) hacia
    target_ref, ajustando referencias relativas/absolutas con Translator, y
    la escribe respetando si es una fórmula normal o de matriz (ArrayFormula)."""
    translated = Translator(tpl_text, origin=origin_ref).translate_formula(target_ref)
    if is_array:
        ws.cell(row=row, column=col, value=ArrayFormula(ref=target_ref, text=translated))
    else:
        ws.cell(row=row, column=col, value=translated)


def _needs_fix(v) -> bool:
    """True si una celda está vacía, o si aún tiene el texto de una fórmula
    (normal o de matriz) sin resolver — openpyxl nunca calcula fórmulas, así
    que cualquier celda así se ve en blanco en pandas/Streamlit hasta que
    alguien la reemplace por su valor literal ya calculado."""
    if v is None or v == "":
        return True
    if isinstance(v, str) and v.startswith("="):
        return True
    if isinstance(v, ArrayFormula):
        return True
    return False


def _last_data_row(ws, key_col: int = 1, header_row: int = 1, upper_bound: Optional[int] = None) -> int:
    """Última fila con datos REALES en `key_col` (no fórmula ni celda vacía
    con solo formato). openpyxl.Worksheet.max_row cuenta cualquier celda que
    alguna vez tuvo estilo/borde aplicado -- en estas plantillas eso deja
    `max_row` muy por encima del último dato real (sobre todo después de
    delete_rows, que no lo recalcula), lo que producía el 'hueco' al agregar
    filas nuevas. Se escanea por datos de verdad en vez de confiar en
    max_row."""
    bound = upper_bound if upper_bound is not None else ws.max_row
    last = header_row
    for r in range(header_row + 1, bound + 1):
        v = ws.cell(row=r, column=key_col).value
        if v is not None and str(v).strip() != "":
            last = r
    return last


def _compute_full_name(first_name, last_name) -> str:
    """Replica en Python la fórmula real de F (Full Name) en 'planta':
    concatena First Name + Last Name."""
    fn = "" if first_name is None else str(first_name).strip()
    ln = "" if last_name is None else str(last_name).strip()
    return f"{fn} {ln}".strip()


def _compute_age(dob) -> Optional[int]:
    """Replica en Python la fórmula real de V (Age) en 'planta': edad en
    años completos a la fecha de hoy (equivalente a DATEDIF(DOB,HOY,"Y"))."""
    import datetime as _dt
    if dob is None or dob == "":
        return None
    if isinstance(dob, str):
        try:
            dob = _dt.datetime.strptime(dob.strip(), "%Y-%m-%d")
        except ValueError:
            try:
                dob = _dt.datetime.strptime(dob.strip(), "%m/%d/%Y")
            except ValueError:
                return None
    if isinstance(dob, _dt.datetime):
        dob = dob.date()
    if not isinstance(dob, _dt.date):
        return None
    today = _dt.date.today()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))


def _eval_simple_lookup_formula(wb, formula_text: str, row: int, home_sheet: str) -> Optional[object]:
    """Evalúa (sin abrir Excel) una fórmula XLOOKUP/VLOOKUP de una sola tabla
    de referencia -- el patrón dominante en estas plantillas -- leyendo los
    valores reales ya presentes en el propio workbook. Así 'se calcula la
    fórmula que es' de forma genérica, sin tener que adivinar de antemano
    qué representa cada columna. Si la fórmula no calza con estos patrones
    conocidos, devuelve None (la fórmula se deja copiada como antes)."""
    if not isinstance(formula_text, str):
        return None
    text = formula_text.lstrip("=").strip()

    def _split_ref(ref: str):
        ref = ref.strip()
        if "!" in ref:
            sheet, cellref = ref.split("!", 1)
            sheet = sheet.strip("'")
        else:
            sheet, cellref = home_sheet, ref
        return sheet, cellref.replace("$", "")

    def _resolve_scalar(ref: str):
        sheet, cellref = _split_ref(ref)
        if sheet not in wb.sheetnames:
            return None
        m = re.match(r"^([A-Z]+)(\d+)$", cellref)
        if not m:
            return None
        col_letters, row_s = m.groups()
        col = column_index_from_string(col_letters)
        r = int(row_s) if row_s != "0" else row
        return wb[sheet].cell(row=r, column=col).value

    def _build_dict(lookup_ref: str, return_ref: str):
        lsheet, lcell = _split_ref(lookup_ref)
        rsheet, rcell = _split_ref(return_ref)
        if lsheet not in wb.sheetnames or rsheet not in wb.sheetnames:
            return {}
        m_l = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", lcell)
        m_r = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", rcell)
        if not m_l or not m_r:
            return {}
        lcol = column_index_from_string(m_l.group(1))
        rcol = column_index_from_string(m_r.group(1))
        lstart = int(m_l.group(2))
        d = {}
        ws_l, ws_r = wb[lsheet], wb[rsheet]
        r_off = int(m_r.group(2)) - lstart
        for rr in range(lstart, ws_l.max_row + 1):
            k = ws_l.cell(row=rr, column=lcol).value
            if k is None:
                continue
            d.setdefault(str(k).strip(), ws_r.cell(row=rr + r_off, column=rcol).value)
        return d

    m = re.match(r"XLOOKUP\(([^,]+),\s*([^,]+),\s*([^,]+)", text, re.IGNORECASE)
    if not m:
        m = re.match(r"VLOOKUP\(([^,]+),\s*([^,]+),\s*(\d+)", text, re.IGNORECASE)
        if m:
            lookup_val_ref, table_ref, col_idx_s = [g.strip().rstrip(")") for g in m.groups()]
            lsheet, lcell = _split_ref(table_ref)
            m_range = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", lcell)
            if not m_range or lsheet not in wb.sheetnames:
                return None
            lcol = column_index_from_string(m_range.group(1))
            rcol = lcol + int(col_idx_s) - 1
            lstart = int(m_range.group(2))
            d = {}
            ws_l = wb[lsheet]
            for rr in range(lstart, ws_l.max_row + 1):
                k = ws_l.cell(row=rr, column=lcol).value
                if k is None:
                    continue
                d.setdefault(str(k).strip(), ws_l.cell(row=rr, column=rcol).value)
            lookup_val = _resolve_scalar(lookup_val_ref)
            return d.get(str(lookup_val).strip()) if lookup_val is not None else None
        return None

    lookup_val_ref, lookup_arr_ref, return_arr_ref = [g.strip().rstrip(")") for g in m.groups()]
    try:
        lookup_val = _resolve_scalar(lookup_val_ref)
        if lookup_val is None:
            return None
        lookup_dict = _build_dict(lookup_arr_ref, return_arr_ref)
        return lookup_dict.get(str(lookup_val).strip())
    except Exception:
        return None


def _table_info(ws, table_name: str):
    """Ubica una Tabla de Excel por nombre (case-insensitive) y devuelve
    (nombre_real, min_col, min_row, max_col, last_row) usando el rango de la
    tabla — NO ws.max_row, porque algunas hojas tienen filas con formato
    "fantasma" más allá de los datos reales que inflan ws.max_row.

    Si openpyxl no detecta NINGUNA tabla registrada en la hoja (puede pasar
    si un guardado anterior dejó la definición de la Tabla dañada, aunque
    Excel la siga mostrando bien porque es más tolerante), se calcula un
    rango equivalente directo desde los datos: encabezados en la fila 1,
    y la última fila con datos reales en la columna A. `match` queda como
    None en ese caso — el código que llama a esto debe evitar actualizar
    `ws.tables[match].ref` cuando match es None."""
    try:
        names = list(ws.tables.keys())
    except AttributeError:
        names = list(ws.tables)
    match = next((n for n in names if n.strip().lower() == table_name.strip().lower()), None)
    if not match and len(names) == 1:
        # Respaldo 1: si el nombre no calzó exacto (espacio invisible, etc.)
        # pero solo hay una tabla en la hoja, es casi seguro que es esa.
        match = names[0]
    if match:
        tbl = ws.tables[match]
        min_col, min_row, max_col, last_row = range_boundaries(tbl.ref)
        return match, min_col, min_row, max_col, last_row
    if not names:
        # Respaldo 2: no hay ninguna Tabla registrada — inferir el rango
        # directo de los datos reales, sin depender de un objeto Tabla.
        max_col = 1
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value not in (None, ""):
                max_col = c
        last_row = 1
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value not in (None, ""):
                last_row = r
        return None, 1, 1, max_col, last_row
    return None


def _planta_fmt_date(v) -> str:
    """Normaliza una fecha de la template (datetime, serial de Excel, o texto) a DD/MM/YYYY."""
    if v is None or (isinstance(v, float) and pd.isna(v)) or v == "":
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, (int, float)):
        try:
            dt = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(v))
            return dt.strftime("%d/%m/%Y")
        except (ValueError, OverflowError):
            return str(v)
    return str(v).strip()


def _build_planta_row(tpl_row: list, nro: int, formula_row_num: int) -> list:
    """Traduce una fila de la Template_planta (empezando en columna B) a una
    fila completa A:AB de BD_PLANTA. La template ya NO trae el numero
    indicativo en la columna B -- ahora trae el Periodo, y la app genera el
    numero indicativo (Nro) ella misma, consecutivo. F (Full Name) y V (Age)
    NO se tocan -- ya tienen formula en la Base y se diligencian solas. AB
    tampoco se escribe. Estas 3 columnas quedan como None en la lista
    devuelta como senal de "no escribir esta celda"."""

    def t(idx):
        v = tpl_row[idx] if idx < len(tpl_row) else ""
        return "" if v is None or (isinstance(v, float) and pd.isna(v)) else v

    row = [""] * 28  # A .. AB

    row[0] = t(0)                                        # A -- Periodo (tpl B, ahora es el Periodo)
    row[1] = nro                                          # B -- Nro (generado por la app)
    row[2] = t(1)                                       # C -- ID Nr.            (tpl C)
    row[3] = t(2)                                       # D -- First Name        (tpl D)
    row[4] = t(3)                                       # E -- Last Name         (tpl E)
    row[5] = None                                        # F -- Full Name (formula existente, NO se escribe)
    row[6] = _planta_fmt_date(t(8))                      # G -- Date of First Appointment (tpl J)
    row[7] = t(9)                                        # H -- Academic Area     (tpl K)
    row[8] = t(13)                                       # I -- Highest Earned Degree (tpl O)
    row[9] = t(14)                                       # J -- Year (Degree)     (tpl P)
    row[10] = t(15)                                      # K -- University        (tpl Q)
    row[11] = t(16)                                      # L -- Region            (tpl R)
    row[12] = t(17)                                      # M -- Highest Degree    (tpl S)
    row[13] = t(18)                                      # N -- International Degree (tpl T)
    row[14] = t(11)                                      # O -- % devoted to Mission (tpl M)
    row[15] = t(10)                                      # P -- Faculty Ranking   (tpl L)
    row[16] = ""                                         # Q -- Subcategorization (vacio)
    row[17] = t(19)                                      # R -- Field             (tpl U)
    row[18] = t(6)                                       # S -- Country of Birth  (tpl H)
    row[19] = t(7)                                       # T -- Double Nationality (tpl I)
    row[20] = _planta_fmt_date(t(4))                     # U -- Date of Birth     (tpl F)
    row[21] = None                                       # V -- Age (formula existente, NO se escribe)
    row[22] = t(5)                                        # W -- Gender            (tpl G)
    row[23] = t(12)                                       # X -- Faculty Qualific. (tpl N)
    row[24] = "P"                                          # Y -- P/S, fijo "P"
    row[25] = t(20)                                       # Z -- Normal professional Resp. (tpl V)
    row[26] = t(21)                                       # AA -- Notes            (tpl W)
    row[27] = None                                        # AB -- ya no se escribe "PLANTA"

    return row


def _planta_note_style(note: str) -> Optional[str]:
    """'IN IN...' → azul y negrilla · 'OUT IN...' → rojo · si no, sin estilo."""
    n = str(note or "").strip().upper()
    if n.startswith("IN IN"):
        return "blue_bold"
    if n.startswith("OUT IN"):
        return "red"
    return None


def _read_planta_template(uploaded_file) -> pd.DataFrame:
    """Lee la Template_planta.xlsx tal como la define el layout original:
    encabezados en la fila 4, datos desde la fila 6, empezando en la columna B."""
    raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)
    headers = raw.iloc[_PLANTA_TEMPLATE_HEADER_ROW - 1, _PLANTA_TEMPLATE_START_COL:].tolist()
    headers = [str(h).replace("\n", " ").strip() if pd.notna(h) else "" for h in headers]
    data = raw.iloc[_PLANTA_TEMPLATE_DATA_ROW - 1:, _PLANTA_TEMPLATE_START_COL:].copy()
    data.columns = headers
    data = data.dropna(how="all").reset_index(drop=True)
    _validate_template_columns(data, ["First Name", "Last Name", "ID Nr."], "Template_planta.xlsx")
    return data


def _style_planta_preview(df: pd.DataFrame):
    def _row_style(row):
        style = _planta_note_style(row.get("Notes", ""))
        if style == "blue_bold":
            return ["color: #1d4ed8; font-weight: 700;"] * len(row)
        if style == "red":
            return ["color: #dc2626;"] * len(row)
        return [""] * len(row)
    return df.style.apply(_row_style, axis=1)


def _drive_upload_file_bytes(file_id: str, content: bytes) -> Tuple[bool, str]:
    """Sube contenido nuevo sobre un archivo YA EXISTENTE en Drive (lo
    sobreescribe completo), autenticado con la service account."""
    token = _get_gspread_access_token()
    if not token:
        return False, "No hay credenciales configuradas (falta st.secrets['gcp_service_account'])."
    url = f"https://www.googleapis.com/upload/drive/v3/files/{file_id}?uploadType=media"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    try:
        resp = requests.patch(url, headers=headers, data=content, timeout=120)
        if resp.status_code == 200:
            return True, ""
        return False, f"HTTP {resp.status_code}: {resp.text[:300]}"
    except Exception as e:
        return False, str(e)


def push_planta_updates(new_rows_df: pd.DataFrame) -> Tuple[bool, str]:
    """Escribe las filas nuevas en la hoja 'planta' de BD_profesores.xlsx.
    La template ahora trae el Periodo en cada fila (columna B), asi que ya
    no se pide un periodo global aparte: se toma directo de cada fila, se
    borran las filas existentes que compartan esos periodos (reemplazo,
    igual que antes), se agregan las nuevas al final con un numero
    indicativo (Nro) consecutivo generado por la app, se aplica el
    resaltado de Notes (IN IN -> azul negrilla, OUT IN -> rojo), y se sube
    el archivo completo de vuelta a Drive.

    Nota: como BD_profesores.xlsx es un archivo .xlsx normal (no un Google
    Sheet nativo), no se puede editar celda por celda via API -- hay que
    descargar el archivo entero, modificarlo con openpyxl, y volver a
    subirlo completo. Esto reemplaza el archivo tal cual queda guardado; si
    alguien mas lo esta editando en Excel/Sheets al mismo tiempo, esos
    cambios se perderian (ultimo en guardar gana)."""
    if not _OPENPYXL_OK:
        return False, "Falta la libreria `openpyxl` en el entorno (agregala a requirements.txt)."

    token = _get_gspread_access_token()
    if not token:
        return False, (
            "No hay credenciales configuradas para escribir en Drive. "
            "Falta `st.secrets['gcp_service_account']` (ver instrucciones abajo)."
        )
    try:
        raw_bytes = _download_drive_file_bytes(PROFESORES_FILE_ID)
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
        if "planta" not in wb.sheetnames:
            return False, "No encontre la hoja 'planta' dentro de BD_profesores.xlsx."
        ws = wb["planta"]

        periodos = set(
            str(r[0]).strip() for r in new_rows_df.itertuples(index=False, name=None)
            if r[0] is not None and str(r[0]).strip() != ""
        )

        # 1) Borra filas existentes con esos periodos (columna A), de abajo hacia arriba
        rows_to_delete = [
            r for r in range(2, ws.max_row + 1)
            if str(ws.cell(row=r, column=1).value or "").strip().replace(".0", "") in periodos
        ]
        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r)

        # 1.5) Numero indicativo (Nro) consecutivo: continua desde el maximo
        # existente en la columna B, ya que la template no lo trae mas.
        max_nro = 0
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=2).value
            try:
                max_nro = max(max_nro, int(float(v)))
            except (TypeError, ValueError):
                continue

        # 2) Agrega las filas nuevas justo debajo del ultimo dato REAL (no de
        # ws.max_row, que queda inflado por formato/bordes que sobreviven a
        # delete_rows y dejaba un hueco de filas vacias antes de esta fila).
        last_real_row = _last_data_row(ws, key_col=1)
        append_start = last_real_row + 1
        rows = [
            _build_planta_row(list(r), max_nro + 1 + i, append_start + i)
            for i, r in enumerate(new_rows_df.itertuples(index=False, name=None))
        ]

        base_font = _BASE_ARIAL_FONT
        blue_bold_font = Font(name="Arial", size=11, color="1D4ED8", bold=True)
        red_font = Font(name="Arial", size=11, color="DC2626")
        thin = Side(style="thin")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fv_fill = PatternFill(fill_type="solid", fgColor="F1CEEE")
        fv_align = Alignment(horizontal="left")

        for i, row_vals in enumerate(rows):
            rn = append_start + i
            style = _planta_note_style(row_vals[26])  # indice 26 = columna AA (Notes)
            font = blue_bold_font if style == "blue_bold" else red_font if style == "red" else base_font
            for c, val in enumerate(row_vals, start=1):
                if val is None:
                    if c == 6:      # F -- Full Name: se calcula en Python (First + Last), no formula
                        val = _compute_full_name(row_vals[3], row_vals[4])
                    elif c == 22:    # V -- Age: se calcula en Python desde Date of Birth (col U, indice 20)
                        val = _compute_age(row_vals[20])
                        if val is None:
                            continue
                    else:
                        continue
                cell = ws.cell(row=rn, column=c, value=val)
                cell.font = font
                cell.border = thin_border
                if c in (6, 22):  # F (Full Name) y V (Age): relleno rosado + alineado a la izquierda
                    cell.fill = fv_fill
                    cell.alignment = fv_align

        # 2.6) Repara de paso cualquier fila EXISTENTE de F/V que haya quedado
        # como texto de formula sin resolver (de cargas anteriores a este
        # cambio) o mal alineada, para que toda la columna quede consistente.
        for r in range(2, append_start):
            for c in (6, 22):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    if c == 6:
                        cell.value = _compute_full_name(ws.cell(row=r, column=4).value, ws.cell(row=r, column=5).value)
                    else:
                        age = _compute_age(ws.cell(row=r, column=21).value)
                        if age is not None:
                            cell.value = age
                cell.alignment = fv_align

        # 3.5) Extiende la Tabla de Excel "tabla_planta" para que incluya las
        # filas nuevas -- sin esto, aunque las celdas queden vacias, Excel no
        # las reconoce como parte de la tabla y no autocompleta las columnas
        # calculadas (F, V).
        new_last_row = append_start + len(rows) - 1 if rows else last_real_row
        try:
            table_names = list(ws.tables.keys())
        except AttributeError:
            table_names = list(ws.tables)
        match = next((n for n in table_names if n.strip().lower() == "tabla_planta"), None)
        if match:
            tbl = ws.tables[match]
            min_col, min_row, max_col, _old_max_row = range_boundaries(tbl.ref)
            tbl.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_last_row}"

        wb.calculation.fullCalcOnLoad = True  # fuerza recalculo de formulas al abrir en Excel
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        # 4) Sube el archivo completo de vuelta a Drive
        ok, err = _drive_upload_file_bytes(PROFESORES_FILE_ID, buf.getvalue())
        if not ok:
            return False, f"Error al subir el archivo actualizado a Drive: {err}"

        load_data.clear()
        qual_load_planta.clear()
        area_load_fulltime.clear()
        demo_load_fulltime.clear()
        _download_drive_file_bytes.clear()

        periodos_txt = ", ".join(sorted(periodos)) if periodos else "?"
        msg = f"\u2713 BD_profesores.xlsx (hoja 'planta') actualizada \u2014 {len(rows)} filas para el/los periodo(s) {periodos_txt}."
        if not match:
            msg += (
                " \u26a0\ufe0f No encontre una Tabla de Excel llamada 'tabla_planta' en la hoja -- "
                "las filas se agregaron igual, pero quedaran fuera de la tabla."
            )
        if not formulas_ok:
            msg += (
                " \u26a0\ufe0f No encontre una formula existente en F/V de la ultima fila para "
                "copiar -- esas columnas quedaron en blanco en las filas nuevas."
            )
        return True, msg
    except Exception as e:
        return False, f"Error al escribir en la hoja 'planta': {e}"


def _read_cartelera_template(uploaded_file) -> pd.DataFrame:
    """Template_cartelera: columnas B..H → Period, Campus, Course, Sec,
    Credits, Full Course Name, Professor (la template ahora trae los
    encabezados en inglés; se renombran aquí a los nombres en español que
    usa el resto del flujo de carga, sin tener que tocar cada referencia)."""
    df_ = _read_generic_template(uploaded_file)
    rename_map = {
        "Period": "Periodo",
        "Course": "Materia",
        "Sec": "Secc",
        "Credits": "Créditos",
        "Full Course Name": "Nombre largo curso",
        "Professor": "Profesor",
    }
    df_ = df_.rename(columns={k: v for k, v in rename_map.items() if k in df_.columns})
    _validate_template_columns(df_, ["Periodo", "Materia", "Profesor"], "Template_cartelera.xlsx")
    return df_


def _read_cursos_nuevos_template(uploaded_file) -> pd.DataFrame:
    """Template_Cursos_Nuevos: columnas B..E → Código Materia, Créditos,
    Nombre largo curso, Area del curso."""
    df_ = _read_generic_template(uploaded_file)
    _validate_template_columns(df_, ["Código Materia", "Nombre largo curso"], "Template_cursos_nuevos.xlsx")
    return df_


@st.cache_data(ttl=60)
def _load_cursos_area_map() -> Dict[str, str]:
    """Código Materia (columna A de 'cursos') → Area del curso (columna D),
    para saber en el preview qué cursos de la template ya existen y cuáles no."""
    raw = io.BytesIO(_download_drive_file_bytes(CARTELERA_FILE_ID))
    dfc = pd.read_excel(raw, sheet_name="cursos")
    dfc.columns = dfc.columns.str.strip()
    key = dfc["Código Materia"].astype(str).str.strip()
    return dict(zip(key, dfc["Area del curso"]))


# ── Profesores (lookup + carga de nuevos) ───────────────────────────────
def _read_profesores_nuevos_template(uploaded_file) -> pd.DataFrame:
    """Template_Profesores_Nuevos: columnas B..T → Profesor, ID, AREA_PROFESOR,
    GÉNERO, TIPO, P/S, PLANTA_CATEDRA, Date of First Appointment, Highest
    Earned Degree, Highest Degree Year Earned, Highest Degree, University,
    Region, International Degree?, Normal Professional Resp., Basis for
    qualification, Nationality, Date of birth, Years Industry experience."""
    df_ = _read_generic_template(uploaded_file)
    _validate_template_columns(df_, ["Profesor", "ID", "AREA_PROFESOR"], "Template_profesores_nuevos.xlsx")
    return df_


@st.cache_data(ttl=60)
def _load_profesores_lookup() -> Dict[str, Tuple]:
    """Profesor (columna A de 'Info. Profesores', normalizado) → (ID,
    AREA_PROFESOR, TIPO, P/S) desde las columnas B, C, E, F."""
    raw = io.BytesIO(_download_drive_file_bytes(PROFESORES_FILE_ID))
    dfp = pd.read_excel(raw, sheet_name="Info. Profesores")
    dfp.columns = dfp.columns.str.strip()
    key = dfp["Profesor"].astype(str).str.strip().str.upper()
    vals = list(zip(dfp["ID"], dfp["AREA_PROFESOR"], dfp["TIPO"], dfp["P/S"]))
    return dict(zip(key, vals))


def _build_prefilled_cursos_template(missing_rows: pd.DataFrame) -> bytes:
    """Descarga la Template_cursos_nuevos.xlsx real y prellena Código Materia
    (B), Créditos (C) y Nombre largo curso (D) con lo que ya se conoce,
    desde la fila 6."""
    raw = _download_drive_file_bytes(TEMPLATE_CURSOS_NUEVOS_FILE_ID)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb[wb.sheetnames[0]]
    for i, (_, row) in enumerate(missing_rows.iterrows()):
        ws.cell(row=6 + i, column=2, value=row["Materia"])
        ws.cell(row=6 + i, column=3, value=row["Créditos"])
        ws.cell(row=6 + i, column=4, value=row["Nombre largo curso"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_cursos_download() -> bytes:
    """Devuelve la hoja 'cursos' completa (de BD_cartelera.xlsx) como .xlsx descargable."""
    raw = io.BytesIO(_download_drive_file_bytes(CARTELERA_FILE_ID))
    dfc = pd.read_excel(raw, sheet_name="cursos")
    return _xlsx_bytes(dfc, sheet_name="cursos")


def _build_info_profesores_download() -> bytes:
    """Devuelve la hoja 'Info. Profesores' completa como un .xlsx descargable."""
    raw = io.BytesIO(_download_drive_file_bytes(PROFESORES_FILE_ID))
    dfi = pd.read_excel(raw, sheet_name="Info. Profesores")
    return _xlsx_bytes(dfi, sheet_name="Info. Profesores")


def _build_prefilled_profesores_template(missing_names: List[str]) -> bytes:
    """Descarga la Template_profesores_nuevos.xlsx real y prellena la columna
    Profesor (B) con los nombres que no se encontraron, desde la fila 6."""
    raw = _download_drive_file_bytes(TEMPLATE_PROFESORES_NUEVOS_FILE_ID)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb[wb.sheetnames[0]]
    for i, name in enumerate(missing_names):
        ws.cell(row=6 + i, column=2, value=name)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def repair_planta_catedra_cache() -> Tuple[bool, str]:
    """Reparación de una sola vez: la columna PLANTA_CATEDRA de 'Faculty
    Distribution' es una fórmula (=IF(COUNTIFS(planta!...)>0,"PLANTA","CÁTEDRA"))
    cuyo valor en caché se perdió en guardados automatizados anteriores —
    openpyxl no recalcula fórmulas, así que quedó en blanco para todos los
    periodos históricos (solo el más reciente, escrito con valor literal,
    se veía bien). Esta función calcula el valor real cruzando (Periodo, ID)
    contra la hoja 'planta' y lo escribe como literal, para todas las filas
    donde la celda siga siendo una fórmula sin resolver."""
    if not _OPENPYXL_OK:
        return False, "Falta la librería `openpyxl` en el entorno."
    token = _get_gspread_access_token()
    if not token:
        return False, "No hay credenciales configuradas para escribir en Drive."
    try:
        raw_bytes = _download_drive_file_bytes(PROFESORES_FILE_ID)
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
        ws_fd = wb["Faculty Distribution"]
        ws_planta = wb["planta"]

        headers_planta = [c.value for c in ws_planta[1]]
        col_periodo_p = headers_planta.index("Periodo") + 1 if "Periodo" in headers_planta else 1
        col_id_p = headers_planta.index("ID Nr.") + 1 if "ID Nr." in headers_planta else 3
        planta_pairs = set()
        for r in range(2, ws_planta.max_row + 1):
            periodo = ws_planta.cell(row=r, column=col_periodo_p).value
            idval = ws_planta.cell(row=r, column=col_id_p).value
            if periodo is not None and idval is not None:
                planta_pairs.add((str(periodo).strip(), str(idval).strip()))

        headers_fd = [c.value for c in ws_fd[1]]
        col_periodo_fd = headers_fd.index("Semestre") + 1
        col_id_fd = headers_fd.index("ID") + 1
        col_pc_fd = headers_fd.index("PLANTA_CATEDRA") + 1

        n_fixed = 0
        for r in range(2, ws_fd.max_row + 1):
            periodo = ws_fd.cell(row=r, column=col_periodo_fd).value
            if periodo is None:
                continue
            current = ws_fd.cell(row=r, column=col_pc_fd).value
            if isinstance(current, str) and current.startswith("="):
                idval = ws_fd.cell(row=r, column=col_id_fd).value
                key = (str(periodo).strip(), str(idval).strip())
                computed = "PLANTA" if key in planta_pairs else "CÁTEDRA"
                cell = ws_fd.cell(row=r, column=col_pc_fd, value=computed)
                cell.font = _BASE_ARIAL_FONT
                n_fixed += 1

        if n_fixed == 0:
            return True, "No había filas con fórmula sin resolver — nada que reparar."

        wb.calculation.fullCalcOnLoad = True
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        ok, err = _drive_upload_file_bytes(PROFESORES_FILE_ID, buf.getvalue())
        if not ok:
            return False, f"Error al subir el archivo reparado a Drive: {err}"

        _download_drive_file_bytes.clear()
        return True, f"✓ Reparadas {n_fixed} fila(s) de PLANTA_CATEDRA en Faculty Distribution."
    except Exception as e:
        return False, f"Error al reparar PLANTA_CATEDRA: {e}"


def push_faculty_distribution_updates(periodo_to_ids: Dict[str, List]) -> Tuple[bool, str]:
    """Agrega a 'Faculty Distribution' (BD_profesores.xlsx) una fila por cada
    ID único que quedó en la cartelera recién cargada, agrupado por periodo:
    A=Periodo, C=ID (valores directos). B,D,E,F,G,H se escriben como VALOR
    LITERAL (no como fórmula copiada) — se calculan en Python con la misma
    lógica que ya tenían esas fórmulas (lookup contra 'Info. Profesores' y
    'planta'), con fondo #caedfb. Se dejó de copiar la fórmula porque
    openpyxl no la recalcula sola: la app (que lee con pandas) veía esas
    columnas en blanco hasta que alguien abría el archivo manualmente en
    Excel. No agrega un (Periodo, ID) que ya exista en la hoja."""
    if not _OPENPYXL_OK:
        return False, "Falta la librería `openpyxl` en el entorno."
    token = _get_gspread_access_token()
    if not token:
        return False, "No hay credenciales configuradas para escribir en Drive."
    try:
        raw_bytes = _download_drive_file_bytes(PROFESORES_FILE_ID)
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
        if "Faculty Distribution" not in wb.sheetnames:
            return False, "No encontré la hoja 'Faculty Distribution' en BD_profesores.xlsx."
        if "Info. Profesores" not in wb.sheetnames or "planta" not in wb.sheetnames:
            return False, "No encontré 'Info. Profesores' y/o 'planta' en BD_profesores.xlsx."
        ws = wb["Faculty Distribution"]
        ws_info = wb["Info. Profesores"]
        ws_planta = wb["planta"]

        info = _table_info(ws, "tabla_faculty_distribution")
        if not info:
            return False, "No encontré la Tabla de Excel 'tabla_faculty_distribution'."
        match, min_col, min_row, max_col, last_row = info

        base_font = _BASE_ARIAL_FONT
        calc_fill = PatternFill(fill_type="solid", fgColor="CAEDFB")

        # ── Tablas de referencia reales, leídas del propio archivo ──
        # Info. Profesores: A=Profesor,B=ID,C=AREA_PROFESOR,D=GÉNERO,E=TIPO,F=P/S
        info_lookup: Dict[str, Tuple] = {}
        for r in range(2, ws_info.max_row + 1):
            pid = ws_info.cell(row=r, column=2).value
            if pid is None:
                continue
            key = str(pid).strip()
            info_lookup[key] = (
                ws_info.cell(row=r, column=1).value,  # Profesor
                ws_info.cell(row=r, column=3).value,  # AREA_PROFESOR
                ws_info.cell(row=r, column=4).value,  # GÉNERO
                ws_info.cell(row=r, column=5).value,  # TIPO
                ws_info.cell(row=r, column=6).value,  # P/S
            )

        # planta: A=Periodo,C=ID Nr.,X=Faculty Qualific.(24),Y=P/S(25)
        planta_lookup: Dict[Tuple[str, str], Tuple] = {}
        for r in range(2, ws_planta.max_row + 1):
            pid = ws_planta.cell(row=r, column=3).value
            per = ws_planta.cell(row=r, column=1).value
            if pid is None or per is None:
                continue
            key = (str(per).strip(), str(pid).strip())
            planta_lookup[key] = (
                ws_planta.cell(row=r, column=24).value,  # Faculty Qualific. (TIPO en planta)
                ws_planta.cell(row=r, column=25).value,  # P/S
            )

        existing_pairs = set()
        for r in range(2, last_row + 1):
            p = str(ws.cell(row=r, column=1).value or "").strip()
            i = str(ws.cell(row=r, column=3).value or "").strip()
            existing_pairs.add((p, i))

        append_start = last_row + 1
        n_written = 0
        for periodo, ids in periodo_to_ids.items():
            for prof_id in ids:
                periodo_s, id_s = str(periodo).strip(), str(prof_id).strip()
                pair = (periodo_s, id_s)
                if pair in existing_pairs:
                    continue
                rn = append_start + n_written
                for col, val in [(1, periodo), (3, prof_id)]:
                    cell = ws.cell(row=rn, column=col, value=val)
                    cell.font = base_font

                info_row = info_lookup.get(id_s, (None, None, None, None, None))
                profesor, area_prof, genero, tipo_info, ps_info = info_row
                planta_vals = planta_lookup.get((periodo_s, id_s))
                planta_catedra = "PLANTA" if planta_vals is not None else "CÁTEDRA"
                if planta_vals is not None:
                    tipo_val, ps_val = planta_vals
                else:
                    tipo_val, ps_val = tipo_info, ps_info

                calc_vals = {2: profesor, 4: area_prof, 5: genero, 6: tipo_val, 7: ps_val, 8: planta_catedra}
                for col, val in calc_vals.items():
                    cell = ws.cell(row=rn, column=col, value=val)
                    cell.font = base_font
                    cell.fill = calc_fill

                existing_pairs.add(pair)
                n_written += 1

        if n_written == 0:
            return True, "✓ Faculty Distribution: no había IDs nuevos que agregar (ya estaban todos)."

        new_last_row = append_start + n_written - 1
        if match:
            ws.tables[match].ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_last_row}"

        wb.calculation.fullCalcOnLoad = True
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        ok, err = _drive_upload_file_bytes(PROFESORES_FILE_ID, buf.getvalue())
        if not ok:
            return False, f"Error al subir el archivo actualizado a Drive: {err}"

        _download_drive_file_bytes.clear()

        return True, f"✓ Faculty Distribution actualizada — {n_written} fila(s) nueva(s)."
    except Exception as e:
        return False, f"Error al escribir en Faculty Distribution: {e}"


def push_profesores_updates(new_profs_df: pd.DataFrame) -> Tuple[bool, str]:
    """Agrega profesores nuevos a la hoja 'Info. Profesores' de
    BD_profesores.xlsx. Columnas A-F, H-Q, S vienen directo de la template
    (los campos opcionales que queden vacíos se completan con "TBD", igual
    que la convención ya usada en el resto del archivo). R (Age) SÍ tiene
    fórmula real (DATEDIF sobre Date of birth) — se copia y traslada igual
    que F/V en planta, con fondo #caedfb."""
    if not _OPENPYXL_OK:
        return False, "Falta la librería `openpyxl` en el entorno."
    token = _get_gspread_access_token()
    if not token:
        return False, "No hay credenciales configuradas para escribir en Drive."
    try:
        raw_bytes = _download_drive_file_bytes(PROFESORES_FILE_ID)
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
        if "Info. Profesores" not in wb.sheetnames:
            return False, "No encontré la hoja 'Info. Profesores' en BD_profesores.xlsx."
        ws = wb["Info. Profesores"]

        info = _table_info(ws, "tabla_profesores")
        if not info:
            return False, "No encontré la Tabla de Excel 'tabla_profesores'."
        match, min_col, min_row, max_col, last_row = info

        base_font = _BASE_ARIAL_FONT
        age_fill = PatternFill(fill_type="solid", fgColor="CAEDFB")

        tpl_age_text, age_is_array = _get_formula_text(ws.cell(row=last_row, column=18))
        age_template_row = last_row
        age_ok = bool(tpl_age_text)

        # Template B..T → Info.Profesores A,B,C,D,E,F,(H sin destino=PLANTA_CATEDRA),G,H,I,J,K,L,M,N,O,P,Q,S
        col_map = {
            0: 1, 1: 2, 2: 3, 3: 4, 4: 5, 5: 6,        # Profesor,ID,AREA_PROFESOR,GÉNERO,TIPO,P/S
            # idx 6 = PLANTA_CATEDRA -> sin columna destino en Info. Profesores, se omite
            7: 7, 8: 8, 9: 9, 10: 10, 11: 11, 12: 12,
            13: 13, 14: 14, 15: 15, 16: 16, 17: 17, 18: 19,  # ...hasta S (Years Industry exp -> col 19)
        }
        required_idx = {0: "Profesor", 1: "ID", 2: "AREA_PROFESOR", 3: "GÉNERO", 4: "TIPO", 5: "P/S"}

        append_start = last_row + 1
        n_written = 0
        for i, r in enumerate(new_profs_df.itertuples(index=False, name=None)):
            rn = append_start + i
            missing_required = [
                label for idx, label in required_idx.items()
                if idx >= len(r) or r[idx] is None or (isinstance(r[idx], float) and pd.isna(r[idx])) or str(r[idx]).strip() == ""
            ]
            if missing_required:
                return False, (
                    f"Fila {i+1} de la template de profesores: faltan campos obligatorios "
                    f"({', '.join(missing_required)}). No se guardó nada — corrige y vuelve a subir."
                )
            for tpl_idx, dest_col in col_map.items():
                val = r[tpl_idx] if tpl_idx < len(r) else None
                if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == "":
                    val = "TBD" if tpl_idx not in required_idx else val
                cell = ws.cell(row=rn, column=dest_col, value=val)
                cell.font = base_font
            # R — Age: fórmula real copiada/trasladada (o "TBD" si no había de dónde copiarla)
            if age_ok:
                _write_translated_formula(ws, 18, rn, tpl_age_text, age_is_array, f"R{age_template_row}", f"R{rn}")
            else:
                ws.cell(row=rn, column=18, value="TBD")
            age_cell = ws.cell(row=rn, column=18)
            age_cell.font = base_font
            age_cell.fill = age_fill
            n_written += 1

        if n_written == 0:
            return True, "No había profesores nuevos que agregar."

        new_last_row = append_start + n_written - 1
        if match:
            ws.tables[match].ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_last_row}"

        wb.calculation.fullCalcOnLoad = True
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        ok, err = _drive_upload_file_bytes(PROFESORES_FILE_ID, buf.getvalue())
        if not ok:
            return False, f"Error al subir el archivo actualizado a Drive: {err}"

        _download_drive_file_bytes.clear()
        _load_profesores_lookup.clear()

        return True, f"✓ Info. Profesores actualizada — {n_written} profesor(es) nuevo(s)."
    except Exception as e:
        return False, f"Error al escribir en Info. Profesores: {e}"


def _build_academic_area_pct_table(df_cart_period: pd.DataFrame) -> pd.DataFrame:
    """Replica independiente de la tabla 'Academic Area / %P / %S / %SA /
    %OTHER' de Qualifications, para un DataFrame de cartelera ya filtrado a
    un solo periodo. No depende del estado de la página en vivo."""
    def _resolve(df, target):
        t = target.strip().casefold()
        for c in df.columns:
            if c.strip().casefold() == t:
                return c
        return None

    def _get_any(df, *cands):
        for c in cands:
            if _resolve(df, c):
                return _resolve(df, c)
        return None

    def normalize_ps(val):
        v = str(val).strip().lower()
        if v in {"p", "participating", "participante", "participating faculty"}:
            return "P"
        if v in {"s", "supporting", "soporte", "supporting faculty"}:
            return "S"
        return ""

    def normalize_tipo(val):
        v = str(val).strip().lower()
        if v in {"sa", "scholarly academics", "scholarly academic"}:
            return "SA"
        if v in {"pa", "practice academics", "practice academic"}:
            return "PA"
        if v in {"sp", "scholarly practitioners", "scholarly practitioner"}:
            return "SP"
        if v in {"ip", "instructional practitioners", "instructional practitioner"}:
            return "IP"
        if v in {"o", "other", "others", "otro", "otros"}:
            return "OTHER"
        m = re.search(r"\b(sa|pa|sp|ip|o|other)\b", v)
        if m:
            code = m.group(1).upper()
            return "OTHER" if code in {"O", "OTHER"} else code
        return "OTHER"

    col_cred = _get_any(df_cart_period, "Créditos", "Creditos", "Credits")
    col_tipo = _get_any(df_cart_period, "Type", "Tipo", "TIPO")
    col_ps = _get_any(df_cart_period, "P/S", "PS")
    col_area = _get_any(df_cart_period, "Course Area", "Area del curso", "Área")
    if not all([col_cred, col_tipo, col_ps, col_area]) or df_cart_period.empty:
        return pd.DataFrame(columns=["Academic Area", "%P", "%S", "%SA", "%OTHER"])

    d = df_cart_period.copy()
    d["_CRED"] = pd.to_numeric(d[col_cred], errors="coerce").fillna(0.0)
    d["_TIPO"] = d[col_tipo].astype(str).map(normalize_tipo)
    d["_PS"] = d[col_ps].astype(str).map(normalize_ps)
    d["_AREA"] = d[col_area].astype(str).str.strip()

    tipo_pivot = d.groupby(["_AREA", "_TIPO"])["_CRED"].sum().unstack(fill_value=0.0)
    for k in ["SA", "PA", "SP", "IP", "OTHER"]:
        if k not in tipo_pivot.columns:
            tipo_pivot[k] = 0.0
    ps_pivot = d.groupby(["_AREA", "_PS"])["_CRED"].sum().unstack(fill_value=0.0)
    for k in ["P", "S"]:
        if k not in ps_pivot.columns:
            ps_pivot[k] = 0.0

    den_ps = (ps_pivot["P"] + ps_pivot["S"]).replace(0, pd.NA)
    p_share = (ps_pivot["P"] / den_ps) * 100
    denom_q = tipo_pivot.sum(axis=1).replace(0, pd.NA)

    out = pd.DataFrame({
        "Academic Area": tipo_pivot.index,
        "%P": p_share,
        "%S": 100 - p_share,
        "%SA": (tipo_pivot["SA"] / denom_q) * 100,
        "%OTHER": (tipo_pivot["OTHER"] / denom_q) * 100,
    }).fillna(0.0).round(1).reset_index(drop=True)

    tot_p, tot_s = ps_pivot["P"].sum(), ps_pivot["S"].sum()
    tot_den = tot_p + tot_s
    p_tot = (tot_p / tot_den * 100) if tot_den else 0.0
    tipo_sums = tipo_pivot.sum(axis=0)
    denom_tot = float(tipo_sums.sum())
    total_row = pd.DataFrame([{
        "Academic Area": "TOTAL",
        "%P": round(p_tot, 1), "%S": round(100 - p_tot, 1),
        "%SA": round((tipo_sums["SA"] / denom_tot * 100) if denom_tot else 0.0, 1),
        "%OTHER": round((tipo_sums["OTHER"] / denom_tot * 100) if denom_tot else 0.0, 1),
    }])
    return pd.concat([out, total_row], ignore_index=True)


def _build_bsq_report_tables(df_fd_period: pd.DataFrame):
    """Replica independiente de las 2 tablas de BSQ Compensation de
    Qualifications (participating/supporting por categoría de calificación),
    para un DataFrame de Faculty Distribution ya filtrado a un solo periodo."""
    cats = ["SA", "PA", "SP", "IP", "OTHER"]

    def _norm_tipo(v):
        v = str(v or "").strip().upper()
        return v if v in cats else "OTHER"

    def _norm_ps(v):
        v = str(v or "").strip().upper()
        return "P" if v.startswith("P") else ("S" if v.startswith("S") else "")

    def _norm_ft(v):
        v = str(v or "").strip().upper()
        return "PLANTA" if "PLANTA" in v else ("CÁTEDRA" if v else "")

    d = df_fd_period.copy()
    tipo_col = next((c for c in d.columns if c.strip().upper() in ("TIPO", "TYPE")), None)
    ps_col = next((c for c in d.columns if c.strip().upper() in ("P/S", "PS")), None)
    ft_col = next((c for c in d.columns if "PLANTA" in c.strip().upper() or "CATEDRA" in c.strip().upper().replace("Á", "A")), None)
    if not tipo_col or not ps_col:
        empty = pd.DataFrame(columns=["Row"] + cats + ["TOTAL"])
        return empty, empty

    d["_TIPO"] = d[tipo_col].map(_norm_tipo)
    d["_PS"] = d[ps_col].map(_norm_ps)
    d["_FT"] = d[ft_col].map(_norm_ft) if ft_col else ""

    def _counts(mask):
        sub = d[mask]
        c = sub["_TIPO"].value_counts()
        row = {k: int(c.get(k, 0)) for k in cats}
        row["TOTAL"] = sum(row.values())
        return row

    r7a, r7b = _counts(d["_PS"] == "P"), _counts(d["_PS"] == "S")
    tbl7 = pd.DataFrame([
        {"Row": "a. Participating faculty members", **r7a},
        {"Row": "b. Supporting faculty members", **r7b},
    ])[["Row"] + cats + ["TOTAL"]]

    r8a = _counts((d["_PS"] == "P") & (d["_FT"] == "PLANTA"))
    r8b = _counts((d["_PS"] == "P") & (d["_FT"] == "CÁTEDRA"))
    r8c = {k: r8a.get(k, 0) + r8b.get(k, 0) for k in cats + ["TOTAL"]}
    r8d = _counts((d["_PS"] == "S") & (d["_FT"] == "PLANTA"))
    r8e = _counts((d["_PS"] == "S") & (d["_FT"] == "CÁTEDRA"))
    r8f = {k: r8d.get(k, 0) + r8e.get(k, 0) for k in cats + ["TOTAL"]}

    tbl8 = pd.DataFrame([
        {"Row": "a. Full-time Participating faculty members", **r8a},
        {"Row": "b. Part-time Participating faculty members", **r8b},
        {"Row": "c. Total Participating faculty members", **r8c},
        {"Row": "d. Full-time Supporting faculty members", **r8d},
        {"Row": "e. Part-time Supporting faculty members", **r8e},
        {"Row": "f. Total Supporting faculty members", **r8f},
    ])[["Row"] + cats + ["TOTAL"]]

    return tbl7, tbl8


def _build_cartelera_save_report(target_period: str, new_courses_df: pd.DataFrame, new_profs_df: pd.DataFrame) -> bytes:
    """Reporte de 5 hojas generado justo después de guardar Cartelera con éxito:
    1) Profesores (Faculty Distribution completa, filtrada al periodo subido)
    2) Profesores nuevos agregados en esta carga
    3) Cartelera completa (sin filtrar)
    4) Cursos nuevos agregados en esta carga
    5) Qualifications — las 3 tablas dinámicas de la página (Academic Area
       %P/%S/%SA/%OTHER, y las 2 de BSQ Compensation), para el mismo periodo"""
    raw_fd = io.BytesIO(_download_drive_file_bytes(PROFESORES_FILE_ID))
    df_fd = pd.read_excel(raw_fd, sheet_name="Faculty Distribution")
    df_fd_period = df_fd[df_fd["Semestre"].astype(str).str.strip() == str(target_period).strip()].copy()

    raw_cart = io.BytesIO(_download_drive_file_bytes(CARTELERA_FILE_ID))
    df_cart_full = pd.read_excel(raw_cart, sheet_name="cartelera")
    period_nodash = str(target_period).strip().replace("-", "")
    cart_period_mask = df_cart_full["Periodo"].astype(str).str.strip().isin([str(target_period).strip(), period_nodash])
    df_cart_period = df_cart_full[cart_period_mask].copy()

    tbl_area = _build_academic_area_pct_table(df_cart_period)
    tbl7, tbl8 = _build_bsq_report_tables(df_fd_period)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf) as writer:
        df_fd_period.to_excel(writer, index=False, sheet_name="Profesores")
        (new_profs_df if new_profs_df is not None else pd.DataFrame()).to_excel(
            writer, index=False, sheet_name="Profesores Nuevos")
        df_cart_period.to_excel(writer, index=False, sheet_name="Cartelera")
        (new_courses_df if new_courses_df is not None else pd.DataFrame()).to_excel(
            writer, index=False, sheet_name="Cursos Nuevos")
        tbl_area.to_excel(writer, index=False, sheet_name="Qualifications", startrow=0)
        tbl7.to_excel(writer, index=False, sheet_name="Qualifications", startrow=len(tbl_area) + 3)
        tbl8.to_excel(writer, index=False, sheet_name="Qualifications", startrow=len(tbl_area) + len(tbl7) + 6)
    buf.seek(0)
    return buf.getvalue()


def push_cartelera_updates(cartelera_df: pd.DataFrame, new_courses_df: pd.DataFrame,
                            profesor_lookup: Optional[Dict[str, Tuple]] = None,
                            area_map: Optional[Dict[str, str]] = None) -> Tuple[bool, str]:
    """1) Si hay cursos nuevos, los agrega a 'cursos' (A-D valores directos,
    sin borde; E-F son fórmulas de matriz existentes, copiadas/trasladadas,
    con fondo #c1f4e5, sin borde).
    2) Agrega las filas de cartelera: A,C-G,L directos (sin fondo especial).
    H (Area del curso) y M,N,O,P (ID, AREA_PROFESOR, TIPO, P/S) se buscan y
    escriben como valor literal, con fondo #f1ceee. B,I,J,K,Q,R,S,T,U,V,W
    TAMBIÉN se escriben como VALOR LITERAL (no como fórmula copiada) — se
    calculan en Python usando las mismas tablas de referencia reales que usan
    esas fórmulas (AD:AE de cartelera, D:E y L:M de cursos, A:C de
    'programas'), con fondo #caedfb. Se dejó de copiar la fórmula porque
    openpyxl no la recalcula: al escribir solo el texto de la fórmula, Excel
    nunca la evalúa hasta que alguien abre el archivo manualmente, y hasta
    entonces la app (que lee con pandas) veía esas columnas en blanco. Sin
    bordes en ninguna celda nueva.
    3) Sube BD_cartelera.xlsx actualizado a Drive."""
    if not _OPENPYXL_OK:
        return False, "Falta la librería `openpyxl` en el entorno."
    token = _get_gspread_access_token()
    if not token:
        return False, (
            "No hay credenciales configuradas para escribir en Drive. "
            "Falta `st.secrets['gcp_service_account']`."
        )
    try:
        raw_bytes = _download_drive_file_bytes(CARTELERA_FILE_ID)
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
        if "cartelera" not in wb.sheetnames or "cursos" not in wb.sheetnames:
            return False, "No encontré las hojas 'cartelera' y/o 'cursos' en BD_cartelera.xlsx."
        ws_cart = wb["cartelera"]
        ws_cursos = wb["cursos"]
        ws_programas = wb["programas"] if "programas" in wb.sheetnames else None

        base_font = _BASE_ARIAL_FONT
        area_fill = PatternFill(fill_type="solid", fgColor="F1CEEE")
        cursos_fill = PatternFill(fill_type="solid", fgColor="C1F4E5")
        calc_fill = PatternFill(fill_type="solid", fgColor="CAEDFB")

        full_area_map = dict(area_map or {})

        n_new_courses = 0
        # 1) Cursos nuevos → hoja 'cursos'
        if new_courses_df is not None and not new_courses_df.empty:
            for r in new_courses_df.itertuples(index=False, name=None):
                full_area_map[str(r[0]).strip()] = r[3]  # Código Materia -> Area del curso

            info = _table_info(ws_cursos, "tabla_cursos")
            if not info:
                return False, "No encontré la Tabla de Excel 'tabla_cursos' en la hoja 'cursos'."
            _, min_col_c, min_row_c, max_col_c, last_row_c = info
            # last_row_c viene del ref de la Tabla de Excel, que puede haber
            # quedado inflado en cargas anteriores (mismo problema de
            # max_row-tras-delete) -- se recalcula contra el dato real.
            last_row_c = _last_data_row(ws_cursos, key_col=1, header_row=min_row_c, upper_bound=max(last_row_c, ws_cursos.max_row))
            template_row_c = last_row_c
            tpl_e_text, e_is_array = _get_formula_text(ws_cursos.cell(row=template_row_c, column=5))
            tpl_f_text, f_is_array = _get_formula_text(ws_cursos.cell(row=template_row_c, column=6))
            ef_ok = bool(tpl_e_text) and bool(tpl_f_text)

            append_start_c = last_row_c + 1
            for i, r in enumerate(new_courses_df.itertuples(index=False, name=None)):
                rn = append_start_c + i
                codigo, creditos, nombre, area = (r + ("", "", "", ""))[:4]
                for col, val in [(1, codigo), (2, creditos), (3, nombre), (4, area)]:
                    cell = ws_cursos.cell(row=rn, column=col, value=val)
                    cell.font = base_font
                if ef_ok:
                    # E y F se CALCULAN en Python (evaluando la fórmula real
                    # contra las tablas de referencia que ya viven en el
                    # workbook), en vez de copiar solo el texto de la
                    # fórmula -- así quedan resueltas de una sin depender de
                    # que alguien abra el archivo en Excel.
                    val_e = _eval_simple_lookup_formula(wb, tpl_e_text, rn, "cursos")
                    val_f = _eval_simple_lookup_formula(wb, tpl_f_text, rn, "cursos")
                    if val_e is None:
                        _write_translated_formula(ws_cursos, 5, rn, tpl_e_text, e_is_array, f"E{template_row_c}", f"E{rn}")
                    else:
                        ws_cursos.cell(row=rn, column=5, value=val_e)
                    if val_f is None:
                        _write_translated_formula(ws_cursos, 6, rn, tpl_f_text, f_is_array, f"F{template_row_c}", f"F{rn}")
                    else:
                        ws_cursos.cell(row=rn, column=6, value=val_f)
                    for col in (5, 6):
                        cell = ws_cursos.cell(row=rn, column=col)
                        cell.font = base_font
                        cell.fill = cursos_fill
                        cell.alignment = Alignment(horizontal="left")
                        cell.fill = cursos_fill
                n_new_courses += 1

            new_last_row_c = append_start_c + n_new_courses - 1
            if info[0]:
                ws_cursos.tables[info[0]].ref = (
                    f"{get_column_letter(min_col_c)}{min_row_c}:{get_column_letter(max_col_c)}{new_last_row_c}"
                )

        # ── Tablas de referencia reales, leídas del propio archivo (no inventadas) ──
        # AD:AE de 'cartelera' → Periodo crudo -> Semestre limpio
        semestre_map: Dict[str, object] = {}
        for r in range(2, ws_cart.max_row + 1):
            k = ws_cart.cell(row=r, column=30).value
            if k is None:
                continue
            semestre_map[str(k).strip()] = ws_cart.cell(row=r, column=31).value

        # I:J de 'cursos' (tabla de referencia real Area del curso -> Field;
        # NO se usa la columna E porque esa también es una fórmula que puede
        # no estar resuelta si la fila se escribió por automatización)
        field_map: Dict[str, object] = {}
        for r in range(2, ws_cursos.max_row + 1):
            k = ws_cursos.cell(row=r, column=9).value
            if k is None:
                continue
            field_map.setdefault(str(k).strip(), ws_cursos.cell(row=r, column=10).value)

        # L:M de 'cursos' → primeros 4 caracteres del código de materia -> Cod program
        codprog_map: Dict[str, object] = {}
        for r in range(2, ws_cursos.max_row + 1):
            k = ws_cursos.cell(row=r, column=12).value
            if k is None:
                continue
            codprog_map.setdefault(str(k).strip(), ws_cursos.cell(row=r, column=13).value)

        # A:C de 'programas' → Cod program -> Program
        program_map: Dict[str, object] = {}
        if ws_programas is not None:
            for r in range(2, ws_programas.max_row + 1):
                k = ws_programas.cell(row=r, column=1).value
                if k is None:
                    continue
                program_map.setdefault(str(k).strip(), ws_programas.cell(row=r, column=3).value)

        # 2) Filas de cartelera
        info_cart = _table_info(ws_cart, "tabla_cartelera")
        if not info_cart:
            return False, "No encontré la Tabla de Excel 'tabla_cartelera' en la hoja 'cartelera'."
        match_cart, min_col_ct, min_row_ct, max_col_ct, last_row_ct = info_cart

        # Periodos presentes en la carga: borra filas existentes con esos periodos primero
        periodos = set(str(p).strip() for p in cartelera_df["Periodo"].dropna().unique())
        rows_to_delete = [
            r for r in range(2, last_row_ct + 1)
            if str(ws_cart.cell(row=r, column=1).value or "").strip() in periodos
        ]
        for r in sorted(rows_to_delete, reverse=True):
            ws_cart.delete_rows(r)

        # ── Auto-reparación silenciosa ──────────────────────────────────
        # Filas de cargas anteriores (antes de este arreglo) pueden haber
        # quedado con fórmula sin resolver en B, I, J, K o Q-W. Cada vez que
        # se guarda algo nuevo, se revisa y corrige TODO lo existente con las
        # mismas tablas de referencia que ya construimos arriba — así no hace
        # falta un botón de reparación aparte; se autocorrige solo.
        def _needs_fix(v):
            if v is None or v == "":
                return True
            if isinstance(v, str) and v.startswith("="):
                return True
            if isinstance(v, ArrayFormula):
                return True
            return False

        heal_info = _table_info(ws_cart, "tabla_cartelera")
        heal_last_row = _last_data_row(ws_cart, key_col=1, header_row=heal_info[2] if heal_info else 1,
                                        upper_bound=ws_cart.max_row) if heal_info else ws_cart.max_row
        for r in range(2, heal_last_row + 1):
            h_periodo = ws_cart.cell(row=r, column=1).value
            h_materia = ws_cart.cell(row=r, column=4).value
            h_materia_key = str(h_materia).strip() if h_materia is not None else ""
            h_area = ws_cart.cell(row=r, column=8).value
            h_ps = str(ws_cart.cell(row=r, column=16).value or "").strip().upper()
            h_tipo = str(ws_cart.cell(row=r, column=15).value or "").strip().upper()
            h_creditos = pd.to_numeric(pd.Series([ws_cart.cell(row=r, column=6).value]), errors="coerce").iloc[0]
            h_creditos = 0 if pd.isna(h_creditos) else h_creditos

            if _needs_fix(ws_cart.cell(row=r, column=2).value):
                c = ws_cart.cell(row=r, column=2, value=semestre_map.get(str(h_periodo).strip(), h_periodo))
                c.font = base_font; c.fill = calc_fill
            if _needs_fix(ws_cart.cell(row=r, column=9).value):
                c = ws_cart.cell(row=r, column=9, value=field_map.get(str(h_area).strip(), ""))
                c.font = base_font; c.fill = calc_fill
            if _needs_fix(ws_cart.cell(row=r, column=10).value):
                h_cp = codprog_map.get(h_materia_key[:4], "")
                c = ws_cart.cell(row=r, column=10, value=h_cp)
                c.font = base_font; c.fill = calc_fill
            if _needs_fix(ws_cart.cell(row=r, column=11).value):
                h_cp2 = ws_cart.cell(row=r, column=10).value
                c = ws_cart.cell(row=r, column=11, value=program_map.get(str(h_cp2).strip(), ""))
                c.font = base_font; c.fill = calc_fill
            heal_breakdown = {
                17: h_creditos if h_ps == "P" else 0,
                18: h_creditos if h_ps == "S" else 0,
                19: h_creditos if h_tipo == "OTHER" else 0,
                20: h_creditos if h_tipo == "SA" else 0,
                21: h_creditos if h_tipo == "PA" else 0,
                22: h_creditos if h_tipo == "IP" else 0,
                23: h_creditos if h_tipo == "SP" else 0,
            }
            for col, val in heal_breakdown.items():
                if _needs_fix(ws_cart.cell(row=r, column=col).value):
                    c = ws_cart.cell(row=r, column=col, value=val)
                    c.font = base_font; c.fill = calc_fill

        info_cart2 = _table_info(ws_cart, "tabla_cartelera")
        _, _, min_row_ct2, _, last_row_ct2 = info_cart2
        # El ref de la Tabla de Excel NO se actualiza solo al borrar filas
        # (delete_rows no lo toca) -- por eso quedaba un hueco de filas
        # vacías antes de las nuevas. Se recalcula contra el dato real.
        last_row_ct2 = _last_data_row(ws_cart, key_col=1, header_row=min_row_ct2, upper_bound=max(last_row_ct2, ws_cart.max_row))
        append_start_ct = last_row_ct2 + 1

        lookup = profesor_lookup or {}
        for i, r in enumerate(cartelera_df.itertuples(index=False, name=None)):
            rn = append_start_ct + i
            periodo, campus, materia, secc, creditos, nombre, profesor = (r + ("",) * 7)[:7]
            materia_key = str(materia).strip()
            direct = {1: periodo, 3: campus, 4: materia, 5: secc, 6: creditos, 7: nombre, 12: profesor}
            for col, val in direct.items():
                cell = ws_cart.cell(row=rn, column=col, value=val)
                cell.font = base_font

            # B — Semestre (limpio vía tabla AD:AE; si el periodo no está ahí
            # registrado, se deja al menos el Periodo crudo — nunca en blanco)
            b_cell = ws_cart.cell(row=rn, column=2, value=semestre_map.get(str(periodo).strip(), periodo))
            b_cell.font = base_font
            b_cell.fill = calc_fill

            # H — Area del curso: valor literal buscado por Código Materia contra 'cursos'
            h_val = full_area_map.get(materia_key, "")
            h_cell = ws_cart.cell(row=rn, column=8, value=h_val)
            h_cell.font = base_font
            h_cell.fill = area_fill

            # I — Field (vía Area del curso -> cursos!D:E)
            i_cell = ws_cart.cell(row=rn, column=9, value=field_map.get(str(h_val).strip(), ""))
            i_cell.font = base_font
            i_cell.fill = calc_fill

            # J — Cod program (vía primeros 4 caracteres de Materia -> cursos!L:M)
            cod_prog = codprog_map.get(materia_key[:4], "")
            j_cell = ws_cart.cell(row=rn, column=10, value=cod_prog)
            j_cell.font = base_font
            j_cell.fill = calc_fill

            # K — Program (vía Cod program -> programas!A:C)
            k_cell = ws_cart.cell(row=rn, column=11, value=program_map.get(str(cod_prog).strip(), ""))
            k_cell.font = base_font
            k_cell.fill = calc_fill

            # M,N,O,P — ID, AREA_PROFESOR, TIPO, P/S (valor literal, buscado por nombre)
            prof_key = str(profesor).strip().upper()
            match_prof = lookup.get(prof_key)
            tipo_val, ps_val = "", ""
            if match_prof:
                for col, val in zip((13, 14, 15, 16), match_prof):
                    cell = ws_cart.cell(row=rn, column=col, value=val)
                    cell.font = base_font
                    cell.fill = area_fill
                tipo_val, ps_val = str(match_prof[2]).strip().upper(), str(match_prof[3]).strip().upper()

            # Q,R,S,T,U,V,W — desglose de créditos por P/S y TIPO (misma lógica que la fórmula real:
            # Créditos si coincide con la etiqueta de la columna, si no 0)
            creditos_num = pd.to_numeric(pd.Series([creditos]), errors="coerce").iloc[0]
            creditos_num = 0 if pd.isna(creditos_num) else creditos_num
            breakdown = {
                17: creditos_num if ps_val == "P" else 0,        # Q
                18: creditos_num if ps_val == "S" else 0,        # R
                19: creditos_num if tipo_val == "OTHER" else 0,  # S
                20: creditos_num if tipo_val == "SA" else 0,     # T
                21: creditos_num if tipo_val == "PA" else 0,     # U
                22: creditos_num if tipo_val == "IP" else 0,     # V
                23: creditos_num if tipo_val == "SP" else 0,     # W
            }
            for col, val in breakdown.items():
                cell = ws_cart.cell(row=rn, column=col, value=val)
                cell.font = base_font
                cell.fill = calc_fill

        new_last_row_ct = append_start_ct + len(cartelera_df) - 1
        if match_cart:
            ws_cart.tables[match_cart].ref = (
                f"{get_column_letter(min_col_ct)}{min_row_ct}:{get_column_letter(max_col_ct)}{new_last_row_ct}"
            )

        wb.calculation.fullCalcOnLoad = True
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        ok, err = _drive_upload_file_bytes(CARTELERA_FILE_ID, buf.getvalue())
        if not ok:
            return False, f"Error al subir el archivo actualizado a Drive: {err}"

        qual_load_cartelera.clear()
        _download_drive_file_bytes.clear()
        _load_cursos_area_map.clear()

        msg = f"✓ BD_cartelera.xlsx actualizada — {len(cartelera_df)} filas en 'cartelera'"
        if n_new_courses:
            msg += f" y {n_new_courses} cursos nuevos en 'cursos'"
        msg += "."
        return True, msg
    except Exception as e:
        return False, f"Error al escribir en BD_cartelera.xlsx: {e}"


def page_update_data():
    _render_header("Update Data", "Sube la Template para actualizar la BD maestra")

    st.markdown(
        "<div style='text-align:center;'>Esta sección reemplaza el antiguo modal "
        "<i>Update data</i> de la web de KPIs. Los cambios que hagas aquí se escriben "
        "directamente en el Google Sheet que alimenta todos los dashboards.</div>",
        unsafe_allow_html=True,
    )

    tab_planta, tab_cartelera, tab_quest = st.tabs(
        ["BD_planta", "BD_cartelera", "BD_Faculty_Questionnaire"]
    )

    # ── BD_planta ───────────────────────────────────────────────────────
    with tab_planta:
        col_h, col_info = st.columns([6, 0.6])
        with col_h:
            st.markdown("#### Actualizar BD_planta")
        with col_info:
            with st.popover("", icon=":material/help:"):
                st.caption("Base de referencia con toda la información de profesores.")
                st.download_button(
                    "Descargar Base Info. Profesores", data=_build_info_profesores_download(),
                    file_name="Info_Profesores.xlsx", key="dl_info_profesores",
                    icon=":material/download:",
                )

        _planta_regular = [p for p in df["Periodo"].dropna().unique().tolist() if "Intersemestral" not in str(p)]
        last_period = sorted(_planta_regular, key=_period_sort_key)[-1] if _planta_regular else "—"
        st.caption(f"Último periodo registrado en la Base: **{last_period}**")

        st.download_button(
            "Descargar Template_planta.xlsx", data=_download_drive_file_bytes(TEMPLATE_PLANTA_FILE_ID),
            file_name="Template_planta.xlsx", key="dl_template_planta", icon=":material/download:",
        )

        up = st.file_uploader("Template_planta.xlsx diligenciada", type=["xlsx"], key="planta_upload")

        if up is not None:
            try:
                tpl_df = _read_planta_template(up)
            except Exception as e:
                st.error(f"No pude leer el archivo: {e}")
                tpl_df = None

            if tpl_df is not None and not tpl_df.empty:
                st.success(f"{len(tpl_df)} filas detectadas en la template.")

                preview_cols = list(tpl_df.columns)
                notes_col = next((c for c in preview_cols if str(c).strip().lower() == "notes"), None)
                if notes_col and notes_col != "Notes":
                    tpl_df = tpl_df.rename(columns={notes_col: "Notes"})

                st.markdown("**Vista previa**")
                st.dataframe(_style_planta_preview(tpl_df), use_container_width=True, hide_index=True)

                if st.button("Guardar en BD_planta", type="primary", icon=":material/save:"):
                    with st.spinner("Escribiendo en Drive…"):
                        ok, msg = push_planta_updates(tpl_df)
                    if ok:
                        st.success(msg)
                        st.balloons()
                    else:
                        st.error(msg)
            elif tpl_df is not None:
                st.warning("No se detectaron filas de datos a partir de la fila 6.")

    # ── BD_cartelera ─────────────────────────────────────────────────────
    with tab_cartelera:
        col_h2, col_info2 = st.columns([6, 0.6])
        with col_h2:
            st.markdown("#### Actualizar BD_cartelera")
        with col_info2:
            with st.popover("", icon=":material/help:"):
                st.caption("Hoja de referencia con los cursos y áreas ya cargados.")
                st.download_button(
                    "Descargar hoja de cursos", data=_build_cursos_download(),
                    file_name="cursos.xlsx", key="dl_cursos_sheet",
                    icon=":material/download:",
                )

        _cart_periods_raw = [
            p for p in qual_load_cartelera()["Semestre"].dropna().unique().tolist()
            if re.fullmatch(r"(?:19|20)\d{2}(10|20)", str(p).strip())
        ]

        def _with_dash(p: str) -> str:
            s = str(p).strip()
            if "Intersemestral" in s or "-" in s:
                return s
            return f"{s[:4]}-{s[4:]}" if len(s) == 6 else s

        last_cart_period = _with_dash(sorted(_cart_periods_raw, key=_period_sort_key)[-1]) if _cart_periods_raw else "—"
        st.caption(f"Último periodo registrado en la Base: **{last_cart_period}**")

        st.download_button(
            "Descargar Template_cartelera.xlsx", data=_download_drive_file_bytes(TEMPLATE_CARTELERA_FILE_ID),
            file_name="Template_cartelera.xlsx", key="dl_template_cartelera", icon=":material/download:",
        )

        up_cart = st.file_uploader("Template_cartelera.xlsx", type=["xlsx"], key="cartelera_upload")

        if up_cart is not None:
            try:
                cart_df = _read_cartelera_template(up_cart)
            except Exception as e:
                st.error(f"No pude leer el archivo: {e}")
                cart_df = None

            if cart_df is not None and not cart_df.empty:
                # --- Área del curso (lookup contra 'cursos') ---
                area_map = _load_cursos_area_map()
                cart_df["Materia"] = cart_df["Materia"].astype(str).str.strip()
                cart_df["Area del curso"] = cart_df["Materia"].map(area_map)
                missing_area_mask = cart_df["Area del curso"].isna()

                # --- Profesor (lookup contra 'Info. Profesores') ---
                prof_lookup = _load_profesores_lookup()
                cart_df["Profesor"] = cart_df["Profesor"].astype(str).str.strip()
                missing_prof_mask = ~cart_df["Profesor"].str.upper().isin(prof_lookup.keys())

                st.success(
                    f"{len(cart_df)} filas detectadas · "
                    f"{(~missing_area_mask).sum()} con área encontrada, {missing_area_mask.sum()} sin área · "
                    f"{(~missing_prof_mask).sum()} con profesor encontrado, {missing_prof_mask.sum()} sin profesor."
                )
                st.markdown("**Vista previa**")
                st.dataframe(cart_df, use_container_width=True, hide_index=True)

                # ============== CURSOS NUEVOS (área) ==============
                missing_courses = (
                    cart_df.loc[missing_area_mask, ["Materia", "Créditos", "Nombre largo curso"]]
                    .drop_duplicates(subset=["Materia"])
                    .reset_index(drop=True)
                )

                new_courses_df = None
                if missing_courses.empty:
                    new_courses_df = pd.DataFrame(columns=["Código Materia", "Créditos", "Nombre largo curso", "Area del curso"])
                else:
                    st.warning(
                        f"⚠️ {len(missing_courses)} curso(s) no están en la hoja 'cursos' — "
                        "hay que asignarles un área antes de poder guardar."
                    )
                    fill_mode = st.radio(
                        "¿Cómo quieres completar las áreas?",
                        ["Seleccionar aquí mismo", "Subir Template_cursos_nuevos.xlsx diligenciada"],
                        key="cartelera_fill_mode", horizontal=True,
                    )

                    if fill_mode == "Seleccionar aquí mismo":
                        picked_areas = {}
                        grid_cols = st.columns(2)
                        for i, (_, row) in enumerate(missing_courses.iterrows()):
                            area_key = f"area_pick_{row['Materia']}"
                            is_empty = st.session_state.get(area_key, "— Selecciona —") == "— Selecciona —"
                            box_key = f"course_box_bad_{i}" if is_empty else f"course_box_ok_{i}"
                            with grid_cols[i % 2]:
                                with st.container(key=box_key):
                                    with st.expander(f"{row['Materia']}", expanded=False, icon=":material/menu_book:"):
                                        c1, c2, c3 = st.columns([1, 3, 0.6])
                                        c1.markdown(f"**Créditos:** {row['Créditos']}")
                                        c2.markdown(row["Nombre largo curso"])
                                        with c3.popover("", icon=":material/person:"):
                                            profs = sorted(cart_df.loc[cart_df["Materia"] == row["Materia"], "Profesor"].dropna().unique().tolist())
                                            st.caption("Profesor(es) y área:")
                                            for p in profs:
                                                info = prof_lookup.get(str(p).strip().upper())
                                                area_txt = info[1] if info else "—"
                                                st.markdown(f"{p} · *{area_txt}*")

                                        color = "#DC2626" if is_empty else "#374151"
                                        st.markdown(
                                            f"<div style='font-size:12px;color:{color};font-weight:600;'>Area *</div>",
                                            unsafe_allow_html=True,
                                        )
                                        picked_areas[row["Materia"]] = st.selectbox(
                                            "Area", options=["— Selecciona —"] + AREA_OPTIONS,
                                            key=area_key, label_visibility="collapsed",
                                        )

                        if all(v != "— Selecciona —" for v in picked_areas.values()):
                            new_courses_df = missing_courses.assign(
                                **{"Area del curso": missing_courses["Materia"].map(picked_areas)}
                            ).rename(columns={"Materia": "Código Materia"})
                    else:
                        st.download_button(
                            "Descargar Template_cursos_nuevos.xlsx (con los datos ya puestos)",
                            data=_build_prefilled_cursos_template(missing_courses),
                            file_name="Template_cursos_nuevos.xlsx",
                            key="cursos_template_dl", icon=":material/download:",
                        )
                        up_new = st.file_uploader(
                            "Template_cursos_nuevos.xlsx diligenciada", type=["xlsx"], key="cursos_nuevos_upload"
                        )
                        if up_new is not None:
                            try:
                                nc = _read_cursos_nuevos_template(up_new)
                                nc.columns = [c.strip() for c in nc.columns]
                                new_courses_df = nc
                                st.dataframe(new_courses_df, use_container_width=True, hide_index=True)
                            except Exception as e:
                                st.error(f"No pude leer la template de cursos nuevos: {e}")

                st.markdown("---")

                # ============== PROFESORES NUEVOS ==============
                missing_profs = sorted(cart_df.loc[missing_prof_mask, "Profesor"].dropna().unique().tolist())
                new_profs_df = None
                if not missing_profs:
                    new_profs_df = pd.DataFrame()
                else:
                    st.warning(
                        f"⚠️ {len(missing_profs)} profesor(es) no están en 'Info. Profesores' — "
                        "hay que completarlos antes de poder guardar."
                    )
                    prof_fill_mode = st.radio(
                        "¿Cómo quieres completar los profesores?",
                        ["Seleccionar aquí mismo", "Subir Template_profesores_nuevos.xlsx diligenciada"],
                        key="profesores_fill_mode", horizontal=True,
                    )

                    if prof_fill_mode == "Seleccionar aquí mismo":
                        def _req_label(text: str, key: str, is_empty) -> None:
                            color = "#DC2626" if is_empty(st.session_state.get(key, "")) else "#374151"
                            st.markdown(
                                f"<div style='font-size:14px;color:{color};font-weight:600;margin-bottom:2px;'>{text} *</div>",
                                unsafe_allow_html=True,
                            )

                        _empty_txt = lambda v: str(v).strip() == ""
                        _empty_sel = lambda v: v in (None, "", "— Selecciona —")

                        picked_profs = {}
                        all_required_filled = True
                        for p_idx, name in enumerate(missing_profs):
                            row_currently_ok = (
                                not _empty_txt(st.session_state.get(f"prof_id_{name}", ""))
                                and not _empty_sel(st.session_state.get(f"prof_area_{name}", ""))
                                and not _empty_sel(st.session_state.get(f"prof_genero_{name}", ""))
                                and not _empty_sel(st.session_state.get(f"prof_tipo_{name}", ""))
                                and not _empty_sel(st.session_state.get(f"prof_ps_{name}", ""))
                                and not _empty_sel(st.session_state.get(f"prof_planta_{name}", ""))
                            )
                            prof_box_key = f"prof_box_ok_{p_idx}" if row_currently_ok else f"prof_box_bad_{p_idx}"
                            with st.container(key=prof_box_key):
                                with st.expander(f"{name}", expanded=False, icon=":material/person:"):
                                    r1c1, r1c2, r1c3 = st.columns(3)
                                    with r1c1:
                                        _req_label("ID / Cédula", f"prof_id_{name}", _empty_txt)
                                        p_id = st.text_input("ID / Cédula", key=f"prof_id_{name}", label_visibility="collapsed")
                                    with r1c2:
                                        _req_label("AREA_PROFESOR", f"prof_area_{name}", _empty_sel)
                                        p_area = st.selectbox("AREA_PROFESOR", ["— Selecciona —"] + AREA_OPTIONS, key=f"prof_area_{name}", label_visibility="collapsed")
                                    with r1c3:
                                        _req_label("GÉNERO", f"prof_genero_{name}", _empty_sel)
                                        p_genero = st.selectbox("GÉNERO", ["— Selecciona —"] + GENERO_OPTIONS, key=f"prof_genero_{name}", label_visibility="collapsed")

                                    r2c1, r2c2, r2c3 = st.columns(3)
                                    with r2c1:
                                        _req_label("TIPO", f"prof_tipo_{name}", _empty_sel)
                                        p_tipo = st.selectbox("TIPO", ["— Selecciona —"] + TIPO_OPTIONS, key=f"prof_tipo_{name}", label_visibility="collapsed")
                                    with r2c2:
                                        _req_label("P/S", f"prof_ps_{name}", _empty_sel)
                                        p_ps = st.selectbox("P/S", ["— Selecciona —"] + PS_OPTIONS, key=f"prof_ps_{name}", label_visibility="collapsed")
                                    with r2c3:
                                        _req_label("PLANTA_CATEDRA", f"prof_planta_{name}", _empty_sel)
                                        p_planta = st.selectbox("PLANTA_CATEDRA", ["— Selecciona —"] + PLANTA_CATEDRA_OPTIONS, key=f"prof_planta_{name}", label_visibility="collapsed")

                                    r3c1, r3c2, r3c3 = st.columns(3)
                                    p_fecha_ingreso = r3c1.text_input("Date of First Appointment (DD/MM/YYYY)", key=f"prof_fecha_ing_{name}")
                                    p_degree = r3c2.text_input("Highest Earned Degree", key=f"prof_degree_{name}")
                                    p_year = r3c3.text_input("Highest Degree, Year Earned", key=f"prof_year_{name}")

                                    r4c1, r4c2, r4c3 = st.columns(3)
                                    p_hd = r4c1.selectbox("Highest Degree", ["— (deja TBD) —"] + HIGHEST_DEGREE_OPTIONS, key=f"prof_hd_{name}")
                                    p_univ = r4c2.text_input("University", key=f"prof_univ_{name}")
                                    p_region = r4c3.selectbox("Region Where it was obtained", ["— (deja TBD) —"] + REGION_OPTIONS, key=f"prof_region_{name}")

                                    r5c1, r5c2, r5c3 = st.columns(3)
                                    p_intl = r5c1.selectbox("International Degree?", ["— (deja TBD) —"] + INTL_DEGREE_OPTIONS, key=f"prof_intl_{name}")
                                    p_resp = r5c2.text_input("Normal Professional Responsibilities", key=f"prof_resp_{name}")
                                    p_basis = r5c3.text_input("Basis for qualification", key=f"prof_basis_{name}")

                                    r6c1, r6c2, r6c3 = st.columns(3)
                                    p_nat = r6c1.text_input("Nationality", key=f"prof_nat_{name}")
                                    p_dob = r6c2.text_input("Date of birth (DD/MM/YYYY)", key=f"prof_dob_{name}")
                                    p_exp = r6c3.text_input("Years Industry experience", key=f"prof_exp_{name}")

                                    with st.popover("Cursos que dicta (según esta carga)", use_container_width=True, icon=":material/menu_book:"):
                                        courses_taught = cart_df.loc[
                                            cart_df["Profesor"] == name, ["Materia", "Créditos", "Nombre largo curso"]
                                        ].drop_duplicates().reset_index(drop=True)
                                        st.dataframe(courses_taught, use_container_width=True, hide_index=True)

                            row_required_ok = (
                                p_id.strip() != "" and p_area != "— Selecciona —" and p_genero != "— Selecciona —"
                                and p_tipo != "— Selecciona —" and p_ps != "— Selecciona —" and p_planta != "— Selecciona —"
                            )
                            all_required_filled = all_required_filled and row_required_ok
                            picked_profs[name] = {
                                "Profesor": name, "ID": p_id, "AREA_PROFESOR": p_area, "GÉNERO": p_genero,
                                "TIPO": p_tipo, "P/S": p_ps, "PLANTA_CATEDRA": p_planta,
                                "Date of First Appointment to the School": p_fecha_ingreso,
                                "Highest Earned Degree": p_degree, "Highest Degree, Year Earned": p_year,
                                "Highest Degree": "" if p_hd.startswith("—") else p_hd,
                                "University": p_univ,
                                "Region Where it was obtained": "" if p_region.startswith("—") else p_region,
                                "International Degree?": "" if p_intl.startswith("—") else p_intl,
                                "Normal Professional Responsibilities": p_resp, "Basis for qualification": p_basis,
                                "Nationality": p_nat, "Date of birth": p_dob, "Years Industry experience": p_exp,
                            }
                        if all_required_filled:
                            new_profs_df = pd.DataFrame(list(picked_profs.values()))
                    else:
                        st.download_button(
                            "Descargar Template_profesores_nuevos.xlsx (con los nombres ya puestos)",
                            data=_build_prefilled_profesores_template(missing_profs),
                            file_name="Template_profesores_nuevos.xlsx",
                            key="prof_template_dl", icon=":material/download:",
                        )
                        up_profs = st.file_uploader(
                            "Template_profesores_nuevos.xlsx diligenciada", type=["xlsx"], key="profesores_nuevos_upload"
                        )
                        if up_profs is not None:
                            try:
                                npf = _read_profesores_nuevos_template(up_profs)
                                npf.columns = [c.strip() for c in npf.columns]
                                new_profs_df = npf
                                st.dataframe(new_profs_df, use_container_width=True, hide_index=True)
                            except Exception as e:
                                st.error(f"No pude leer la template de profesores nuevos: {e}")

                ready = new_courses_df is not None and new_profs_df is not None
                if not ready:
                    st.info("Completa las áreas y/o los profesores pendientes antes de guardar.")

                if st.button("Guardar en BD_Cartelera", type="primary", disabled=not ready, icon=":material/save:"):
                    save_df = cart_df.drop(columns=["Area del curso"])
                    with st.spinner("Escribiendo en Drive…"):
                        combined_lookup = dict(prof_lookup)
                        if new_profs_df is not None and not new_profs_df.empty:
                            ok_p, msg_p = push_profesores_updates(new_profs_df)
                            if not ok_p:
                                st.error(msg_p)
                                st.stop()
                            st.success(msg_p)
                            for r in new_profs_df.itertuples(index=False, name=None):
                                name_key = str(r[0]).strip().upper()
                                combined_lookup[name_key] = (r[1], r[2], r[4], r[5])  # ID, AREA_PROFESOR, TIPO, P/S
                        ok, msg = push_cartelera_updates(save_df, new_courses_df, combined_lookup, area_map)
                    if ok:
                        st.success(msg)
                        # Faculty Distribution: un ID único por periodo, tomado de la cartelera recién guardada
                        periodo_to_ids: Dict[str, List] = {}
                        for periodo_val, prof_name in zip(save_df["Periodo"], save_df["Profesor"]):
                            m = combined_lookup.get(str(prof_name).strip().upper())
                            if not m:
                                continue
                            periodo_to_ids.setdefault(str(periodo_val).strip(), set()).add(m[0])
                        periodo_to_ids = {p: sorted(ids, key=str) for p, ids in periodo_to_ids.items()}
                        if periodo_to_ids:
                            with st.spinner("Actualizando Faculty Distribution…"):
                                ok_fd, msg_fd = push_faculty_distribution_updates(periodo_to_ids)
                            if ok_fd:
                                st.success(msg_fd)
                            else:
                                st.error(msg_fd)
                        st.balloons()

                        target_period = sorted(periodo_to_ids.keys(), key=_period_sort_key)[-1] if periodo_to_ids else None
                        if target_period:
                            with st.spinner("Preparando reporte…"):
                                report_bytes = _build_cartelera_save_report(target_period, new_courses_df, new_profs_df)
                            st.markdown(
                                "<style>div[data-testid='stDownloadButton'] button{"
                                "background-color:#16A34A !important;color:#FFFFFF !important;"
                                "border:none !important;font-weight:700 !important;}"
                                "div[data-testid='stDownloadButton'] button:hover{background-color:#15803D !important;}</style>",
                                unsafe_allow_html=True,
                            )
                            st.download_button(
                                "Descargar reporte en Excel", data=report_bytes,
                                file_name=f"Reporte_Cartelera_{target_period}.xlsx".replace(" ", "_"),
                                key="dl_cartelera_report", icon=":material/download:",
                                use_container_width=True,
                            )
                    else:
                        st.error(msg)
            elif cart_df is not None:
                st.warning("No se detectaron filas de datos a partir de la fila 6.")

    # ── BD_Faculty_Questionnaire ─────────────────────────────────────────
    with tab_quest:
        pass


# Navegación multipágina — menú nativo oculto; desplegable sutil (flecha) con los enlaces
pages = [
    st.Page(page_composition, title="Composition", icon="🎓", url_path="composition", default=True),
    st.Page(page_staffing, title="Staffing Levels", icon="📊", url_path="staffing"),
    st.Page(page_area, title="By Area", icon="🏛️", url_path="area"),
    st.Page(page_demographics, title="Demographics", icon="🧑‍🤝‍🧑", url_path="demographics"),
    st.Page(page_activities, title="Activities", icon="🧭", url_path="activities"),
    st.Page(page_qualifications, title="Qualifications", icon="📚", url_path="qualifications"),
    st.Page(page_update_data, title="Update Data", icon="🔄", url_path="update-data"),
]
pg = st.navigation(pages, position="hidden")
IS_UPDATE_PAGE = pg is pages[-1]  # comparación por identidad, más confiable que el título

if not IS_UPDATE_PAGE:
    # Recuerda en qué página estaba el usuario antes de ir a Update Data,
    # para que "Go to Faculty Dashboard" lo regrese exactamente ahí.
    st.session_state["_return_page_idx"] = pages.index(pg)


def _period_sort_key(p):
    s = str(p).strip()
    try:
        return (int(s[:4]), 30 if "Intersemestral" in s else int(s[-2:].replace("-", "")))
    except (ValueError, IndexError):
        return (-1, -1)  # valores no reconocibles (vacíos, ruido de datos) quedan al final al ordenar


if not IS_UPDATE_PAGE:
    with st.sidebar:
        col_logo, col_title = st.columns([1, 3])
        with col_logo:
            st.image("imagenes/logo.png", width=65)
        with col_title:
            st.markdown(
                '<div style="padding-top:10px;color:#004d47;font-size:24px;'
                'font-weight:800;line-height:1.1;">UASM Faculty KPIs</div>',
                unsafe_allow_html=True,
            )
            st.caption("Analytics Dashboard")
        st.markdown("---")

    # "Other sections": simple, nativo de Streamlit, siempre visible.
    with st.container(key="nav_toggle"):
        nav_cols = st.columns(6)
        for col, page_obj in zip(nav_cols, pages[:-1]):
            with col:
                st.page_link(page_obj)

pg.run()

if IS_UPDATE_PAGE:
    # Update Data: logo + título pegados justo encima del botón, todo el
    # grupo centrado a media altura del sidebar — nada arriba por separado.
    # Logo y texto quedan planos (sin link); solo el botón es clicable.
    return_idx = st.session_state.get("_return_page_idx", 0)
    return_page = pages[return_idx]
    with st.sidebar:
        st.markdown('<div style="height:26vh;"></div>', unsafe_allow_html=True)
        with st.container(key="update_sidebar_group"):
            st.image("imagenes/logo.png", width=65)
            st.markdown(
                '<div style="color:#004d47;font-size:22px;font-weight:800;'
                'line-height:1.15;margin-top:8px;">UASM Faculty KPIs</div>'
                '<div style="color:#6b7280;font-size:13px;margin-bottom:16px;">Analytics Dashboard</div>',
                unsafe_allow_html=True,
            )
            with st.container(key="go_to_dashboard_btn"):
                st.page_link(return_page, label="Go to Faculty Dashboard", icon=":material/bar_chart:", use_container_width=True)
else:
    # Resto de páginas: Download + botón de Update, uno al lado del otro, al fondo del sidebar
    with st.sidebar:
        st.markdown("---")
        col_dl, col_upd = st.columns(2)
        with col_dl:
            with st.expander("Download", expanded=False, icon=":material/download:"):
                _dl_periods = sorted(df["Periodo"].dropna().unique().tolist(), key=_period_sort_key, reverse=True)
                dl_period = st.selectbox("Period", _dl_periods, index=0, key="dl_db_period")
                dl_scope = st.radio("Scope", ["Full-time", "Part-time", "Catalog"], key="dl_db_scope", horizontal=True)

                def _build_db_download(period: str, scope: str) -> bytes:
                    buf = io.BytesIO()
                    with pd.ExcelWriter(buf) as writer:
                        if scope == "Full-time":
                            df[df["Periodo"].astype(str) == str(period)].to_excel(writer, index=False, sheet_name="Full-time")
                        elif scope == "Part-time":
                            df_cat = demo_load_parttime()
                            period_nodash = str(period).replace("-", "")
                            df_cat[df_cat["Periodo"].astype(str) == period_nodash].to_excel(
                                writer, index=False, sheet_name="Part-time"
                            )
                        else:  # Catalog — cartelera completa del periodo (incluye Field/Program y el
                            # desglose de créditos P/S/OTHER/SA/PA/IP/SP ya calculados).
                            df_cart_all = qual_load_cartelera()
                            period_nodash = str(period).replace("-", "")
                            cart_mask = df_cart_all["Periodo"].astype(str).isin([str(period), period_nodash])
                            df_cart_all[cart_mask].to_excel(writer, index=False, sheet_name="Catalog")
                    buf.seek(0)
                    return buf.getvalue()

                st.download_button(
                    "Download", data=_build_db_download(dl_period, dl_scope),
                    file_name=f"BD_{dl_scope}_{dl_period}.xlsx".replace(" ", "_"),
                    key="dl_db_btn", use_container_width=True,
                )
        with col_upd:
            with st.container(key="go_to_update_btn"):
                st.page_link(pages[-1], label="Update", icon=":material/sync:", use_container_width=True)
